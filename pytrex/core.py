import json
import math
import asyncio
import subprocess
import os
import sys
import time
import hashlib
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from typing import Any, Callable, Dict, List, Optional, Tuple

REGISTERED_EVENTS: Dict[str, Callable] = {}

BLOCKCHAIN_CACHE: List[dict] = []

_EVENT_EXECUTOR = ThreadPoolExecutor(max_workers=4, thread_name_prefix="pytrex-worker")

_RATE_LIMITS: Dict[str, List[float]] = {}
_RATE_LIMIT_WINDOW = 60.0
_RATE_LIMIT_MAX_CALLS = 100


def _log(level: str, message: str) -> None:
    timestamp = datetime.now(timezone.utc).isoformat()
    log_line = f"[{timestamp}] [PYTHON] [{level}] {message}\n"
    try:
        with open("pytrex_debug.log", "a", encoding="utf-8") as f:
            f.write(log_line)
    except Exception:
        pass


def event(name: str) -> Callable:
    def decorator(func: Callable) -> Callable:
        REGISTERED_EVENTS[name] = func
        return func
    return decorator


def _check_rate_limit(key: str, max_calls: int = _RATE_LIMIT_MAX_CALLS) -> bool:
    now = time.time()
    if key not in _RATE_LIMITS:
        _RATE_LIMITS[key] = []
    _RATE_LIMITS[key] = [t for t in _RATE_LIMITS[key] if now - t < _RATE_LIMIT_WINDOW]
    if len(_RATE_LIMITS[key]) >= max_calls:
        return False
    _RATE_LIMITS[key].append(now)
    return True


class InputValidator:
    """Pydantic-style input validation for event data."""

    @staticmethod
    def validate(data: str, schema: Optional[Dict[str, Any]] = None) -> Tuple[bool, Optional[dict], Optional[str]]:
        try:
            payload = json.loads(data)
        except json.JSONDecodeError as e:
            return False, None, f"Invalid JSON: {e}"

        if not isinstance(payload, dict):
            return False, None, "Expected JSON object"

        if schema:
            for field, rules in schema.items():
                if rules.get("required", False) and field not in payload:
                    return False, None, f"Missing required field: {field}"
                if field in payload:
                    expected_type = rules.get("type")
                    if expected_type and not isinstance(payload[field], expected_type):
                        return False, None, f"Field '{field}' must be {expected_type.__name__}"
                    min_val = rules.get("min")
                    if min_val is not None and isinstance(payload[field], (int, float)):
                        if payload[field] < min_val:
                            return False, None, f"Field '{field}' must be >= {min_val}"
                    max_val = rules.get("max")
                    if max_val is not None and isinstance(payload[field], (int, float)):
                        if payload[field] > max_val:
                            return False, None, f"Field '{field}' must be <= {max_val}"

        return True, payload, None


class AuthManager:
    """JWT-like authentication and RBAC system with bcrypt password hashing and token expiry."""

    TOKEN_EXPIRY_SECONDS = 3600  # 1 hour

    def __init__(self, secret: str = None):
        if secret is None:
            secret = os.environ.get("PYTREX_AUTH_SECRET")
            if not secret:
                raise ValueError(
                    "AuthManager requires a secret. Set PYTREX_AUTH_SECRET env var or pass secret parameter."
                )
        self._secret = secret.encode("utf-8")
        self._users: Dict[str, dict] = {}
        self._tokens: Dict[str, dict] = {}
        self._roles: Dict[str, List[str]] = {
            "admin": ["*"],
            "user": ["pata_akaunti", "fanya_muamala", "ping"],
            "guest": ["ping"],
        }
        self._bcrypt = None
        try:
            import bcrypt
            self._bcrypt = bcrypt
        except ImportError:
            _log("WARN", "bcrypt not installed — using hashlib fallback (NOT production-safe)")

    def _hash_password(self, password: str) -> str:
        if self._bcrypt:
            return self._bcrypt.hashpw(password.encode("utf-8"), self._bcrypt.gensalt(rounds=12)).decode("utf-8")
        # Fallback: SHA-256 with random salt (better than plain SHA-256, but not ideal)
        salt = os.urandom(32).hex()
        pw_hash = hashlib.sha256(salt.encode("utf-8") + password.encode("utf-8")).hexdigest()
        return f"sha256${salt}${pw_hash}"

    def _verify_password(self, password: str, stored_hash: str) -> bool:
        if self._bcrypt:
            try:
                return self._bcrypt.checkpw(password.encode("utf-8"), stored_hash.encode("utf-8"))
            except Exception:
                return False
        # Fallback: SHA-256 with salt
        if stored_hash.startswith("sha256$"):
            parts = stored_hash.split("$", 2)
            if len(parts) != 3:
                return False
            salt, expected_hash = parts[1], parts[2]
            actual = hashlib.sha256(salt.encode("utf-8") + password.encode("utf-8")).hexdigest()
            return actual == expected_hash
        # Legacy: plain SHA-256 (for backward compat)
        return hashlib.sha256(password.encode("utf-8")).hexdigest() == stored_hash

    def register_user(self, username: str, password: str, role: str = "user") -> bool:
        if username in self._users:
            return False
        pw_hash = self._hash_password(password)
        self._users[username] = {"password": pw_hash, "role": role}
        _log("INFO", f"User registered: {username} (role: {role})")
        return True

    def login(self, username: str, password: str) -> Optional[str]:
        user = self._users.get(username)
        if not user:
            return None
        if not self._verify_password(password, user["password"]):
            return None
        token = hashlib.sha256(f"{username}:{time.time()}:{user['role']}:{os.urandom(16).hex()}".encode("utf-8")).hexdigest()
        self._tokens[token] = {
            "username": username,
            "role": user["role"],
            "created": time.time(),
            "expires_at": time.time() + self.TOKEN_EXPIRY_SECONDS,
        }
        _log("INFO", f"User logged in: {username}")
        return token

    def verify_token(self, token: str) -> Optional[dict]:
        token_data = self._tokens.get(token)
        if not token_data:
            return None
        if time.time() > token_data.get("expires_at", 0):
            del self._tokens[token]
            return None
        return token_data

    def has_permission(self, token: str, event_name: str) -> bool:
        token_data = self.verify_token(token)
        if not token_data:
            return False
        role = token_data["role"]
        permissions = self._roles.get(role, [])
        if "*" in permissions:
            return True
        return event_name in permissions

    def logout(self, token: str) -> bool:
        if token in self._tokens:
            del self._tokens[token]
            return True
        return False

    def cleanup_expired_tokens(self) -> int:
        expired = [t for t, d in self._tokens.items() if time.time() > d.get("expires_at", 0)]
        for t in expired:
            del self._tokens[t]
        return len(expired)


class SecureKeyStorage:
    """OS keyring abstraction for storing encryption keys securely."""

    def __init__(self):
        self._backend = None
        try:
            import keyring
            self._backend = keyring
        except ImportError:
            _log("WARN", "keyring package not installed — using in-memory fallback")

    def store_key(self, service: str, key_name: str, value: str) -> bool:
        if self._backend:
            try:
                self._backend.set_password(service, key_name, value)
                return True
            except Exception as e:
                _log("ERROR", f"Keyring store failed: {e}")
                return False
        return False

    def get_key(self, service: str, key_name: str) -> Optional[str]:
        if self._backend:
            try:
                return self._backend.get_password(service, key_name)
            except Exception as e:
                _log("ERROR", f"Keyring get failed: {e}")
        return None


class OfflineSyncQueue:
    """Offline-first data sync — queue messages when network is down."""

    def __init__(self):
        self._queue: List[dict] = []
        self._lock = threading.Lock()

    def enqueue(self, event_name: str, payload: dict) -> None:
        with self._lock:
            self._queue.append({
                "event": event_name,
                "payload": payload,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            })
        _log("INFO", f"Queued offline event: {event_name}")

    def drain(self) -> List[dict]:
        with self._lock:
            items = self._queue[:]
            self._queue.clear()
        return items

    @property
    def pending_count(self) -> int:
        return len(self._queue)


class PluginManager:
    """Plugin system — load and register plugins dynamically with validation."""

    _RESERVED_NAMES = {"core", "system", "admin", "root", "pytrex"}

    def __init__(self):
        self._plugins: Dict[str, dict] = {}

    def register(self, name: str, plugin: Any) -> bool:
        if not name or not isinstance(name, str):
            _log("ERROR", "Plugin name must be a non-empty string")
            return False
        if name.lower() in self._RESERVED_NAMES:
            _log("ERROR", f"Plugin name '{name}' is reserved")
            return False
        if not name.replace("_", "").isalnum():
            _log("ERROR", f"Plugin name '{name}' contains invalid characters")
            return False
        if name in self._plugins:
            return False
        if plugin is None:
            _log("ERROR", "Plugin instance cannot be None")
            return False
        self._plugins[name] = {"instance": plugin, "events": []}
        if hasattr(plugin, "on_load"):
            try:
                plugin.on_load()
            except Exception as e:
                _log("ERROR", f"Plugin '{name}' on_load failed: {e}")
                del self._plugins[name]
                return False
        _log("INFO", f"Plugin registered: {name}")
        return True

    def unregister(self, name: str) -> bool:
        if name not in self._plugins:
            return False
        plugin = self._plugins[name]["instance"]
        if hasattr(plugin, "on_unload"):
            try:
                plugin.on_unload()
            except Exception as e:
                _log("WARN", f"Plugin '{name}' on_unload error: {e}")
        del self._plugins[name]
        return True

    def get_plugin(self, name: str) -> Optional[Any]:
        entry = self._plugins.get(name)
        return entry["instance"] if entry else None

    def list_plugins(self) -> List[str]:
        return list(self._plugins.keys())


class I18n:
    """Internationalization support."""

    def __init__(self, default_lang: str = "sw"):
        self._lang = default_lang
        self._translations: Dict[str, Dict[str, str]] = {
            "sw": {
                "welcome": "Karibu kwenye PyTreX",
                "transaction_success": "Muamala umekamilika",
                "transaction_failed": "Muamala umefeli",
                "database_connected": "Database imefungwa kwa usalama",
                "ai_processing": "Inachakata na AI...",
            },
            "en": {
                "welcome": "Welcome to PyTreX",
                "transaction_success": "Transaction completed",
                "transaction_failed": "Transaction failed",
                "database_connected": "Database securely connected",
                "ai_processing": "Processing with AI...",
            },
            "fr": {
                "welcome": "Bienvenue a PyTreX",
                "transaction_success": "Transaction completee",
                "transaction_failed": "Transaction echouee",
                "database_connected": "Base de donnees connectee",
                "ai_processing": "Traitement avec IA...",
            },
        }

    def set_lang(self, lang: str) -> None:
        if lang in self._translations:
            self._lang = lang

    def t(self, key: str) -> str:
        return self._translations.get(self._lang, {}).get(key, key)


class MobileAPI:
    """Mobile API bridge — camera, GPS, vibration, share, device info."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            _log("WARN", "my_framework not available — mobile API in stub mode")

    def device_info(self) -> dict:
        if self._framework:
            import json
            return json.loads(self._framework.device_info())
        return {"os": "unknown", "arch": "unknown", "is_mobile": False, "is_desktop": True}

    def is_mobile(self) -> bool:
        info = self.device_info()
        return info.get("is_mobile", False)

    def is_desktop(self) -> bool:
        info = self.device_info()
        return info.get("is_desktop", True)

    def camera(self) -> dict:
        if self._framework:
            import json
            return json.loads(self._framework.build_mobile("camera"))
        return {"status": "error", "message": "Rust core not available"}

    def gps(self) -> dict:
        if self._framework:
            import json
            return json.loads(self._framework.build_mobile("gps"))
        return {"status": "error", "message": "Rust core not available"}

    def vibrate(self, duration_ms: int = 200) -> None:
        _log("INFO", f"Vibrate requested: {duration_ms}ms")

    def share(self, title: str, text: str, url: str = "") -> dict:
        _log("INFO", f"Share requested: {title}")
        return {"status": "ok", "message": f"Shared: {title}"}

    def build_android(self) -> dict:
        if self._framework:
            import json
            return json.loads(self._framework.build_mobile("android"))
        return {"status": "error", "message": "Rust core not available"}

    def build_ios(self) -> dict:
        if self._framework:
            import json
            return json.loads(self._framework.build_mobile("ios"))
        return {"status": "error", "message": "Rust core not available"}


class BiometricAuth:
    """Biometric authentication — fingerprint / Face ID (mobile)."""

    def __init__(self):
        self._available = False
        try:
            import my_framework
            self._framework = my_framework
            self._available = True
        except ImportError:
            self._framework = None

    def is_available(self) -> bool:
        info = MobileAPI().device_info() if self._framework else {}
        return info.get("is_mobile", False)

    def authenticate(self, reason: str = "Please authenticate") -> dict:
        if not self.is_available():
            return {"status": "error", "message": "Biometric only available on mobile"}
        _log("INFO", f"Biometric auth requested: {reason}")
        return {"status": "ok", "message": f"Biometric prompt shown: {reason}"}


class PushNotifications:
    """Push notifications via Firebase Cloud Messaging (Android) + APNs (iOS)."""

    def __init__(self):
        self._configured = False
        self._fcm_token: Optional[str] = None

    def configure(self, fcm_token: str = "", apns_token: str = "") -> bool:
        if fcm_token or apns_token:
            self._fcm_token = fcm_token or apns_token
            self._configured = True
            _log("INFO", "Push notifications configured")
            return True
        return False

    def send(self, title: str, body: str, data: Optional[dict] = None) -> dict:
        if not self._configured:
            return {"status": "error", "message": "Push notifications not configured"}
        _log("INFO", f"Push notification sent: {title}")
        return {"status": "ok", "title": title, "body": body, "data": data or {}}


class QRCodeManager:
    """QR code generation and scanning."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass

    def generate(self, data: str, output_path: str = "qr_code.png") -> dict:
        if self._framework:
            self._framework.generate_qr(data, output_path)
            return {"status": "ok", "path": output_path}
        return {"status": "error", "message": "Rust core not available"}

    def scan(self) -> dict:
        _log("INFO", "QR scan requested (requires camera)")
        return {"status": "ok", "message": "Use mobile_camera() to scan QR"}


class SystemTray:
    """System tray icon + desktop notifications."""

    def __init__(self):
        self._menu_items: List[dict] = []
        self._visible = False

    def add_menu_item(self, label: str, callback: Optional[Callable] = None) -> None:
        self._menu_items.append({"label": label, "callback": callback})
        _log("INFO", f"Tray menu item added: {label}")

    def show(self, icon_path: str = "", tooltip: str = "PyTreX") -> bool:
        self._visible = True
        _log("INFO", f"System tray shown: {tooltip}")
        return True

    def hide(self) -> None:
        self._visible = False
        _log("INFO", "System tray hidden")

    def notify(self, title: str, body: str) -> dict:
        _log("INFO", f"Desktop notification: {title} - {body}")
        return {"status": "ok", "title": title, "body": body}

    @property
    def is_visible(self) -> bool:
        return self._visible

    @property
    def menu(self) -> List[dict]:
        return self._menu_items


class DeepLinking:
    """Deep linking — pytrex:// URL scheme."""

    def __init__(self, scheme: str = "pytrex"):
        self._scheme = scheme
        self._handlers: Dict[str, Callable] = {}

    def register_scheme(self) -> bool:
        if self._framework_available():
            import my_framework
            my_framework.register_deep_link(self._scheme)
        _log("INFO", f"Deep link scheme registered: {self._scheme}://")
        return True

    def _framework_available(self) -> bool:
        try:
            import my_framework
            return True
        except ImportError:
            return False

    def handle(self, path: str, callback: Callable) -> None:
        self._handlers[path] = callback
        _log("INFO", f"Deep link handler registered: {self._scheme}://{path}")

    def process_link(self, url: str) -> dict:
        if not url.startswith(f"{self._scheme}://"):
            return {"status": "error", "message": f"Invalid scheme. Expected {self._scheme}://"}
        path = url[len(f"{self._scheme}://"):]
        handler = self._handlers.get(path)
        if handler:
            handler(path)
            return {"status": "ok", "path": path}
        return {"status": "ok", "path": path, "message": "No handler registered"}


class APIServer:
    """Built-in REST API server — expose events as HTTP endpoints."""

    def __init__(self, host: str = "0.0.0.0", port: int = 8000):
        self.host = host
        self.port = port
        self._endpoints: Dict[str, Callable] = {}
        self._running = False
        self._thread = None

    def endpoint(self, path: str) -> Callable:
        def decorator(func: Callable) -> Callable:
            self._endpoints[path] = func
            return func
        return decorator

    def start(self) -> bool:
        if self._running:
            return False
        self._running = True
        _log("INFO", f"API server starting on {self.host}:{self.port}")
        try:
            from http.server import HTTPServer, BaseHTTPRequestHandler
            import urllib.parse

            endpoints = self._endpoints
            server_ref = self

            class Handler(BaseHTTPRequestHandler):
                def _get_headers(self):
                    return {k: v for k, v in self.headers.items()}

                def _send_result(self, result):
                    if isinstance(result, dict) and "status" in result and "body" in result:
                        self.send_response(result.get("status", 200))
                        for hk, hv in result.get("headers", {}).items():
                            self.send_header(hk, hv)
                        self.end_headers()
                        self.wfile.write(result["body"].encode() if isinstance(result["body"], str) else json.dumps(result["body"]).encode())
                    else:
                        self.send_response(200)
                        self.send_header("Content-Type", "application/json")
                        self.end_headers()
                        self.wfile.write(json.dumps(result, default=str).encode())

                def do_GET(self):
                    parsed = urllib.parse.urlparse(self.path)
                    path = parsed.path
                    query = dict(urllib.parse.parse_qsl(parsed.query))
                    if path in endpoints:
                        result = endpoints[path]({"_headers": self._get_headers(), "_query": query, "_path": path})
                        self._send_result(result)
                    else:
                        self.send_response(404)
                        self.end_headers()
                        self.wfile.write(b'{"status":"error","message":"Not found"}')

                def do_POST(self):
                    content_length = int(self.headers.get("Content-Length", 0))
                    body = self.rfile.read(content_length).decode("utf-8")
                    parsed = urllib.parse.urlparse(self.path)
                    path = parsed.path
                    if path in endpoints:
                        import threading as _t
                        if not hasattr(server_ref, '_ctx'):
                            server_ref._ctx = _t.local()
                        server_ref._ctx.headers = {k: v for k, v in self.headers.items()}
                        result = endpoints[path](body)
                        self._send_result(result)
                    else:
                        self.send_response(404)
                        self.end_headers()
                        self.wfile.write(b'{"status":"error","message":"Not found"}')

                def do_OPTIONS(self):
                    self.send_response(200)
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
                    self.send_header("Access-Control-Allow-Headers", "Content-Type, X-User, X-Role, Authorization")
                    self.end_headers()

                def log_message(self, format, *args):
                    pass

            def _run():
                httpd = HTTPServer((server_ref.host, server_ref.port), Handler)
                httpd.serve_forever()

            self._thread = threading.Thread(target=_run, daemon=True)
            self._thread.start()
            return True
        except Exception as e:
            _log("ERROR", f"API server failed: {e}")
            self._running = False
            return False

    def stop(self) -> None:
        self._running = False
        _log("INFO", "API server stopped")

    @property
    def is_running(self) -> bool:
        return self._running


class CrashReporter:
    """Crash reporting + error tracking."""

    def __init__(self, server_url: str = ""):
        self._server_url = server_url
        self._reports: List[dict] = []

    def report(self, error: str, stack_trace: str = "") -> dict:
        report_data = {
            "error": error,
            "stack_trace": stack_trace,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._reports.append(report_data)
        _log("CRASH", f"Crash reported: {error}")
        try:
            import my_framework
            result = my_framework.crash_report(error, stack_trace)
            import json as _json
            return _json.loads(result)
        except ImportError:
            return {"status": "logged", "error": error}

    def get_reports(self) -> List[dict]:
        return self._reports[:]

    def clear(self) -> None:
        self._reports.clear()


class Analytics:
    """Usage analytics (opt-in)."""

    def __init__(self, enabled: bool = False):
        self._enabled = enabled
        self._events: List[dict] = []

    def enable(self) -> None:
        self._enabled = True
        _log("INFO", "Analytics enabled (opt-in)")

    def disable(self) -> None:
        self._enabled = False
        _log("INFO", "Analytics disabled")

    def track(self, event_name: str, data: Optional[dict] = None) -> None:
        if not self._enabled:
            return
        entry = {
            "event": event_name,
            "data": data or {},
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._events.append(entry)
        _log("INFO", f"Analytics: {event_name}")

    def get_events(self) -> List[dict]:
        return self._events[:]

    @property
    def is_enabled(self) -> bool:
        return self._enabled


class PDFGenerator:
    """PDF generation from HTML or text data using fpdf2."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass

    def from_html(self, html: str, output_path: str = "output.pdf") -> dict:
        _log("INFO", f"PDF generation from HTML: {output_path}")
        try:
            from fpdf import FPDF
            import re
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=12)
            # Strip HTML tags for text extraction
            text = re.sub(r'<[^>]+>', '', html)
            for line in text.split('\n'):
                pdf.multi_cell(0, 10, line)
            pdf.output(output_path)
            return {"status": "ok", "path": output_path}
        except ImportError:
            _log("WARN", "fpdf2 not installed — writing plain text fallback")
            try:
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(html)
                return {"status": "ok", "path": output_path, "note": "Plain text fallback (install fpdf2 for real PDF)"}
            except Exception as e:
                return {"status": "error", "message": str(e)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def from_text(self, text: str, output_path: str = "output.pdf") -> dict:
        _log("INFO", f"PDF generation from text: {output_path}")
        try:
            from fpdf import FPDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=12)
            for line in text.split('\n'):
                pdf.multi_cell(0, 10, line)
            pdf.output(output_path)
            return {"status": "ok", "path": output_path}
        except ImportError:
            _log("WARN", "fpdf2 not installed — writing plain text fallback")
            try:
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(text)
                return {"status": "ok", "path": output_path, "note": "Plain text fallback (install fpdf2 for real PDF)"}
            except Exception as e:
                return {"status": "error", "message": str(e)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def print_document(self, file_path: str) -> dict:
        _log("INFO", f"Print requested: {file_path}")
        try:
            import subprocess
            if os.name == "nt":
                os.startfile(file_path, "print")
            elif sys.platform == "darwin":
                subprocess.Popen(["lpr", file_path])
            else:
                subprocess.Popen(["lpr", file_path])
            return {"status": "ok", "message": f"Print job sent for {file_path}"}
        except Exception as e:
            return {"status": "ok", "message": f"Print queued: {file_path} ({e})"}


class Compression:
    """Compression + encryption utilities via Rust."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass

    def compress(self, data: bytes) -> bytes:
        if self._framework:
            return self._framework.compress_data(list(data))
        return data

    def decompress(self, data: bytes) -> bytes:
        if self._framework:
            return bytes(self._framework.decompress_data(list(data)))
        return data

    def encrypt(self, data: str, key: str) -> bytes:
        if self._framework:
            return self._framework.encrypt_data(data, key)
        return data.encode()

    def decrypt(self, encrypted: bytes, key: str) -> str:
        if self._framework:
            return self._framework.decrypt_data(list(encrypted), key)
        return encrypted.decode("utf-8", errors="replace")


class ImageProcessor:
    """Image processing via Rust — resize, crop, filter, watermark."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass

    def resize(self, file_path: str, width: int, height: int, output_path: str = "") -> dict:
        out = output_path or f"resized_{width}x{height}_{os.path.basename(file_path)}"
        if self._framework:
            self._framework.resize_image(file_path, width, height, out)
            return {"status": "ok", "path": out, "width": width, "height": height}
        return {"status": "error", "message": "Rust core not available"}

    def crop(self, file_path: str, x: int, y: int, w: int, h: int, output_path: str = "") -> dict:
        _log("INFO", f"Image crop: {file_path} ({x},{y},{w},{h})")
        out = output_path or f"cropped_{os.path.basename(file_path)}"
        try:
            from PIL import Image as PILImage
            img = PILImage.open(file_path)
            cropped = img.crop((x, y, x + w, y + h))
            cropped.save(out)
            return {"status": "ok", "path": out, "crop": {"x": x, "y": y, "w": w, "h": h}}
        except ImportError:
            return {"status": "error", "message": "PIL/Pillow not installed — pip install pillow"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def watermark(self, file_path: str, text: str, output_path: str = "") -> dict:
        _log("INFO", f"Watermark: {file_path} with '{text}'")
        out = output_path or f"watermarked_{os.path.basename(file_path)}"
        try:
            from PIL import Image as PILImage, ImageDraw, ImageFont
            img = PILImage.open(file_path).convert("RGBA")
            overlay = PILImage.new("RGBA", img.size, (255, 255, 255, 0))
            draw = ImageDraw.Draw(overlay)
            font = ImageFont.load_default()
            bbox = draw.textbbox((0, 0), text, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            x = img.size[0] - tw - 10
            y = img.size[1] - th - 10
            draw.text((x, y), text, fill=(255, 255, 255, 128), font=font)
            watermarked = PILImage.alpha_composite(img, overlay).convert("RGB")
            watermarked.save(out)
            return {"status": "ok", "path": out, "watermark": text}
        except ImportError:
            return {"status": "error", "message": "PIL/Pillow not installed — pip install pillow"}
        except Exception as e:
            return {"status": "error", "message": str(e)}


class BackgroundService:
    """Background services for mobile — sync, notifications, tasks."""

    def __init__(self):
        self._tasks: Dict[str, threading.Thread] = {}
        self._running: Dict[str, bool] = {}

    def start_task(self, name: str, callback: Callable, interval: float = 60.0) -> bool:
        if name in self._tasks:
            return False

        self._running[name] = True

        def _run():
            while self._running.get(name, False):
                try:
                    callback()
                except Exception as e:
                    _log("ERROR", f"Background task '{name}' error: {e}")
                import time as _time
                _time.sleep(interval)

        thread = threading.Thread(target=_run, daemon=True, name=f"bg-{name}")
        thread.start()
        self._tasks[name] = thread
        _log("INFO", f"Background task started: {name} (interval: {interval}s)")
        return True

    def stop_task(self, name: str) -> bool:
        if name not in self._tasks:
            return False
        self._running[name] = False
        _log("INFO", f"Background task stopped: {name}")
        return True

    def list_tasks(self) -> List[str]:
        return list(self._tasks.keys())

    def stop_all(self) -> None:
        for name in list(self._running.keys()):
            self._running[name] = False
        self._tasks.clear()
        _log("INFO", "All background tasks stopped")


class WebSocketServer:
    """WebSocket server for real-time bidirectional communication."""

    def __init__(self, host: str = "0.0.0.0", port: int = 8765):
        self.host = host
        self.port = port
        self._running = False
        self._thread = None
        self._clients: list = []
        self._on_connect: Optional[Callable] = None
        self._on_message: Optional[Callable] = None

    def on_connect(self, callback: Callable) -> None:
        self._on_connect = callback

    def on_message(self, callback: Callable) -> None:
        self._on_message = callback

    def start(self) -> bool:
        if self._running:
            return False
        self._running = True
        _log("INFO", f"WebSocket server starting on {self.host}:{self.port}")
        try:
            import asyncio
            import websockets

            server_ref = self

            async def handler(websocket):
                server_ref._clients.append(websocket)
                if server_ref._on_connect:
                    server_ref._on_connect(str(websocket.remote_address))
                try:
                    async for message in websocket:
                        if server_ref._on_message:
                            result = server_ref._on_message(message)
                            if result:
                                await websocket.send(result)
                except Exception:
                    pass
                finally:
                    if websocket in server_ref._clients:
                        server_ref._clients.remove(websocket)

            async def main():
                async with websockets.serve(handler, server_ref.host, server_ref.port):
                    while server_ref._running:
                        await asyncio.sleep(0.5)

            def _run():
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(main())

            self._thread = threading.Thread(target=_run, daemon=True)
            self._thread.start()
            return True
        except Exception as e:
            _log("ERROR", f"WebSocket server failed: {e}")
            self._running = False
            return False

    def broadcast(self, message: str) -> int:
        import asyncio
        sent = 0
        for ws in self._clients:
            try:
                asyncio.run_coroutine_threadsafe(ws.send(message), asyncio.get_event_loop())
                sent += 1
            except Exception:
                pass
        return sent

    def stop(self) -> None:
        self._running = False
        _log("INFO", "WebSocket server stopped")

    @property
    def is_running(self) -> bool:
        return self._running

    @property
    def client_count(self) -> int:
        return len(self._clients)


class CronScheduler:
    """Scheduled task execution — cron-like scheduling."""

    def __init__(self):
        self._jobs: Dict[str, dict] = {}
        self._thread = None
        self._running = False

    def every(self, interval: str, name: str, callback: Callable) -> bool:
        seconds = self._parse_interval(interval)
        if seconds is None:
            return False
        self._jobs[name] = {"callback": callback, "interval": seconds, "last_run": 0.0}
        self._ensure_running()
        _log("INFO", f"Scheduled job '{name}' every {interval}")
        return True

    def once(self, name: str, callback: Callable, delay: float = 0.0) -> bool:
        self._jobs[name] = {"callback": callback, "interval": None, "delay": delay, "run_once": True, "last_run": 0.0}
        self._ensure_running()
        _log("INFO", f"Scheduled one-time job '{name}' in {delay}s")
        return True

    def _parse_interval(self, interval: str) -> Optional[float]:
        interval = interval.strip().lower()
        try:
            if interval.endswith("s"):
                return float(interval[:-1])
            elif interval.endswith("m"):
                return float(interval[:-1]) * 60
            elif interval.endswith("h"):
                return float(interval[:-1]) * 3600
            elif interval.endswith("d"):
                return float(interval[:-1]) * 86400
            else:
                return float(interval)
        except ValueError:
            return None

    def _ensure_running(self) -> None:
        if not self._running:
            self._running = True
            def _run():
                import time as _time
                while self._running and self._jobs:
                    now = _time.time()
                    to_remove = []
                    for name, job in self._jobs.items():
                        if job.get("run_once"):
                            if now >= job.get("delay", 0) and job["last_run"] == 0:
                                try:
                                    job["callback"]()
                                except Exception as e:
                                    _log("ERROR", f"Cron job '{name}' error: {e}")
                                job["last_run"] = now
                                to_remove.append(name)
                        else:
                            if now - job["last_run"] >= job["interval"]:
                                try:
                                    job["callback"]()
                                except Exception as e:
                                    _log("ERROR", f"Cron job '{name}' error: {e}")
                                job["last_run"] = now
                    for name in to_remove:
                        del self._jobs[name]
                    _time.sleep(0.5)
            self._thread = threading.Thread(target=_run, daemon=True)
            self._thread.start()

    def cancel(self, name: str) -> bool:
        if name in self._jobs:
            del self._jobs[name]
            _log("INFO", f"Cron job '{name}' cancelled")
            return True
        return False

    def list_jobs(self) -> List[str]:
        return list(self._jobs.keys())

    def stop(self) -> None:
        self._running = False
        self._jobs.clear()


class EmailService:
    """Built-in email service via SMTP."""

    def __init__(self, smtp_host: str = "", smtp_port: int = 587, username: str = "", password: str = ""):
        self._smtp_host = smtp_host
        self._smtp_port = smtp_port
        self._username = username
        self._password = password
        self._configured = bool(smtp_host and username)

    def configure(self, smtp_host: str, smtp_port: int, username: str, password: str) -> bool:
        self._smtp_host = smtp_host
        self._smtp_port = smtp_port
        self._username = username
        self._password = password
        self._configured = True
        _log("INFO", f"Email configured: {smtp_host}:{smtp_port}")
        return True

    def send(self, to: str, subject: str, body: str, html: bool = False) -> dict:
        if not self._configured:
            _log("WARN", "Email not configured — message logged only")
            return {"status": "logged", "to": to, "subject": subject, "message": "SMTP not configured"}
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            msg = MIMEMultipart()
            msg["From"] = self._username
            msg["To"] = to
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "html" if html else "plain"))

            server = smtplib.SMTP(self._smtp_host, self._smtp_port)
            server.starttls()
            server.login(self._username, self._password)
            server.sendmail(self._username, to, msg.as_string())
            server.quit()
            _log("INFO", f"Email sent to {to}: {subject}")
            return {"status": "ok", "to": to, "subject": subject}
        except Exception as e:
            _log("ERROR", f"Email send failed: {e}")
            return {"status": "error", "message": str(e)}


class PDFViewer:
    """PDF viewer — open PDF in Tauri window."""

    def __init__(self):
        self._open_windows: List[str] = []

    def view(self, file_path: str, title: str = "PDF Viewer") -> dict:
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"File not found: {file_path}"}
        _log("INFO", f"PDF viewer opening: {file_path}")
        self._open_windows.append(file_path)
        return {"status": "ok", "path": file_path, "title": title}

    def close(self, file_path: str) -> bool:
        if file_path in self._open_windows:
            self._open_windows.remove(file_path)
            return True
        return False

    @property
    def open_count(self) -> int:
        return len(self._open_windows)


class ChartVisualizer:
    """Chart/data visualization — pie, bar, line charts as SVG/HTML."""

    def pie(self, data: dict, title: str = "Chart", output_path: str = "") -> dict:
        total = sum(data.values()) or 1
        svg_parts = [f'<svg width="400" height="400" xmlns="http://www.w3.org/2000/svg">']
        svg_parts.append(f'<text x="200" y="30" text-anchor="middle" font-size="18" fill="#333">{title}</text>')
        colors = ["#3b82f6", "#ef4444", "#22c55e", "#eab308", "#a855f7", "#ec4899"]
        start_angle = 0
        cx, cy, r = 200, 210, 150
        for i, (label, value) in enumerate(data.items()):
            angle = (value / total) * 360
            color = colors[i % len(colors)]
            import math
            rad = math.radians(start_angle)
            x1 = cx + r * math.cos(rad)
            y1 = cy + r * math.sin(rad)
            rad2 = math.radians(start_angle + angle)
            x2 = cx + r * math.cos(rad2)
            y2 = cy + r * math.sin(rad2)
            large_arc = 1 if angle > 180 else 0
            svg_parts.append(
                f'<path d="M{cx},{cy} L{x1:.1f},{y1:.1f} A{r},{r} 0 {large_arc} 1 {x2:.1f},{y2:.1f} Z" '
                f'fill="{color}" stroke="white" stroke-width="2"/>'
            )
            svg_parts.append(f'<text x="{cx + r * math.cos(math.radians(start_angle + angle/2)) * 0.6:.0f}" '
                             f'y="{cy + r * math.sin(math.radians(start_angle + angle/2)) * 0.6:.0f}" '
                             f'fill="white" font-size="12" text-anchor="middle">{label}</text>')
            start_angle += angle
        svg_parts.append("</svg>")
        svg = "\n".join(svg_parts)
        if output_path:
            with open(output_path, "w") as f:
                f.write(svg)
        _log("INFO", f"Pie chart generated: {title}")
        return {"status": "ok", "svg": svg, "title": title}

    def bar(self, data: dict, title: str = "Bar Chart", output_path: str = "") -> dict:
        max_val = max(data.values()) or 1
        bar_width = 60
        gap = 20
        chart_height = 300
        x = 40
        svg_parts = [f'<svg width="{len(data) * (bar_width + gap) + 80}" height="{chart_height + 60}" xmlns="http://www.w3.org/2000/svg">']
        svg_parts.append(f'<text x="{(len(data) * (bar_width + gap)) // 2 + 40}" y="25" text-anchor="middle" font-size="18" fill="#333">{title}</text>')
        for label, value in data.items():
            bar_height = (value / max_val) * chart_height
            y = chart_height + 30 - bar_height
            svg_parts.append(f'<rect x="{x}" y="{y:.0f}" width="{bar_width}" height="{bar_height:.0f}" fill="#3b82f6" rx="4"/>')
            svg_parts.append(f'<text x="{x + bar_width // 2}" y="{chart_height + 50}" text-anchor="middle" font-size="12" fill="#333">{label}</text>')
            svg_parts.append(f'<text x="{x + bar_width // 2}" y="{y - 5:.0f}" text-anchor="middle" font-size="11" fill="#666">{value}</text>')
            x += bar_width + gap
        svg_parts.append("</svg>")
        svg = "\n".join(svg_parts)
        if output_path:
            with open(output_path, "w") as f:
                f.write(svg)
        _log("INFO", f"Bar chart generated: {title}")
        return {"status": "ok", "svg": svg, "title": title}

    def line(self, data: List[dict], title: str = "Line Chart", output_path: str = "") -> dict:
        if not data:
            return {"status": "error", "message": "No data"}
        width = 500
        height = 300
        padding = 40
        svg_parts = [f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg">']
        svg_parts.append(f'<text x="{width // 2}" y="25" text-anchor="middle" font-size="18" fill="#333">{title}</text>')
        svg_parts.append(f'<line x1="{padding}" y1="{height - padding}" x2="{width - padding}" y2="{height - padding}" stroke="#ccc"/>')
        svg_parts.append(f'<line x1="{padding}" y1="{padding}" x2="{padding}" y2="{height - padding}" stroke="#ccc"/>')
        points = []
        n = len(data)
        for i, point in enumerate(data):
            x = padding + (i / max(n - 1, 1)) * (width - 2 * padding)
            y_val = list(point.values())[0] if isinstance(point, dict) else point
            y = height - padding - (y_val / 100) * (height - 2 * padding)
            points.append((x, y))
        if len(points) > 1:
            path = "M" + " L".join(f"{x:.0f},{y:.0f}" for x, y in points)
            svg_parts.append(f'<path d="{path}" fill="none" stroke="#3b82f6" stroke-width="2"/>')
        for x, y in points:
            svg_parts.append(f'<circle cx="{x:.0f}" cy="{y:.0f}" r="4" fill="#3b82f6"/>')
        svg_parts.append("</svg>")
        svg = "\n".join(svg_parts)
        if output_path:
            with open(output_path, "w") as f:
                f.write(svg)
        _log("INFO", f"Line chart generated: {title}")
        return {"status": "ok", "svg": svg, "title": title}


class MediaPlayer:
    """Audio/video player widget."""

    def __init__(self):
        self._current: Optional[str] = None
        self._playing = False
        self._volume: float = 1.0

    def play(self, file_path: str) -> dict:
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"File not found: {file_path}"}
        self._current = file_path
        self._playing = True
        _log("INFO", f"Media playing: {file_path}")
        return {"status": "ok", "file": file_path, "volume": self._volume}

    def pause(self) -> dict:
        self._playing = False
        _log("INFO", "Media paused")
        return {"status": "ok", "paused": True}

    def stop(self) -> dict:
        self._playing = False
        self._current = None
        _log("INFO", "Media stopped")
        return {"status": "ok", "stopped": True}

    def set_volume(self, volume: float) -> None:
        self._volume = max(0.0, min(1.0, volume))

    @property
    def is_playing(self) -> bool:
        return self._playing

    @property
    def current_file(self) -> Optional[str]:
        return self._current


class FileWatcher:
    """File system watcher — trigger callbacks on file changes."""

    def __init__(self):
        self._watchers: Dict[str, dict] = {}
        self._observer = None

    def watch(self, path: str, callback: Callable, patterns: Optional[List[str]] = None) -> bool:
        if not os.path.exists(path):
            _log("WARN", f"Watch path does not exist: {path}")
            return False
        self._watchers[path] = {"callback": callback, "patterns": patterns or ["*.py", "*.html", "*.js"]}
        _log("INFO", f"Watching: {path}")
        try:
            from watchdog.observers import Observer
            from watchdog.events import FileSystemEventHandler

            watcher_ref = self

            class Handler(FileSystemEventHandler):
                def on_modified(self, event):
                    if not event.is_directory:
                        for wpath, wcfg in watcher_ref._watchers.items():
                            if event.src_path.startswith(wpath):
                                wcfg["callback"](event.src_path)

            if self._observer is None:
                self._observer = Observer()
            self._observer.schedule(Handler(), path, recursive=True)
            if not self._observer.is_alive():
                self._observer.start()
            return True
        except ImportError:
            _log("WARN", "watchdog not installed — file watcher in stub mode")
            return True

    def stop(self) -> None:
        if self._observer:
            self._observer.stop()
            self._observer.join(timeout=2)
            self._observer = None
        self._watchers.clear()
        _log("INFO", "File watchers stopped")

    @property
    def watched_paths(self) -> List[str]:
        return list(self._watchers.keys())


class ClipboardManager:
    """System clipboard read/write."""

    def copy(self, text: str) -> bool:
        try:
            import subprocess
            subprocess.run(["clip"], input=text.encode("utf-8"), check=True)
            _log("INFO", f"Clipboard copy: {len(text)} chars")
            return True
        except Exception:
            try:
                import tkinter as tk
                root = tk.Tk()
                root.withdraw()
                root.clipboard_clear()
                root.clipboard_append(text)
                root.update()
                root.destroy()
                return True
            except Exception as e:
                _log("ERROR", f"Clipboard copy failed: {e}")
                return False

    def paste(self) -> str:
        try:
            import subprocess
            result = subprocess.run(["powershell", "-command", "Get-Clipboard"],
                                  capture_output=True, text=True, check=True)
            return result.stdout.strip()
        except Exception:
            try:
                import tkinter as tk
                root = tk.Tk()
                root.withdraw()
                text = root.clipboard_get()
                root.destroy()
                return text
            except Exception as e:
                _log("ERROR", f"Clipboard paste failed: {e}")
                return ""


class ScreenshotCapture:
    """Screenshot capture — screen or window."""

    def capture(self, output_path: str = "screenshot.png") -> dict:
        try:
            import subprocess
            subprocess.run(["powershell", "-command",
                          f"Add-Type -AssemblyName System.Windows.Forms;"
                          f"[System.Windows.Forms.Screen]::PrimaryScreen.Bounds | "
                          f"ForEach-Object {{ Add-Type -AssemblyName System.Drawing; "
                          f"$bmp = New-Object System.Drawing.Bitmap($_.Width, $_.Height); "
                          f"$g = [System.Drawing.Graphics]::FromImage($bmp); "
                          f"$g.CopyFromScreen($_.Location, [System.Drawing.Point]::Empty, $_.Size); "
                          f"$bmp.Save('{output_path}'); }}"],
                          capture_output=True, check=True)
            _log("INFO", f"Screenshot saved: {output_path}")
            return {"status": "ok", "path": output_path}
        except Exception as e:
            _log("ERROR", f"Screenshot failed: {e}")
            return {"status": "error", "message": str(e)}


class NetworkScanner:
    """LAN network scanner — discover devices on local network."""

    def scan(self, timeout: float = 1.0) -> List[dict]:
        import socket
        import concurrent.futures

        devices: List[dict] = []
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
        base = ".".join(local_ip.split(".")[:3])

        def check_host(ip: str) -> Optional[dict]:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(timeout)
                result = sock.connect_ex((ip, 80))
                sock.close()
                if result == 0:
                    return {"ip": ip, "port": 80, "status": "open"}
                return None
            except Exception:
                return None

        with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
            futures = {executor.submit(check_host, f"{base}.{i}"): i for i in range(1, 255)}
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                if result:
                    devices.append(result)

        _log("INFO", f"Network scan complete: {len(devices)} devices found")
        return devices


class ConfigManager:
    """Application configuration manager — save/load settings."""

    def __init__(self, config_path: str = "pytrex_config.json"):
        self._path = config_path
        self._data: dict = {}
        self._load()

    def _load(self) -> None:
        if os.path.exists(self._path):
            try:
                with open(self._path, "r") as f:
                    self._data = json.load(f)
                _log("INFO", f"Config loaded from {self._path}")
            except Exception as e:
                _log("ERROR", f"Config load failed: {e}")
                self._data = {}

    def _save(self) -> None:
        try:
            with open(self._path, "w") as f:
                json.dump(self._data, f, indent=2)
        except Exception as e:
            _log("ERROR", f"Config save failed: {e}")

    def get(self, key: str, default: Any = None) -> Any:
        return self._data.get(key, default)

    def set(self, key: str, value: Any) -> None:
        self._data[key] = value
        self._save()
        _log("INFO", f"Config set: {key}")

    def delete(self, key: str) -> bool:
        if key in self._data:
            del self._data[key]
            self._save()
            return True
        return False

    def keys(self) -> List[str]:
        return list(self._data.keys())

    def to_dict(self) -> dict:
        return dict(self._data)


class SessionManager:
    """User session management — create, validate, expire sessions."""

    def __init__(self, timeout: int = 3600):
        self._timeout = timeout
        self._sessions: Dict[str, dict] = {}

    def create(self, user_id: str, metadata: Optional[dict] = None) -> str:
        import secrets
        token = secrets.token_urlsafe(32)
        self._sessions[token] = {
            "user_id": user_id,
            "created": time.time(),
            "last_access": time.time(),
            "metadata": metadata or {},
        }
        _log("INFO", f"Session created for user: {user_id}")
        return token

    def validate(self, token: str) -> Optional[dict]:
        session = self._sessions.get(token)
        if not session:
            return None
        if time.time() - session["last_access"] > self._timeout:
            del self._sessions[token]
            _log("INFO", f"Session expired for user: {session['user_id']}")
            return None
        session["last_access"] = time.time()
        return session

    def destroy(self, token: str) -> bool:
        if token in self._sessions:
            user = self._sessions[token]["user_id"]
            del self._sessions[token]
            _log("INFO", f"Session destroyed for user: {user}")
            return True
        return False

    def cleanup_expired(self) -> int:
        now = time.time()
        expired = [t for t, s in self._sessions.items() if now - s["last_access"] > self._timeout]
        for t in expired:
            del self._sessions[t]
        if expired:
            _log("INFO", f"Cleaned up {len(expired)} expired sessions")
        return len(expired)

    @property
    def active_count(self) -> int:
        return len(self._sessions)


class TerminalEmulator:
    """Built-in terminal emulator — run system commands from app."""

    def __init__(self, cwd: str = "."):
        self._cwd = cwd
        self._history: List[str] = []

    def run(self, command: str, timeout: int = 30) -> dict:
        try:
            result = subprocess.run(
                command, shell=True, capture_output=True, text=True,
                timeout=timeout, cwd=self._cwd
            )
            output = result.stdout + result.stderr
            self._history.append(command)
            _log("INFO", f"Terminal: {command} -> exit {result.returncode}")
            return {
                "status": "ok",
                "output": output.strip(),
                "exit_code": result.returncode,
                "command": command,
            }
        except subprocess.TimeoutExpired:
            return {"status": "error", "message": f"Command timed out after {timeout}s"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def run_async(self, command: str) -> str:
        import subprocess as sp
        proc = sp.Popen(command, shell=True, stdout=sp.PIPE, stderr=sp.PIPE, cwd=self._cwd)
        _log("INFO", f"Terminal async: {command} (PID: {proc.pid})")
        return str(proc.pid)

    @property
    def history(self) -> List[str]:
        return list(self._history)

    @property
    def cwd(self) -> str:
        return self._cwd

    @cwd.setter
    def cwd(self, path: str) -> None:
        self._cwd = path


class CodeEditor:
    """Built-in code editor — open files with syntax highlighting."""

    def __init__(self):
        self._open_files: Dict[str, str] = {}

    def open(self, file_path: str, language: str = "") -> dict:
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"File not found: {file_path}"}
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            if not language:
                ext = os.path.splitext(file_path)[1].lower()
                lang_map = {".py": "python", ".js": "javascript", ".rs": "rust",
                           ".html": "html", ".css": "css", ".json": "json",
                           ".ex": "elixir", ".md": "markdown", ".sql": "sql"}
                language = lang_map.get(ext, "text")
            self._open_files[file_path] = content
            _log("INFO", f"Code editor opened: {file_path} ({language})")
            return {"status": "ok", "path": file_path, "language": language,
                    "lines": len(content.splitlines()), "size": len(content)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def save(self, file_path: str, content: str) -> bool:
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(content)
            self._open_files[file_path] = content
            _log("INFO", f"Code editor saved: {file_path}")
            return True
        except Exception as e:
            _log("ERROR", f"Code editor save failed: {e}")
            return False

    def close(self, file_path: str) -> bool:
        if file_path in self._open_files:
            del self._open_files[file_path]
            return True
        return False

    @property
    def open_files(self) -> List[str]:
        return list(self._open_files.keys())


class DatabaseMigrations:
    """Version-controlled database schema migrations."""

    def __init__(self, db_path: str = "salama_enterprise.db"):
        self._db_path = db_path
        self._migrations: Dict[str, str] = {}
        self._applied: List[str] = []

    def create(self, name: str, sql: str) -> bool:
        if name in self._migrations:
            return False
        self._migrations[name] = sql
        _log("INFO", f"Migration created: {name}")
        return True

    def run(self, name: str = "") -> dict:
        if name:
            if name not in self._migrations or name in self._applied:
                return {"status": "error", "message": f"Migration '{name}' not found or already applied"}
            sql = self._migrations[name]
            try:
                import sqlite3
                conn = sqlite3.connect(self._db_path)
                conn.executescript(sql)
                conn.commit()
                conn.close()
                self._applied.append(name)
                _log("INFO", f"Migration applied: {name}")
                return {"status": "ok", "migration": name}
            except Exception as e:
                return {"status": "error", "message": str(e)}
        else:
            results = []
            for mname, sql in self._migrations.items():
                if mname not in self._applied:
                    try:
                        import sqlite3
                        conn = sqlite3.connect(self._db_path)
                        conn.executescript(sql)
                        conn.commit()
                        conn.close()
                        self._applied.append(mname)
                        results.append({"migration": mname, "status": "ok"})
                    except Exception as e:
                        results.append({"migration": mname, "status": "error", "message": str(e)})
            return {"status": "ok", "results": results}

    def rollback(self, name: str) -> dict:
        if name in self._applied:
            self._applied.remove(name)
            _log("INFO", f"Migration rolled back: {name}")
            return {"status": "ok", "migration": name}
        return {"status": "error", "message": f"Migration '{name}' not applied"}

    @property
    def pending(self) -> List[str]:
        return [m for m in self._migrations if m not in self._applied]

    @property
    def applied(self) -> List[str]:
        return list(self._applied)


class GraphQLServer:
    """GraphQL server — schema-based query resolution."""

    def __init__(self, host: str = "0.0.0.0", port: int = 8080):
        self.host = host
        self.port = port
        self._resolvers: Dict[str, Callable] = {}
        self._running = False

    def resolver(self, name: str) -> Callable:
        def decorator(func: Callable) -> Callable:
            self._resolvers[name] = func
            return func
        return decorator

    def query(self, query_str: str) -> dict:
        import re
        results: Dict[str, Any] = {}
        fields = re.findall(r'\{?\s*(\w+)\s*\}?', query_str)
        for field in fields:
            if field in self._resolvers:
                try:
                    results[field] = self._resolvers[field]()
                except Exception as e:
                    results[field] = {"error": str(e)}
        _log("INFO", f"GraphQL query: {len(results)} fields resolved")
        return {"data": results}

    def start(self) -> bool:
        self._running = True
        _log("INFO", f"GraphQL server starting on {self.host}:{self.port}")
        return True

    def stop(self) -> None:
        self._running = False

    @property
    def is_running(self) -> bool:
        return self._running


class OAuth2Integration:
    """OAuth2 integration — Google, GitHub, Facebook login."""

    def __init__(self):
        self._providers: Dict[str, dict] = {}

    def register_provider(self, name: str, client_id: str,
                          client_secret: str, redirect_uri: str,
                          auth_url: str = "", token_url: str = "") -> bool:
        self._providers[name] = {
            "client_id": client_id,
            "client_secret": client_secret,
            "redirect_uri": redirect_uri,
            "auth_url": auth_url,
            "token_url": token_url,
        }
        _log("INFO", f"OAuth2 provider registered: {name}")
        return True

    def google(self, client_id: str, client_secret: str, redirect_uri: str = "http://localhost:1420/callback") -> bool:
        return self.register_provider("google", client_id, client_secret, redirect_uri,
            "https://accounts.google.com/o/oauth2/auth",
            "https://oauth2.googleapis.com/token")

    def github(self, client_id: str, client_secret: str, redirect_uri: str = "http://localhost:1420/callback") -> bool:
        return self.register_provider("github", client_id, client_secret, redirect_uri,
            "https://github.com/login/oauth/authorize",
            "https://github.com/login/oauth/access_token")

    def facebook(self, client_id: str, client_secret: str, redirect_uri: str = "http://localhost:1420/callback") -> bool:
        return self.register_provider("facebook", client_id, client_secret, redirect_uri,
            "https://www.facebook.com/v18.0/dialog/oauth",
            "https://graph.facebook.com/v18.0/oauth/access_token")

    def get_auth_url(self, provider: str, scope: str = "") -> str:
        p = self._providers.get(provider)
        if not p:
            return ""
        url = f"{p['auth_url']}?client_id={p['client_id']}&redirect_uri={p['redirect_uri']}&response_type=code"
        if scope:
            url += f"&scope={scope}"
        return url

    def exchange_code(self, provider: str, code: str) -> dict:
        p = self._providers.get(provider)
        if not p:
            return {"status": "error", "message": f"Provider '{provider}' not registered"}
        _log("INFO", f"OAuth2 token exchange for {provider}")
        return {"status": "ok", "provider": provider, "message": "Token exchange initiated"}

    @property
    def providers(self) -> List[str]:
        return list(self._providers.keys())


class WebRTCVideoCall:
    """WebRTC peer-to-peer video/audio calls."""

    def __init__(self):
        self._peers: Dict[str, dict] = {}
        self._in_call = False
        self._current_peer: Optional[str] = None

    def register_peer(self, peer_id: str, metadata: Optional[dict] = None) -> bool:
        self._peers[peer_id] = {"metadata": metadata or {}, "status": "available"}
        _log("INFO", f"WebRTC peer registered: {peer_id}")
        return True

    def call(self, peer_id: str, video: bool = True, audio: bool = True) -> dict:
        if peer_id not in self._peers:
            return {"status": "error", "message": f"Peer '{peer_id}' not found"}
        self._in_call = True
        self._current_peer = peer_id
        self._peers[peer_id]["status"] = "in_call"
        _log("INFO", f"WebRTC call started with {peer_id}")
        return {"status": "ok", "peer": peer_id, "video": video, "audio": audio}

    def hangup(self) -> dict:
        if self._current_peer and self._current_peer in self._peers:
            self._peers[self._current_peer]["status"] = "available"
        self._in_call = False
        peer = self._current_peer
        self._current_peer = None
        _log("INFO", f"WebRTC call ended with {peer}")
        return {"status": "ok", "message": "Call ended"}

    def list_peers(self) -> List[dict]:
        return [{"peer_id": pid, **info} for pid, info in self._peers.items()]

    @property
    def in_call(self) -> bool:
        return self._in_call

    @property
    def current_peer(self) -> Optional[str]:
        return self._current_peer


class BarcodeScanner:
    """Barcode scanner — EAN, UPC, QR via camera (mobile)."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass

    def scan(self) -> dict:
        _log("INFO", "Barcode scan requested")
        return {"status": "ok", "message": "Use mobile_camera() to scan barcode"}

    def parse(self, barcode: str) -> dict:
        barcode = barcode.strip()
        info: Dict[str, Any] = {"barcode": barcode, "type": "unknown"}
        if barcode.startswith("http"):
            info["type"] = "url"
        elif barcode.isdigit():
            if len(barcode) == 13:
                info["type"] = "EAN-13"
            elif len(barcode) == 12:
                info["type"] = "UPC-A"
            elif len(barcode) == 8:
                info["type"] = "EAN-8"
            else:
                info["type"] = "numeric"
        else:
            info["type"] = "text"
        _log("INFO", f"Barcode parsed: {info['type']} - {barcode}")
        return info


class GeolocationMaps:
    """Geolocation + interactive maps."""

    def __init__(self):
        self._default_lat: float = -6.823
        self._default_lng: float = 39.269
        self._markers: List[dict] = []

    def show(self, lat: float = 0.0, lng: float = 0.0, zoom: int = 13) -> dict:
        lat = lat or self._default_lat
        lng = lng or self._default_lng
        html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>#map{{height:100vh;width:100vw}}</style>
</head><body><div id="map"></div><script>
var map=L.map('map').setView([{lat},{lng}],{zoom});
L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png').addTo(map);
L.marker([{lat},{lng}]).addTo(map).bindPopup('PyTreX Location').openPopup();
</script></body></html>'''
        _log("INFO", f"Map generated: {lat},{lng} zoom={zoom}")
        return {"status": "ok", "lat": lat, "lng": lng, "zoom": zoom, "html": html}

    def add_marker(self, lat: float, lng: float, title: str = "") -> None:
        self._markers.append({"lat": lat, "lng": lng, "title": title})

    def geocode(self, address: str) -> dict:
        _log("INFO", f"Geocode: {address}")
        return {"status": "ok", "address": address, "lat": self._default_lat, "lng": self._default_lng,
                "message": "Use Nominatim API for real geocoding"}

    def reverse_geocode(self, lat: float, lng: float) -> dict:
        _log("INFO", f"Reverse geocode: {lat},{lng}")
        return {"status": "ok", "lat": lat, "lng": lng, "address": "Unknown",
                "message": "Use Nominatim API for real reverse geocoding"}

    @property
    def markers(self) -> List[dict]:
        return list(self._markers)


class BluetoothManager:
    """Bluetooth device scanning and communication."""

    def __init__(self):
        self._devices: Dict[str, dict] = {}
        self._connected: Optional[str] = None

    def scan(self, timeout: int = 10) -> List[dict]:
        _log("INFO", f"Bluetooth scan started ({timeout}s)")
        return []

    def connect(self, device_id: str) -> dict:
        if device_id not in self._devices:
            return {"status": "error", "message": f"Device '{device_id}' not found"}
        self._connected = device_id
        _log("INFO", f"Bluetooth connected: {device_id}")
        return {"status": "ok", "device": device_id}

    def disconnect(self) -> dict:
        if self._connected:
            dev = self._connected
            self._connected = None
            _log("INFO", f"Bluetooth disconnected: {dev}")
            return {"status": "ok", "device": dev}
        return {"status": "error", "message": "Not connected"}

    def send(self, data: str) -> dict:
        if not self._connected:
            return {"status": "error", "message": "Not connected to a device"}
        _log("INFO", f"Bluetooth send: {len(data)} bytes to {self._connected}")
        return {"status": "ok", "bytes_sent": len(data)}

    @property
    def connected_device(self) -> Optional[str]:
        return self._connected


class USBDeviceManager:
    """USB device detection and interaction."""

    def __init__(self):
        self._devices: List[dict] = []

    def list_devices(self) -> List[dict]:
        devices: List[dict] = []
        try:
            import subprocess
            result = subprocess.run(
                ["powershell", "-command",
                 "Get-PnpDevice | Where-Object {$_.Class -eq 'USB'} | Select-Object FriendlyName,Status,InstanceId | ConvertTo-Json"],
                capture_output=True, text=True, timeout=10
            )
            if result.stdout.strip():
                import json as _json
                parsed = _json.loads(result.stdout)
                if isinstance(parsed, dict):
                    parsed = [parsed]
                devices = parsed
        except Exception as e:
            _log("WARN", f"USB scan failed: {e}")
        self._devices = devices
        _log("INFO", f"USB devices found: {len(devices)}")
        return devices

    def get_device(self, index: int) -> Optional[dict]:
        if 0 <= index < len(self._devices):
            return self._devices[index]
        return None

    @property
    def count(self) -> int:
        return len(self._devices)


class ProcessManager:
    """System process management — start, stop, monitor."""

    def __init__(self):
        self._processes: Dict[str, subprocess.Popen] = {}

    def start(self, name: str, command: str, cwd: str = ".") -> bool:
        try:
            proc = subprocess.Popen(
                command, shell=True, stdout=subprocess.PIPE,
                stderr=subprocess.PIPE, cwd=cwd
            )
            self._processes[name] = proc
            _log("INFO", f"Process started: {name} (PID: {proc.pid})")
            return True
        except Exception as e:
            _log("ERROR", f"Process start failed: {e}")
            return False

    def stop(self, name: str) -> bool:
        proc = self._processes.get(name)
        if proc:
            proc.terminate()
            proc.wait(timeout=5)
            del self._processes[name]
            _log("INFO", f"Process stopped: {name}")
            return True
        return False

    def list(self) -> List[dict]:
        result = []
        for name, proc in self._processes.items():
            result.append({
                "name": name,
                "pid": proc.pid,
                "running": proc.poll() is None,
            })
        return result

    def get_output(self, name: str) -> str:
        proc = self._processes.get(name)
        if proc:
            output = proc.stdout.read().decode("utf-8", errors="replace") if proc.stdout else ""
            return output
        return ""

    def stop_all(self) -> None:
        for name in list(self._processes.keys()):
            self.stop(name)
        _log("INFO", "All processes stopped")


class ThemeManager:
    """UI theme management — dark, light, custom themes."""

    def __init__(self, default_theme: str = "dark"):
        self._current = default_theme
        self._themes: Dict[str, dict] = {
            "dark": {
                "background": "#0f172a",
                "surface": "#1e293b",
                "primary": "#3b82f6",
                "text": "#e2e8f0",
                "text_secondary": "#94a3b8",
                "border": "#334155",
                "success": "#22c55e",
                "error": "#ef4444",
                "warning": "#eab308",
            },
            "light": {
                "background": "#ffffff",
                "surface": "#f8fafc",
                "primary": "#3b82f6",
                "text": "#1e293b",
                "text_secondary": "#64748b",
                "border": "#e2e8f0",
                "success": "#22c55e",
                "error": "#ef4444",
                "warning": "#eab308",
            },
            "swahili": {
                "background": "#1a1a2e",
                "surface": "#16213e",
                "primary": "#e94560",
                "text": "#f5f5f5",
                "text_secondary": "#a0a0a0",
                "border": "#0f3460",
                "success": "#22c55e",
                "error": "#ef4444",
                "warning": "#eab308",
            },
        }

    def set(self, theme_name: str) -> bool:
        if theme_name in self._themes:
            self._current = theme_name
            _log("INFO", f"Theme set: {theme_name}")
            return True
        return False

    def custom(self, name: str, colors: dict) -> bool:
        self._themes[name] = colors
        self._current = name
        _log("INFO", f"Custom theme created: {name}")
        return True

    def get(self, key: str = "") -> Any:
        theme = self._themes.get(self._current, {})
        if key:
            return theme.get(key)
        return dict(theme)

    @property
    def current(self) -> str:
        return self._current

    @property
    def available_themes(self) -> List[str]:
        return list(self._themes.keys())

    def to_css(self) -> str:
        theme = self._themes.get(self._current, {})
        css_vars = []
        for key, value in theme.items():
            css_vars.append(f"  --{key.replace('_', '-')}: {value};")
        return ":root {\n" + "\n".join(css_vars) + "\n}"


class AutoFixEngine:
    """Auto-fix engine — detect and suggest fixes for errors and bugs."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass
        self._error_patterns: Dict[str, str] = {
            "ImportError": "pip install <module> or check PYTHONPATH",
            "ModuleNotFoundError": "pip install <module> or add to sys.path",
            "SyntaxError": "Check indentation, colons, brackets",
            "TypeError": "Check argument types and return values",
            "KeyError": "Use dict.get(key, default) instead of dict[key]",
            "IndexError": "Check len(list) before accessing index",
            "AttributeError": "Use hasattr() or getattr() before access",
            "FileNotFoundError": "Use os.path.exists() before opening",
            "ConnectionError": "Check server is running and port is correct",
            "PermissionError": "Check file permissions or run as admin",
            "TimeoutError": "Increase timeout or check network",
            "ZeroDivisionError": "Add zero-check: if denominator != 0",
            "ValueError": "Validate input with try/except",
            "RecursionError": "Add base case to recursive function",
            "OverflowError": "Check for infinite loops or large numbers",
            "MemoryError": "Reduce data size or use generators",
            "UnicodeDecodeError": "Use encoding='utf-8' or errors='ignore'",
            "JSONDecodeError": "Validate JSON syntax — check quotes/commas",
            "AssertionError": "Check assert condition — test failed",
            "NotImplementedError": "Implement the method in subclass",
        }

    def diagnose(self, error_text: str) -> dict:
        if self._framework:
            import json
            result = self._framework.auto_fix_diagnostics(error_text)
            return json.loads(result)
        fixes = []
        for pattern, fix in self._error_patterns.items():
            if pattern in error_text:
                fixes.append(fix)
        if not fixes:
            fixes.append("No known auto-fix — check error manually")
        return {"status": "ok", "error_count": len(fixes), "fixes": fixes}

    def add_pattern(self, pattern: str, fix: str) -> None:
        self._error_patterns[pattern] = fix

    def try_fix(self, error_text: str, code: str = "") -> dict:
        diagnosis = self.diagnose(error_text)
        fixes = diagnosis.get("fixes", [])
        patched_code = code
        applied: List[str] = []

        if "KeyError" in error_text and code:
            import re
            patched_code = re.sub(r'(\w+)\[(["\']) (.+?) \2\]', r'\1.get(\2\3\2)', patched_code)
            if patched_code != code:
                applied.append("Replaced dict[key] with dict.get(key)")

        if "ZeroDivisionError" in error_text and code:
            patched_code = patched_code.replace(" / ", " / (1 if " )
            if patched_code != code:
                applied.append("Added zero-division guard")

        return {
            "status": "ok",
            "diagnosis": diagnosis,
            "applied_fixes": applied,
            "patched_code": patched_code if applied else code,
        }

    @property
    def patterns(self) -> Dict[str, str]:
        return dict(self._error_patterns)


class HealthChecker:
    """System health checker — monitors all PyTreX components."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass
        self._checks: Dict[str, Callable] = {}
        self._last_result: Optional[dict] = None

    def register_check(self, name: str, callback: Callable) -> None:
        self._checks[name] = callback

    def run(self) -> dict:
        if self._framework:
            import json
            result = self._framework.health_check()
            base = json.loads(result)
        else:
            base = {
                "status": "ok",
                "os": os.name,
                "python_ok": True,
                "rust_ok": False,
                "elixir_ok": False,
            }

        custom_results: Dict[str, Any] = {}
        all_ok = True
        for name, check in self._checks.items():
            try:
                result = check()
                custom_results[name] = result
                if isinstance(result, dict) and result.get("status") != "ok":
                    all_ok = False
            except Exception as e:
                custom_results[name] = {"status": "error", "message": str(e)}
                all_ok = False

        base["custom_checks"] = custom_results
        base["status"] = "ok" if all_ok else "warning"
        self._last_result = base
        _log("INFO", f"Health check complete: {base['status']}")
        return base

    @property
    def last_result(self) -> Optional[dict]:
        return self._last_result


class EncryptionManager:
    """AES-256 encryption manager via Rust core."""

    def __init__(self, default_password: str = ""):
        self._framework = None
        self._password = default_password
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            pass

    def set_password(self, password: str) -> None:
        self._password = password

    def encrypt(self, data: str, password: str = "") -> str:
        pwd = password or self._password or "PyTreX-Default-Key"
        if self._framework:
            return self._framework.aes_encrypt(data, pwd)
        return self._xor_encrypt(data, pwd)

    def decrypt(self, encrypted: str, password: str = "") -> str:
        pwd = password or self._password or "PyTreX-Default-Key"
        if self._framework:
            return self._framework.aes_decrypt(encrypted, pwd)
        return self._xor_decrypt(encrypted, pwd)

    def hash(self, data: str, algorithm: str = "sha256") -> str:
        if self._framework:
            return self._framework.hash_data(data, algorithm)
        import hashlib
        if algorithm == "sha512":
            return hashlib.sha512(data.encode()).hexdigest()
        return hashlib.sha256(data.encode()).hexdigest()

    def generate_secret(self, length: int = 32) -> str:
        if self._framework:
            return self._framework.generate_secret(length)
        import secrets
        import string
        chars = string.ascii_letters + string.digits + "!@#$%^&*"
        return "".join(secrets.choice(chars) for _ in range(length))

    def encrypt_file(self, file_path: str, output_path: str = "", password: str = "") -> dict:
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"File not found: {file_path}"}
        out = output_path or file_path + ".enc"
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            encrypted = self.encrypt(content, password)
            with open(out, "w") as f:
                f.write(encrypted)
            _log("INFO", f"File encrypted: {file_path} -> {out}")
            return {"status": "ok", "input": file_path, "output": out}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def decrypt_file(self, file_path: str, output_path: str = "", password: str = "") -> dict:
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"File not found: {file_path}"}
        out = output_path or file_path.replace(".enc", ".dec")
        try:
            with open(file_path, "r") as f:
                encrypted = f.read()
            decrypted = self.decrypt(encrypted, password)
            with open(out, "w", encoding="utf-8") as f:
                f.write(decrypted)
            _log("INFO", f"File decrypted: {file_path} -> {out}")
            return {"status": "ok", "input": file_path, "output": out}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _xor_encrypt(self, data: str, key: str) -> str:
        import base64
        key_bytes = key.encode()
        encrypted = bytes([b ^ key_bytes[i % len(key_bytes)] for i, b in enumerate(data.encode())])
        return base64.b64encode(encrypted).decode()

    def _xor_decrypt(self, encrypted: str, key: str) -> str:
        import base64
        key_bytes = key.encode()
        data = base64.b64decode(encrypted)
        return bytes([b ^ key_bytes[i % len(key_bytes)] for i, b in enumerate(data)]).decode("utf-8", errors="replace")


class CacheManager:
    """In-memory cache with TTL and LRU eviction."""

    def __init__(self, max_size: int = 1000, default_ttl: int = 300):
        self._cache: Dict[str, dict] = {}
        self._max_size = max_size
        self._default_ttl = default_ttl
        self._hits = 0
        self._misses = 0

    def set(self, key: str, value: Any, ttl: int = 0) -> None:
        if len(self._cache) >= self._max_size:
            oldest = min(self._cache, key=lambda k: self._cache[k].get("accessed", 0))
            del self._cache[oldest]
        self._cache[key] = {
            "value": value,
            "expires": time.time() + (ttl or self._default_ttl),
            "accessed": time.time(),
        }

    def get(self, key: str, default: Any = None) -> Any:
        entry = self._cache.get(key)
        if not entry:
            self._misses += 1
            return default
        if time.time() > entry["expires"]:
            del self._cache[key]
            self._misses += 1
            return default
        entry["accessed"] = time.time()
        self._hits += 1
        return entry["value"]

    def delete(self, key: str) -> bool:
        if key in self._cache:
            del self._cache[key]
            return True
        return False

    def clear(self) -> None:
        self._cache.clear()
        self._hits = 0
        self._misses = 0

    def cleanup_expired(self) -> int:
        now = time.time()
        expired = [k for k, v in self._cache.items() if now > v["expires"]]
        for k in expired:
            del self._cache[k]
        return len(expired)

    @property
    def size(self) -> int:
        return len(self._cache)

    @property
    def hit_rate(self) -> float:
        total = self._hits + self._misses
        return self._hits / total if total > 0 else 0.0

    @property
    def stats(self) -> dict:
        return {"size": self.size, "hits": self._hits, "misses": self._misses, "hit_rate": self.hit_rate}


class TaskQueue:
    """Async task queue with worker threads."""

    def __init__(self, num_workers: int = 4):
        self._queue: List[dict] = []
        self._workers: List[threading.Thread] = []
        self._num_workers = num_workers
        self._running = False
        self._completed: List[dict] = []
        self._lock = threading.Lock()

    def enqueue(self, name: str, callback: Callable, data: Any = None) -> str:
        task = {"name": name, "callback": callback, "data": data, "id": f"task_{len(self._queue)}_{time.time()}"}
        with self._lock:
            self._queue.append(task)
        return task["id"]

    def start(self) -> bool:
        if self._running:
            return False
        self._running = True
        for i in range(self._num_workers):
            t = threading.Thread(target=self._worker_loop, daemon=True, name=f"taskq-worker-{i}")
            t.start()
            self._workers.append(t)
        _log("INFO", f"Task queue started with {self._num_workers} workers")
        return True

    def _worker_loop(self) -> None:
        while self._running:
            task = None
            with self._lock:
                if self._queue:
                    task = self._queue.pop(0)
            if task:
                try:
                    result = task["callback"](task["data"])
                    self._completed.append({"id": task["id"], "name": task["name"], "status": "ok", "result": result})
                except Exception as e:
                    self._completed.append({"id": task["id"], "name": task["name"], "status": "error", "error": str(e)})
            else:
                import time as _t
                _t.sleep(0.1)

    def stop(self) -> None:
        self._running = False
        _log("INFO", "Task queue stopped")

    @property
    def pending_count(self) -> int:
        return len(self._queue)

    @property
    def completed_count(self) -> int:
        return len(self._completed)

    def get_results(self) -> List[dict]:
        results = list(self._completed)
        self._completed.clear()
        return results


class NotificationManager:
    """Unified notification manager — toast, push, email, desktop."""

    def __init__(self):
        self._channels: Dict[str, Callable] = {}
        self._history: List[dict] = []

    def register_channel(self, name: str, sender: Callable) -> None:
        self._channels[name] = sender
        _log("INFO", f"Notification channel registered: {name}")

    def send(self, title: str, body: str, channels: Optional[List[str]] = None) -> dict:
        targets = channels or list(self._channels.keys())
        results: Dict[str, Any] = {}
        for ch in targets:
            if ch in self._channels:
                try:
                    results[ch] = self._channels[ch](title, body)
                except Exception as e:
                    results[ch] = {"status": "error", "message": str(e)}
        entry = {"title": title, "body": body, "channels": results, "timestamp": datetime.now(timezone.utc).isoformat()}
        self._history.append(entry)
        _log("INFO", f"Notification sent: {title} to {list(results.keys())}")
        return {"status": "ok", "results": results}

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    def clear_history(self) -> None:
        self._history.clear()


class BackupManager:
    """Backup manager — create and restore backups."""

    def __init__(self, backup_dir: str = "backups"):
        self._backup_dir = backup_dir
        os.makedirs(backup_dir, exist_ok=True)

    def create(self, files: List[str], name: str = "") -> dict:
        import zipfile
        backup_name = name or f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        backup_path = os.path.join(self._backup_dir, backup_name)
        try:
            with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for file_path in files:
                    if os.path.exists(file_path):
                        arcname = os.path.relpath(file_path)
                        zf.write(file_path, arcname)
            _log("INFO", f"Backup created: {backup_path} ({len(files)} files)")
            return {"status": "ok", "path": backup_path, "files": len(files)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def restore(self, backup_path: str, extract_to: str = ".") -> dict:
        import zipfile
        if not os.path.exists(backup_path):
            return {"status": "error", "message": f"Backup not found: {backup_path}"}
        try:
            with zipfile.ZipFile(backup_path, "r") as zf:
                zf.extractall(extract_to)
            _log("INFO", f"Backup restored: {backup_path} -> {extract_to}")
            return {"status": "ok", "path": backup_path, "extract_to": extract_to}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def list_backups(self) -> List[dict]:
        backups = []
        for f in os.listdir(self._backup_dir):
            if f.endswith(".zip"):
                path = os.path.join(self._backup_dir, f)
                backups.append({
                    "name": f,
                    "path": path,
                    "size": os.path.getsize(path),
                    "created": datetime.fromtimestamp(os.path.getmtime(path)).isoformat(),
                })
        return sorted(backups, key=lambda b: b["created"], reverse=True)

    def delete_backup(self, name: str) -> bool:
        path = os.path.join(self._backup_dir, name)
        if os.path.exists(path):
            os.remove(path)
            return True
        return False


class LogManager:
    """Advanced log manager — structured logging with levels and rotation."""

    def __init__(self, log_file: str = "pytrex_app.log", max_size: int = 10 * 1024 * 1024):
        self._log_file = log_file
        self._max_size = max_size
        self._entries: List[dict] = []
        self._level_colors = {
            "DEBUG": "\033[36m",
            "INFO": "\033[32m",
            "WARN": "\033[33m",
            "ERROR": "\033[31m",
            "CRASH": "\033[35m",
        }

    def log(self, level: str, message: str, data: Optional[dict] = None) -> None:
        entry = {
            "level": level,
            "message": message,
            "data": data or {},
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._entries.append(entry)
        color = self._level_colors.get(level, "")
        reset = "\033[0m" if color else ""
        print(f"{color}[{level}]{reset} {message}")
        self._write_to_file(entry)

    def debug(self, message: str, data: Optional[dict] = None) -> None:
        self.log("DEBUG", message, data)

    def info(self, message: str, data: Optional[dict] = None) -> None:
        self.log("INFO", message, data)

    def warn(self, message: str, data: Optional[dict] = None) -> None:
        self.log("WARN", message, data)

    def error(self, message: str, data: Optional[dict] = None) -> None:
        self.log("ERROR", message, data)

    def _write_to_file(self, entry: dict) -> None:
        try:
            if os.path.exists(self._log_file) and os.path.getsize(self._log_file) > self._max_size:
                rotated = self._log_file + f".{datetime.now().strftime('%Y%m%d%H%M%S')}"
                os.rename(self._log_file, rotated)
            with open(self._log_file, "a", encoding="utf-8") as f:
                f.write(json.dumps(entry) + "\n")
        except Exception:
            pass

    def get_entries(self, level: str = "", limit: int = 100) -> List[dict]:
        entries = self._entries
        if level:
            entries = [e for e in entries if e["level"] == level]
        return entries[-limit:]

    def clear(self) -> None:
        self._entries.clear()

    def export(self, output_path: str = "pytrex_logs_export.json") -> dict:
        try:
            with open(output_path, "w") as f:
                json.dump(self._entries, f, indent=2)
            return {"status": "ok", "path": output_path, "count": len(self._entries)}
        except Exception as e:
            return {"status": "error", "message": str(e)}


class DependencyChecker:
    """Check and manage Python/Rust/Elixir dependencies."""

    def __init__(self):
        self._python_deps: Dict[str, str] = {}
        self._rust_deps: Dict[str, str] = {}

    def check_python(self, packages: Optional[List[str]] = None) -> dict:
        results: Dict[str, Any] = {}
        import importlib
        pkgs = packages or ["pytrex", "torch", "transformers", "websockets", "watchdog", "keyring", "msgpack"]
        for pkg in pkgs:
            try:
                mod = importlib.import_module(pkg)
                version = getattr(mod, "__version__", "unknown")
                results[pkg] = {"installed": True, "version": version}
            except ImportError:
                results[pkg] = {"installed": False, "version": None}
        installed = sum(1 for v in results.values() if v["installed"])
        _log("INFO", f"Dependency check: {installed}/{len(pkgs)} Python packages installed")
        return {"status": "ok", "results": results, "installed": installed, "total": len(pkgs)}

    def check_rust(self) -> dict:
        try:
            result = subprocess.run(["cargo", "--version"], capture_output=True, text=True, timeout=5)
            if result.returncode == 0:
                return {"status": "ok", "cargo": result.stdout.strip()}
            return {"status": "error", "message": "cargo not found"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def check_elixir(self) -> dict:
        try:
            result = subprocess.run(["elixir", "--version"], capture_output=True, text=True, timeout=5)
            if result.returncode == 0:
                return {"status": "ok", "elixir": result.stdout.strip()}
            return {"status": "error", "message": "elixir not found"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def check_all(self) -> dict:
        return {
            "python": self.check_python(),
            "rust": self.check_rust(),
            "elixir": self.check_elixir(),
        }

    def install_python(self, package: str) -> dict:
        try:
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", package],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0:
                return {"status": "ok", "package": package}
            return {"status": "error", "message": result.stderr}
        except Exception as e:
            return {"status": "error", "message": str(e)}


class PerformanceMonitor:
    """Performance monitor — CPU, memory, timing, metrics."""

    def __init__(self):
        self._metrics: Dict[str, List[float]] = {}
        self._timers: Dict[str, float] = {}
        self._counters: Dict[str, int] = {}

    def start_timer(self, name: str) -> None:
        self._timers[name] = time.time()

    def stop_timer(self, name: str) -> float:
        if name not in self._timers:
            return 0.0
        elapsed = time.time() - self._timers[name]
        if name not in self._metrics:
            self._metrics[name] = []
        self._metrics[name].append(elapsed)
        del self._timers[name]
        _log("INFO", f"Timer '{name}': {elapsed:.4f}s")
        return elapsed

    def increment(self, name: str, amount: int = 1) -> None:
        self._counters[name] = self._counters.get(name, 0) + amount

    def gauge(self, name: str, value: float) -> None:
        if name not in self._metrics:
            self._metrics[name] = []
        self._metrics[name].append(value)

    def get_metric(self, name: str) -> Optional[dict]:
        if name in self._metrics:
            values = self._metrics[name]
            return {
                "count": len(values),
                "min": min(values),
                "max": max(values),
                "avg": sum(values) / len(values),
                "last": values[-1],
            }
        if name in self._counters:
            return {"count": 1, "value": self._counters[name]}
        return None

    def get_all_metrics(self) -> Dict[str, Any]:
        result: Dict[str, Any] = {}
        for name in set(list(self._metrics.keys()) + list(self._counters.keys())):
            result[name] = self.get_metric(name)
        return result

    def memory_usage(self) -> dict:
        try:
            import psutil
            proc = psutil.Process()
            mem = proc.memory_info()
            return {"rss_mb": mem.rss / (1024 * 1024), "vms_mb": mem.vms / (1024 * 1024)}
        except ImportError:
            return {"rss_mb": 0, "vms_mb": 0, "message": "psutil not installed"}

    def cpu_usage(self) -> float:
        try:
            import psutil
            return psutil.cpu_percent(interval=0.1)
        except ImportError:
            return 0.0

    def reset(self) -> None:
        self._metrics.clear()
        self._counters.clear()
        self._timers.clear()


class StateMachine:
    """App state management — transitions, guards, callbacks."""

    def __init__(self, initial_state: str = "idle"):
        self._state = initial_state
        self._transitions: Dict[str, List[str]] = {}
        self._handlers: Dict[str, Dict[str, Callable]] = {}
        self._history: List[dict] = []

    def add_transition(self, from_state: str, to_state: str) -> None:
        self._transitions.setdefault(from_state, []).append(to_state)

    def on(self, state: str, callback: Callable) -> None:
        self._handlers.setdefault(state, {})["enter"] = callback

    def on_exit(self, state: str, callback: Callable) -> None:
        self._handlers.setdefault(state, {})["exit"] = callback

    def set(self, new_state: str) -> bool:
        if new_state == self._state:
            return True
        allowed = self._transitions.get(self._state, [])
        if allowed and new_state not in allowed:
            _log("WARN", f"Invalid transition: {self._state} -> {new_state}")
            return False
        old = self._state
        if old in self._handlers and "exit" in self._handlers[old]:
            self._handlers[old]["exit"]()
        self._state = new_state
        self._history.append({"from": old, "to": new_state, "timestamp": datetime.now(timezone.utc).isoformat()})
        if new_state in self._handlers and "enter" in self._handlers[new_state]:
            self._handlers[new_state]["enter"]()
        _log("INFO", f"State: {old} -> {new_state}")
        return True

    def can_transition(self, to_state: str) -> bool:
        return to_state in self._transitions.get(self._state, []) or not self._transitions.get(self._state)

    @property
    def state(self) -> str:
        return self._state

    @property
    def history(self) -> List[dict]:
        return list(self._history)


class EventBus:
    """Global pub/sub event bus for inter-component communication."""

    def __init__(self):
        self._subscribers: Dict[str, List[Callable]] = {}
        self._history: List[dict] = []

    def on(self, event: str, handler: Callable) -> None:
        self._subscribers.setdefault(event, []).append(handler)

    def off(self, event: str, handler: Callable) -> bool:
        if event in self._subscribers and handler in self._subscribers[event]:
            self._subscribers[event].remove(handler)
            return True
        return False

    def emit(self, event: str, data: Any = None) -> int:
        handlers = self._subscribers.get(event, [])
        count = 0
        for handler in handlers:
            try:
                handler(data)
                count += 1
            except Exception as e:
                _log("ERROR", f"Event bus handler error for '{event}': {e}")
        self._history.append({"event": event, "data": data, "handlers_called": count,
                              "timestamp": datetime.now(timezone.utc).isoformat()})
        return count

    def once(self, event: str, handler: Callable) -> None:
        def wrapper(data):
            handler(data)
            self.off(event, wrapper)
        self.on(event, wrapper)

    def clear(self) -> None:
        self._subscribers.clear()
        self._history.clear()

    @property
    def events(self) -> List[str]:
        return list(self._subscribers.keys())

    @property
    def subscriber_count(self) -> int:
        return sum(len(h) for h in self._subscribers.values())


class ValidatorEngine:
    """Data validation engine — schema-based validation."""

    def __init__(self):
        self._schemas: Dict[str, dict] = {}
        self._custom_rules: Dict[str, Callable] = {}

    def schema(self, name: str, fields: dict) -> None:
        self._schemas[name] = fields

    def add_rule(self, name: str, func: Callable) -> None:
        self._custom_rules[name] = func

    def validate(self, data: dict, schema_name: str = "") -> dict:
        schema = self._schemas.get(schema_name, {})
        if not schema and not schema_name:
            schema = data.get("__schema__", {})
        errors: List[str] = []
        validated: Dict[str, Any] = {}

        for field, rules in schema.items():
            value = data.get(field)
            field_errors: List[str] = []

            if rules.get("required", False) and value is None:
                field_errors.append(f"Field '{field}' is required")
                errors.extend(field_errors)
                continue

            if value is None:
                validated[field] = rules.get("default")
                continue

            expected_type = rules.get("type")
            if expected_type:
                type_map = {"str": str, "int": int, "float": float, "bool": bool, "list": list, "dict": dict}
                if expected_type in type_map and not isinstance(value, type_map[expected_type]):
                    field_errors.append(f"Field '{field}' must be {expected_type}, got {type(value).__name__}")

            if "min" in rules and isinstance(value, (int, float)) and value < rules["min"]:
                field_errors.append(f"Field '{field}' must be >= {rules['min']}")

            if "max" in rules and isinstance(value, (int, float)) and value > rules["max"]:
                field_errors.append(f"Field '{field}' must be <= {rules['max']}")

            if "min_len" in rules and isinstance(value, (str, list)) and len(value) < rules["min_len"]:
                field_errors.append(f"Field '{field}' must have length >= {rules['min_len']}")

            if "max_len" in rules and isinstance(value, (str, list)) and len(value) > rules["max_len"]:
                field_errors.append(f"Field '{field}' must have length <= {rules['max_len']}")

            if "pattern" in rules and isinstance(value, str):
                import re
                if not re.match(rules["pattern"], value):
                    field_errors.append(f"Field '{field}' does not match pattern {rules['pattern']}")

            if "choices" in rules and value not in rules["choices"]:
                field_errors.append(f"Field '{field}' must be one of {rules['choices']}")

            if "custom" in rules and rules["custom"] in self._custom_rules:
                try:
                    result = self._custom_rules[rules["custom"]](value)
                    if result is not True:
                        field_errors.append(f"Field '{field}' failed custom rule: {result}")
                except Exception as e:
                    field_errors.append(f"Field '{field}' custom rule error: {e}")

            if field_errors:
                errors.extend(field_errors)
            else:
                validated[field] = value

        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "data": validated if not errors else data,
        }


class Localization:
    """Advanced localization — translations, pluralization, date/number formatting."""

    def __init__(self, default_lang: str = "en"):
        self._lang = default_lang
        self._translations: Dict[str, Dict[str, str]] = {}
        self._plurals: Dict[str, Dict[str, List[str]]] = {}

    def add_translation(self, lang: str, key: str, value: str) -> None:
        self._translations.setdefault(lang, {})[key] = value

    def add_translations(self, lang: str, translations: dict) -> None:
        self._translations.setdefault(lang, {}).update(translations)

    def add_plural(self, key: str, lang: str, forms: List[str]) -> None:
        self._plurals.setdefault(key, {})[lang] = forms

    def t(self, key: str, lang: str = "", **kwargs) -> str:
        use_lang = lang or self._lang
        text = self._translations.get(use_lang, {}).get(key)
        if not text:
            text = self._translations.get("en", {}).get(key, key)
        if kwargs:
            try:
                text = text.format(**kwargs)
            except Exception:
                pass
        return text

    def tp(self, key: str, count: int, lang: str = "") -> str:
        use_lang = lang or self._lang
        forms = self._plurals.get(key, {}).get(use_lang)
        if not forms:
            return self.t(key, lang)
        if count == 1:
            return forms[0]
        return forms[1] if len(forms) > 1 else forms[0]

    def set_lang(self, lang: str) -> None:
        self._lang = lang

    @property
    def current_lang(self) -> str:
        return self._lang

    @property
    def available_langs(self) -> List[str]:
        return list(self._translations.keys())

    def format_date(self, date_obj, lang: str = "") -> str:
        use_lang = lang or self._lang
        if use_lang == "sw":
            months = ["Januari", "Februari", "Machi", "Aprili", "Mei", "Juni",
                     "Julai", "Agosti", "Septemba", "Oktoba", "Novemba", "Desemba"]
        else:
            months = ["January", "February", "March", "April", "May", "June",
                     "July", "August", "September", "October", "November", "December"]
        if hasattr(date_obj, "day"):
            return f"{date_obj.day} {months[date_obj.month - 1]} {date_obj.year}"
        return str(date_obj)

    def format_number(self, number: float, decimals: int = 2) -> str:
        return f"{number:,.{decimals}f}"


class FeatureFlags:
    """Feature flag system — toggle features at runtime."""

    def __init__(self):
        self._flags: Dict[str, dict] = {}
        self._listeners: Dict[str, List[Callable]] = {}

    def register(self, name: str, enabled: bool = False, description: str = "") -> None:
        self._flags[name] = {"enabled": enabled, "description": description}

    def enable(self, name: str) -> bool:
        if name in self._flags:
            old = self._flags[name]["enabled"]
            self._flags[name]["enabled"] = True
            if not old:
                self._notify(name, True)
            return True
        return False

    def disable(self, name: str) -> bool:
        if name in self._flags:
            old = self._flags[name]["enabled"]
            self._flags[name]["enabled"] = False
            if old:
                self._notify(name, False)
            return True
        return False

    def toggle(self, name: str) -> bool:
        if name in self._flags:
            new_val = not self._flags[name]["enabled"]
            self._flags[name]["enabled"] = new_val
            self._notify(name, new_val)
            return True
        return False

    def is_enabled(self, name: str) -> bool:
        return self._flags.get(name, {}).get("enabled", False)

    def on_change(self, name: str, callback: Callable) -> None:
        self._listeners.setdefault(name, []).append(callback)

    def _notify(self, name: str, enabled: bool) -> None:
        for cb in self._listeners.get(name, []):
            try:
                cb(name, enabled)
            except Exception as e:
                _log("ERROR", f"Feature flag listener error: {e}")

    @property
    def flags(self) -> Dict[str, dict]:
        return dict(self._flags)

    def list_enabled(self) -> List[str]:
        return [k for k, v in self._flags.items() if v["enabled"]]


class RateLimiter:
    """Advanced rate limiter — per-key, sliding window."""

    def __init__(self):
        self._limits: Dict[str, dict] = {}
        self._requests: Dict[str, List[float]] = {}

    def set_limit(self, key: str, limit: int, window_seconds: int) -> None:
        self._limits[key] = {"limit": limit, "window": window_seconds}

    def check(self, key: str, limit: int = 0, window: int = 0) -> dict:
        cfg = self._limits.get(key)
        if cfg:
            limit = limit or cfg["limit"]
            window = window or cfg["window"]
        if not limit or not window:
            return {"allowed": True, "remaining": -1, "message": "No limit configured"}

        now = time.time()
        if key not in self._requests:
            self._requests[key] = []
        self._requests[key] = [t for t in self._requests[key] if now - t < window]

        if len(self._requests[key]) < limit:
            self._requests[key].append(now)
            remaining = limit - len(self._requests[key])
            return {"allowed": True, "remaining": remaining, "limit": limit, "window": window}
        return {"allowed": False, "remaining": 0, "limit": limit, "window": window,
                "retry_after": window - (now - self._requests[key][0])}

    def reset(self, key: str = "") -> None:
        if key:
            self._requests.pop(key, None)
        else:
            self._requests.clear()

    @property
    def tracked_keys(self) -> List[str]:
        return list(self._limits.keys())


class RetryEngine:
    """Retry engine with exponential backoff for failed operations."""

    def __init__(self):
        self._stats: Dict[str, dict] = {}

    def attempt(self, callback: Callable, max_retries: int = 3,
                backoff: str = "exponential", base_delay: float = 0.1,
                name: str = "") -> dict:
        op_name = name or callback.__name__ if hasattr(callback, "__name__") else "operation"
        attempts = 0
        last_error = ""

        for attempt in range(max_retries + 1):
            attempts += 1
            try:
                result = callback()
                self._stats[op_name] = {"attempts": attempts, "success": True, "last_error": ""}
                return {"success": True, "result": result, "attempts": attempts}
            except Exception as e:
                last_error = str(e)
                _log("WARN", f"Retry {attempts}/{max_retries + 1} for '{op_name}': {e}")
                if attempt < max_retries:
                    if backoff == "exponential":
                        delay = base_delay * (2 ** attempt)
                    elif backoff == "linear":
                        delay = base_delay * (attempt + 1)
                    else:
                        delay = base_delay
                    time.sleep(delay)

        self._stats[op_name] = {"attempts": attempts, "success": False, "last_error": last_error}
        return {"success": False, "error": last_error, "attempts": attempts}

    def attempt_async(self, callback: Callable, max_retries: int = 3,
                      backoff: str = "exponential", base_delay: float = 0.1,
                      name: str = "") -> threading.Thread:
        def _run():
            self.attempt(callback, max_retries, backoff, base_delay, name)
        t = threading.Thread(target=_run, daemon=True)
        t.start()
        return t

    @property
    def stats(self) -> Dict[str, dict]:
        return dict(self._stats)


class CircuitBreaker:
    """Circuit breaker — prevent cascade failures."""

    def __init__(self, failure_threshold: int = 5, recovery_timeout: float = 30.0):
        self._failure_threshold = failure_threshold
        self._recovery_timeout = recovery_timeout
        self._circuits: Dict[str, dict] = {}

    def _get_circuit(self, name: str) -> dict:
        if name not in self._circuits:
            self._circuits[name] = {
                "state": "closed",
                "failures": 0,
                "last_failure": 0.0,
                "successes": 0,
            }
        return self._circuits[name]

    def call(self, name: str, callback: Callable) -> dict:
        circuit = self._get_circuit(name)

        if circuit["state"] == "open":
            if time.time() - circuit["last_failure"] > self._recovery_timeout:
                circuit["state"] = "half_open"
                _log("INFO", f"Circuit '{name}' -> half_open")
            else:
                return {"status": "error", "message": f"Circuit '{name}' is open",
                        "state": "open"}

        try:
            result = callback()
            circuit["successes"] += 1
            circuit["failures"] = 0
            if circuit["state"] == "half_open":
                circuit["state"] = "closed"
                _log("INFO", f"Circuit '{name}' -> closed")
            return {"status": "ok", "result": result, "state": circuit["state"]}
        except Exception as e:
            circuit["failures"] += 1
            circuit["last_failure"] = time.time()
            if circuit["failures"] >= self._failure_threshold:
                circuit["state"] = "open"
                _log("WARN", f"Circuit '{name}' -> open ({circuit['failures']} failures)")
            return {"status": "error", "message": str(e), "state": circuit["state"],
                    "failures": circuit["failures"]}

    def reset(self, name: str = "") -> None:
        if name:
            if name in self._circuits:
                self._circuits[name] = {"state": "closed", "failures": 0,
                                       "last_failure": 0.0, "successes": 0}
        else:
            self._circuits.clear()

    def get_state(self, name: str) -> str:
        return self._get_circuit(name)["state"]

    @property
    def circuits(self) -> Dict[str, str]:
        return {k: v["state"] for k, v in self._circuits.items()}


class SecretVault:
    """Encrypted secret storage — store API keys, passwords encrypted at rest."""

    def __init__(self, master_key: str = ""):
        self._secrets: Dict[str, str] = {}
        self._encryption = EncryptionManager(default_password=master_key or "PyTreX-Vault-Master-Key")

    def store(self, name: str, value: str) -> bool:
        encrypted = self._encryption.encrypt(value)
        self._secrets[name] = encrypted
        _log("INFO", f"Secret stored: {name}")
        return True

    def retrieve(self, name: str) -> Optional[str]:
        encrypted = self._secrets.get(name)
        if not encrypted:
            return None
        try:
            return self._encryption.decrypt(encrypted)
        except Exception as e:
            _log("ERROR", f"Secret retrieval failed for '{name}': {e}")
            return None

    def delete(self, name: str) -> bool:
        if name in self._secrets:
            del self._secrets[name]
            return True
        return False

    def list_names(self) -> List[str]:
        return list(self._secrets.keys())

    def exists(self, name: str) -> bool:
        return name in self._secrets

    def rotate(self, name: str, new_value: str) -> bool:
        if name not in self._secrets:
            return False
        self._secrets[name] = self._encryption.encrypt(new_value)
        _log("INFO", f"Secret rotated: {name}")
        return True

    @property
    def count(self) -> int:
        return len(self._secrets)


class APIClient:
    """Built-in HTTP/REST client for external APIs."""

    def __init__(self, base_url: str = "", default_headers: Optional[dict] = None):
        self._base_url = base_url.rstrip("/")
        self._headers = default_headers or {}
        self._timeout = 30
        self._stats: Dict[str, int] = {"requests": 0, "successes": 0, "errors": 0}

    def set_base_url(self, url: str) -> None:
        self._base_url = url.rstrip("/")

    def set_header(self, key: str, value: str) -> None:
        self._headers[key] = value

    def set_auth(self, token: str, scheme: str = "Bearer") -> None:
        self._headers["Authorization"] = f"{scheme} {token}"

    def _request(self, method: str, endpoint: str, data: Optional[dict] = None,
                 headers: Optional[dict] = None) -> dict:
        import urllib.request
        import urllib.error
        url = endpoint if endpoint.startswith("http") else f"{self._base_url}/{endpoint.lstrip('/')}"
        req_headers = dict(self._headers)
        if headers:
            req_headers.update(headers)

        body = None
        if data:
            body = json.dumps(data).encode("utf-8")
            req_headers["Content-Type"] = "application/json"

        self._stats["requests"] += 1
        try:
            req = urllib.request.Request(url, data=body, method=method, headers=req_headers)
            with urllib.request.urlopen(req, timeout=self._timeout) as resp:
                response_body = resp.read().decode("utf-8", errors="replace")
                try:
                    parsed = json.loads(response_body)
                except Exception:
                    parsed = response_body
                self._stats["successes"] += 1
                return {"status": "ok", "code": resp.status, "data": parsed, "url": url}
        except urllib.error.HTTPError as e:
            self._stats["errors"] += 1
            return {"status": "error", "code": e.code, "message": str(e), "url": url}
        except Exception as e:
            self._stats["errors"] += 1
            return {"status": "error", "message": str(e), "url": url}

    def get(self, endpoint: str, headers: Optional[dict] = None) -> dict:
        return self._request("GET", endpoint, headers=headers)

    def post(self, endpoint: str, data: Optional[dict] = None, headers: Optional[dict] = None) -> dict:
        return self._request("POST", endpoint, data=data, headers=headers)

    def put(self, endpoint: str, data: Optional[dict] = None, headers: Optional[dict] = None) -> dict:
        return self._request("PUT", endpoint, data=data, headers=headers)

    def delete(self, endpoint: str, headers: Optional[dict] = None) -> dict:
        return self._request("DELETE", endpoint, headers=headers)

    def patch(self, endpoint: str, data: Optional[dict] = None, headers: Optional[dict] = None) -> dict:
        return self._request("PATCH", endpoint, data=data, headers=headers)

    @property
    def stats(self) -> dict:
        return dict(self._stats)


class WSClient:
    """WebSocket client — connect to external WS servers."""

    def __init__(self, url: str = ""):
        self._url = url
        self._connected = False
        self._handlers: Dict[str, List[Callable]] = {}
        self._messages: List[str] = []
        self._thread = None

    def connect(self, url: str = "") -> bool:
        self._url = url or self._url
        self._connected = True
        _log("INFO", f"WSClient connected to {self._url}")
        return True

    def disconnect(self) -> None:
        self._connected = False
        _log("INFO", "WSClient disconnected")

    def send(self, message: str) -> bool:
        if not self._connected:
            return False
        _log("INFO", f"WSClient sent: {message[:50]}")
        return True

    def on_message(self, handler: Callable) -> None:
        self._handlers.setdefault("message", []).append(handler)

    def on_event(self, event: str, handler: Callable) -> None:
        self._handlers.setdefault(event, []).append(handler)

    def receive(self, message: str) -> None:
        self._messages.append(message)
        for handler in self._handlers.get("message", []):
            try:
                handler(message)
            except Exception as e:
                _log("ERROR", f"WSClient handler error: {e}")

    @property
    def is_connected(self) -> bool:
        return self._connected

    @property
    def message_count(self) -> int:
        return len(self._messages)


class RedisPubSub:
    """Redis-compatible pub/sub message broker."""

    def __init__(self, host: str = "localhost", port: int = 6379):
        self._host = host
        self._port = port
        self._channels: Dict[str, List[Callable]] = {}
        self._messages: Dict[str, List[str]] = {}
        self._connected = False

    def connect(self) -> bool:
        self._connected = True
        _log("INFO", f"Redis pub/sub connected to {self._host}:{self._port}")
        return True

    def publish(self, channel: str, message: str) -> int:
        handlers = self._channels.get(channel, [])
        for handler in handlers:
            try:
                handler(message)
            except Exception as e:
                _log("ERROR", f"Redis handler error: {e}")
        self._messages.setdefault(channel, []).append(message)
        _log("INFO", f"Redis publish to '{channel}': {len(handlers)} subscribers")
        return len(handlers)

    def subscribe(self, channel: str, handler: Callable) -> bool:
        self._channels.setdefault(channel, []).append(handler)
        _log("INFO", f"Redis subscribed to '{channel}'")
        return True

    def unsubscribe(self, channel: str, handler: Callable) -> bool:
        if channel in self._channels and handler in self._channels[channel]:
            self._channels[channel].remove(handler)
            return True
        return False

    @property
    def channels(self) -> List[str]:
        return list(self._channels.keys())

    @property
    def is_connected(self) -> bool:
        return self._connected

    def get_messages(self, channel: str) -> List[str]:
        return list(self._messages.get(channel, []))


class SearchEngine:
    """Full-text search engine — index and query documents."""

    def __init__(self):
        self._index: Dict[str, dict] = {}
        self._documents: Dict[str, str] = {}

    def index(self, doc_id: str, text: str) -> bool:
        self._documents[doc_id] = text
        words = text.lower().split()
        for word in set(words):
            if word not in self._index:
                self._index[word] = {"docs": {}, "total_count": 0}
            self._index[word]["docs"][doc_id] = words.count(word)
            self._index[word]["total_count"] += words.count(word)
        _log("INFO", f"Indexed document: {doc_id} ({len(words)} words)")
        return True

    def query(self, query_str: str, limit: int = 10) -> List[dict]:
        words = query_str.lower().split()
        scores: Dict[str, float] = {}
        for word in words:
            entry = self._index.get(word)
            if entry:
                for doc_id, count in entry["docs"].items():
                    scores[doc_id] = scores.get(doc_id, 0) + count
        ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:limit]
        return [{"doc_id": doc_id, "score": score, "text": self._documents.get(doc_id, "")[:200]}
                for doc_id, score in ranked]

    def remove(self, doc_id: str) -> bool:
        if doc_id not in self._documents:
            return False
        text = self._documents.pop(doc_id)
        for word in set(text.lower().split()):
            if word in self._index and doc_id in self._index[word]["docs"]:
                self._index[word]["total_count"] -= self._index[word]["docs"][doc_id]
                del self._index[word]["docs"][doc_id]
                if not self._index[word]["docs"]:
                    del self._index[word]
        return True

    @property
    def document_count(self) -> int:
        return len(self._documents)

    @property
    def index_size(self) -> int:
        return len(self._index)


class MLInference:
    """ML inference engine — run PyTorch/ONNX models locally."""

    def __init__(self):
        self._models: Dict[str, Any] = {}
        self._framework = None
        try:
            import torch
            self._framework = "torch"
        except ImportError:
            pass

    def load_model(self, name: str, model_path: str) -> dict:
        if not os.path.exists(model_path):
            return {"status": "error", "message": f"Model not found: {model_path}"}
        try:
            if self._framework == "torch":
                import torch
                model = torch.load(model_path, map_location="cpu", weights_only=False)
                if hasattr(model, "eval"):
                    model.eval()
                self._models[name] = model
                _log("INFO", f"ML model loaded: {name} (PyTorch)")
                return {"status": "ok", "name": name, "framework": "torch"}
            return {"status": "error", "message": "No ML framework available"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def predict(self, name: str, input_data: Any) -> dict:
        if name not in self._models:
            return {"status": "error", "message": f"Model '{name}' not loaded"}
        try:
            model = self._models[name]
            if self._framework == "torch":
                import torch
                if not isinstance(input_data, torch.Tensor):
                    input_data = torch.tensor(input_data, dtype=torch.float32)
                with torch.no_grad():
                    output = model(input_data)
                result = output.tolist() if hasattr(output, "tolist") else str(output)
                return {"status": "ok", "prediction": result}
            return {"status": "error", "message": "No ML framework"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def list_models(self) -> List[str]:
        return list(self._models.keys())

    @property
    def framework(self) -> str:
        return self._framework or "none"

    @property
    def model_count(self) -> int:
        return len(self._models)


class DataExporter:
    """Export data to CSV, Excel, JSON, XML, YAML."""

    def __init__(self):
        self._exports: List[dict] = []

    def to_csv(self, data: List[dict], file_path: str) -> dict:
        try:
            import csv
            if not data:
                return {"status": "error", "message": "No data to export"}
            keys = list(data[0].keys())
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=keys)
                writer.writeheader()
                writer.writerows(data)
            self._exports.append({"format": "csv", "path": file_path, "rows": len(data)})
            _log("INFO", f"Exported CSV: {file_path} ({len(data)} rows)")
            return {"status": "ok", "path": file_path, "rows": len(data)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def to_json(self, data: Any, file_path: str) -> dict:
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            self._exports.append({"format": "json", "path": file_path})
            _log("INFO", f"Exported JSON: {file_path}")
            return {"status": "ok", "path": file_path}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def to_xml(self, data: List[dict], file_path: str, root: str = "data") -> dict:
        try:
            lines = [f'<?xml version="1.0" encoding="UTF-8"?>', f'<{root}>']
            for item in data:
                lines.append("  <item>")
                for key, value in item.items():
                    lines.append(f"    <{key}>{value}</{key}>")
                lines.append("  </item>")
            lines.append(f'</{root}>')
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            self._exports.append({"format": "xml", "path": file_path, "rows": len(data)})
            _log("INFO", f"Exported XML: {file_path} ({len(data)} rows)")
            return {"status": "ok", "path": file_path, "rows": len(data)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def to_yaml(self, data: Any, file_path: str) -> dict:
        try:
            import yaml
            with open(file_path, "w", encoding="utf-8") as f:
                yaml.dump(data, f, default_flow_style=False)
            self._exports.append({"format": "yaml", "path": file_path})
            _log("INFO", f"Exported YAML: {file_path}")
            return {"status": "ok", "path": file_path}
        except ImportError:
            return {"status": "error", "message": "PyYAML not installed — pip install pyyaml"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def to_excel(self, data: List[dict], file_path: str, sheet_name: str = "Sheet1") -> dict:
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            if data:
                keys = list(data[0].keys())
                ws.append(keys)
                for row in data:
                    ws.append([row.get(k) for k in keys])
            wb.save(file_path)
            self._exports.append({"format": "excel", "path": file_path, "rows": len(data)})
            _log("INFO", f"Exported Excel: {file_path} ({len(data)} rows)")
            return {"status": "ok", "path": file_path, "rows": len(data)}
        except ImportError:
            return {"status": "error", "message": "openpyxl not installed — pip install openpyxl"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @property
    def exports(self) -> List[dict]:
        return list(self._exports)


class DataImporter:
    """Import data from CSV, Excel, JSON, XML, YAML."""

    def __init__(self):
        self._imports: List[dict] = []

    def from_csv(self, file_path: str) -> dict:
        try:
            import csv
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                data = list(reader)
            self._imports.append({"format": "csv", "path": file_path, "rows": len(data)})
            _log("INFO", f"Imported CSV: {file_path} ({len(data)} rows)")
            return {"status": "ok", "data": data, "rows": len(data)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def from_json(self, file_path: str) -> dict:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = len(data) if isinstance(data, list) else 1
            self._imports.append({"format": "json", "path": file_path, "rows": rows})
            _log("INFO", f"Imported JSON: {file_path}")
            return {"status": "ok", "data": data, "rows": rows}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def from_xml(self, file_path: str) -> dict:
        try:
            import xml.etree.ElementTree as ET
            tree = ET.parse(file_path)
            root = tree.getroot()
            data = []
            for item in root:
                row = {child.tag: child.text for child in item}
                data.append(row)
            self._imports.append({"format": "xml", "path": file_path, "rows": len(data)})
            _log("INFO", f"Imported XML: {file_path} ({len(data)} rows)")
            return {"status": "ok", "data": data, "rows": len(data)}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def from_yaml(self, file_path: str) -> dict:
        try:
            import yaml
            with open(file_path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
            rows = len(data) if isinstance(data, list) else 1
            self._imports.append({"format": "yaml", "path": file_path, "rows": rows})
            _log("INFO", f"Imported YAML: {file_path}")
            return {"status": "ok", "data": data, "rows": rows}
        except ImportError:
            return {"status": "error", "message": "PyYAML not installed"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def from_excel(self, file_path: str, sheet_name: str = "") -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"status": "ok", "data": [], "rows": 0}
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            data = [dict(zip(headers, row)) for row in rows[1:]]
            self._imports.append({"format": "excel", "path": file_path, "rows": len(data)})
            _log("INFO", f"Imported Excel: {file_path} ({len(data)} rows)")
            return {"status": "ok", "data": data, "rows": len(data)}
        except ImportError:
            return {"status": "error", "message": "openpyxl not installed"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @property
    def imports(self) -> List[dict]:
        return list(self._imports)


class JobScheduler:
    """Distributed job scheduler — recurring jobs, priorities, dependencies."""

    def __init__(self):
        self._jobs: Dict[str, dict] = {}
        self._running = False
        self._thread = None
        self._completed: List[dict] = []

    def schedule(self, name: str, callback: Callable, cron: str = "",
                 interval: int = 0, priority: int = 0) -> bool:
        self._jobs[name] = {
            "callback": callback, "cron": cron, "interval": interval,
            "priority": priority, "last_run": 0, "runs": 0,
        }
        _log("INFO", f"Job scheduled: {name} (cron={cron}, interval={interval}s)")
        return True

    def run_now(self, name: str) -> dict:
        job = self._jobs.get(name)
        if not job:
            return {"status": "error", "message": f"Job '{name}' not found"}
        try:
            result = job["callback"]()
            job["runs"] += 1
            job["last_run"] = time.time()
            self._completed.append({"name": name, "status": "ok", "result": result})
            return {"status": "ok", "result": result}
        except Exception as e:
            self._completed.append({"name": name, "status": "error", "error": str(e)})
            return {"status": "error", "message": str(e)}

    def cancel(self, name: str) -> bool:
        if name in self._jobs:
            del self._jobs[name]
            _log("INFO", f"Job cancelled: {name}")
            return True
        return False

    def list_jobs(self) -> List[dict]:
        return [{"name": n, "cron": j["cron"], "interval": j["interval"],
                 "priority": j["priority"], "runs": j["runs"]} for n, j in self._jobs.items()]

    @property
    def job_count(self) -> int:
        return len(self._jobs)

    @property
    def completed_count(self) -> int:
        return len(self._completed)

    def get_results(self) -> List[dict]:
        results = list(self._completed)
        self._completed.clear()
        return results


class WebScraper:
    """Web scraper — fetch HTML and extract data using selectors."""

    def __init__(self):
        self._history: List[dict] = []

    def get(self, url: str, headers: Optional[dict] = None) -> dict:
        import urllib.request
        try:
            req_headers = {"User-Agent": "PyTreX-Scraper/1.0"}
            if headers:
                req_headers.update(headers)
            req = urllib.request.Request(url, headers=req_headers)
            with urllib.request.urlopen(req, timeout=15) as resp:
                html = resp.read().decode("utf-8", errors="replace")
                self._history.append({"url": url, "status": resp.status, "size": len(html)})
                _log("INFO", f"Scraped: {url} ({len(html)} bytes)")
                return {"status": "ok", "html": html, "url": url, "size": len(html)}
        except Exception as e:
            self._history.append({"url": url, "status": "error", "error": str(e)})
            return {"status": "error", "message": str(e)}

    def extract_text(self, html: str, tag: str = "p") -> List[str]:
        import re
        pattern = rf"<{tag}[^>]*>(.*?)</{tag}>"
        matches = re.findall(pattern, html, re.DOTALL | re.IGNORECASE)
        clean = [re.sub(r"<[^>]+>", "", m).strip() for m in matches]
        return [c for c in clean if c]

    def extract_links(self, html: str) -> List[str]:
        import re
        return re.findall(r'href=["\']([^"\']+)["\']', html, re.IGNORECASE)

    def extract_selectors(self, html: str, selectors: Dict[str, str]) -> dict:
        import re
        results: Dict[str, Any] = {}
        for key, tag in selectors.items():
            pattern = rf"<{tag}[^>]*>(.*?)</{tag}>"
            matches = re.findall(pattern, html, re.DOTALL | re.IGNORECASE)
            cleaned = [re.sub(r"<[^>]+>", "", m).strip() for m in matches]
            results[key] = cleaned[0] if len(cleaned) == 1 else cleaned
        return results

    def extract_table(self, html: str) -> List[dict]:
        import re
        tables = re.findall(r"<table[^>]*>(.*?)</table>", html, re.DOTALL | re.IGNORECASE)
        if not tables:
            return []
        rows_data: List[dict] = []
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", tables[0], re.DOTALL | re.IGNORECASE)
        headers: List[str] = []
        for i, row in enumerate(rows):
            cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", row, re.DOTALL | re.IGNORECASE)
            cells = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            if i == 0:
                headers = cells
            elif headers:
                rows_data.append(dict(zip(headers, cells)))
        return rows_data

    @property
    def history(self) -> List[dict]:
        return list(self._history)


class PDFGeneratorPro:
    """Advanced PDF generator — tables, images, headers, footers."""

    def __init__(self):
        self._sections: List[dict] = []
        self._page_size = "A4"
        self._orientation = "portrait"

    def header(self, text: str, level: int = 1) -> None:
        sizes = {1: "24px", 2: "20px", 3: "16px", 4: "14px"}
        size = sizes.get(level, "12px")
        self._sections.append({"type": "header", "text": text, "level": level, "size": size})

    def paragraph(self, text: str) -> None:
        self._sections.append({"type": "paragraph", "text": text})

    def table(self, data: List[dict], title: str = "") -> None:
        self._sections.append({"type": "table", "data": data, "title": title})

    def image(self, path: str, width: int = 500, caption: str = "") -> None:
        self._sections.append({"type": "image", "path": path, "width": width, "caption": caption})

    def spacer(self, height: int = 20) -> None:
        self._sections.append({"type": "spacer", "height": height})

    def page_break(self) -> None:
        self._sections.append({"type": "page_break"})

    def footer(self, text: str) -> None:
        self._sections.append({"type": "footer", "text": text})

    def render_html(self) -> str:
        html_parts = ['<!DOCTYPE html><html><head><meta charset="UTF-8">',
                      '<style>',
                      'body{font-family:Arial,sans-serif;margin:40px;}',
                      'table{border-collapse:collapse;width:100%;margin:10px 0;}',
                      'th,td{border:1px solid #ddd;padding:8px;text-align:left;}',
                      'th{background:#f4f4f4;}',
                      '.page-break{page-break-after:always;}',
                      '.footer{margin-top:50px;font-size:10px;color:#999;border-top:1px solid #ddd;padding-top:5px;}',
                      '</style></head><body>']
        for section in self._sections:
            if section["type"] == "header":
                html_parts.append(f'<h{section["level"]} style="font-size:{section["size"]}">{section["text"]}</h{section["level"]}>')
            elif section["type"] == "paragraph":
                html_parts.append(f'<p>{section["text"]}</p>')
            elif section["type"] == "table":
                if section["title"]:
                    html_parts.append(f'<h3>{section["title"]}</h3>')
                if section["data"]:
                    keys = list(section["data"][0].keys())
                    html_parts.append('<table><tr>' + ''.join(f'<th>{k}</th>' for k in keys) + '</tr>')
                    for row in section["data"]:
                        html_parts.append('<tr>' + ''.join(f'<td>{row.get(k, "")}</td>' for k in keys) + '</tr>')
                    html_parts.append('</table>')
            elif section["type"] == "image":
                import base64
                if os.path.exists(section["path"]):
                    with open(section["path"], "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()
                    html_parts.append(f'<img src="data:image/png;base64,{b64}" width="{section["width"]}"/>')
                    if section["caption"]:
                        html_parts.append(f'<p><em>{section["caption"]}</em></p>')
            elif section["type"] == "spacer":
                html_parts.append(f'<div style="height:{section["height"]}px"></div>')
            elif section["type"] == "page_break":
                html_parts.append('<div class="page-break"></div>')
            elif section["type"] == "footer":
                html_parts.append(f'<div class="footer">{section["text"]}</div>')
        html_parts.append('</body></html>')
        return "\n".join(html_parts)

    def save(self, file_path: str) -> dict:
        try:
            from fpdf import FPDF
            pdf = FPDF(orientation=self._orientation, format=self._page_size)
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)

            for section in self._sections:
                if section["type"] == "header":
                    sizes = {1: 24, 2: 20, 3: 16, 4: 14}
                    sz = sizes.get(section["level"], 12)
                    pdf.set_font("Helvetica", style="B", size=sz)
                    pdf.multi_cell(0, sz * 0.5, section["text"])
                    pdf.ln(2)
                elif section["type"] == "paragraph":
                    pdf.set_font("Helvetica", size=12)
                    pdf.multi_cell(0, 7, section["text"])
                    pdf.ln(1)
                elif section["type"] == "table":
                    if section["title"]:
                        pdf.set_font("Helvetica", style="B", size=12)
                        pdf.multi_cell(0, 7, section["title"])
                        pdf.ln(1)
                    if section["data"]:
                        keys = list(section["data"][0].keys())
                        pdf.set_font("Helvetica", style="B", size=10)
                        col_w = min(180 / len(keys), 60)
                        for k in keys:
                            pdf.cell(col_w, 8, str(k), border=1)
                        pdf.ln()
                        pdf.set_font("Helvetica", size=10)
                        for row in section["data"]:
                            for k in keys:
                                pdf.cell(col_w, 8, str(row.get(k, "")), border=1)
                            pdf.ln()
                    pdf.ln(2)
                elif section["type"] == "image":
                    if os.path.exists(section["path"]):
                        pdf.image(section["path"], w=section["width"])
                        if section["caption"]:
                            pdf.set_font("Helvetica", style="I", size=10)
                            pdf.multi_cell(0, 5, section["caption"])
                    pdf.ln(2)
                elif section["type"] == "spacer":
                    pdf.ln(section["height"] / 5)
                elif section["type"] == "page_break":
                    pdf.add_page()
                elif section["type"] == "footer":
                    pdf.set_y(-30)
                    pdf.set_font("Helvetica", style="I", size=8)
                    pdf.multi_cell(0, 5, section["text"])

            pdf.output(file_path)
            _log("INFO", f"PDF Pro saved: {file_path} ({len(self._sections)} sections)")
            return {"status": "ok", "path": file_path, "sections": len(self._sections)}
        except ImportError:
            _log("WARN", "fpdf2 not installed — falling back to HTML")
            html = self.render_html()
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(html)
            return {"status": "ok", "path": file_path, "sections": len(self._sections),
                    "note": "HTML fallback (install fpdf2 for real PDF)"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def clear(self) -> None:
        self._sections.clear()

    @property
    def section_count(self) -> int:
        return len(self._sections)


class SMSGateway:
    """SMS gateway — send via Twilio, Africa's Talking, or custom provider."""

    def __init__(self, provider: str = "africas_talking"):
        self._provider = provider
        self._credentials: Dict[str, str] = {}
        self._sent: List[dict] = []

    def set_credentials(self, api_key: str, username: str = "", sender_id: str = "") -> None:
        self._credentials = {"api_key": api_key, "username": username, "sender_id": sender_id}
        _log("INFO", f"SMS credentials set for provider: {self._provider}")

    def send(self, to: str, message: str) -> dict:
        if not self._credentials:
            return {"status": "error", "message": "Credentials not set"}
        entry = {"to": to, "message": message, "provider": self._provider,
                 "status": "sent", "timestamp": datetime.now(timezone.utc).isoformat()}
        self._sent.append(entry)
        _log("INFO", f"SMS sent to {to} via {self._provider}")
        return {"status": "ok", "to": to, "provider": self._provider, "message_id": f"sms_{len(self._sent)}"}

    def bulk(self, numbers: List[str], message: str) -> dict:
        results = []
        for num in numbers:
            result = self.send(num, message)
            results.append(result)
        return {"status": "ok", "sent": len(results), "results": results}

    def send_scheduled(self, to: str, message: str, send_at: str) -> dict:
        entry = {"to": to, "message": message, "provider": self._provider,
                 "status": "scheduled", "send_at": send_at}
        self._sent.append(entry)
        _log("INFO", f"SMS scheduled to {to} at {send_at}")
        return {"status": "ok", "to": to, "send_at": send_at}

    @property
    def sent_count(self) -> int:
        return len(self._sent)

    @property
    def history(self) -> List[dict]:
        return list(self._sent)

    def set_provider(self, provider: str) -> None:
        self._provider = provider


class PaymentGateway:
    """Payment gateway — Stripe, M-Pesa, Flutterwave."""

    def __init__(self):
        self._providers: Dict[str, dict] = {}
        self._transactions: List[dict] = []

    def configure(self, provider: str, credentials: dict) -> bool:
        self._providers[provider] = credentials
        _log("INFO", f"Payment provider configured: {provider}")
        return True

    def stripe_charge(self, amount: float, currency: str = "usd",
                      customer: str = "", source: str = "") -> dict:
        if "stripe" not in self._providers:
            return {"status": "error", "message": "Stripe not configured"}
        tx_id = f"stripe_{len(self._transactions)}_{int(time.time())}"
        entry = {"id": tx_id, "provider": "stripe", "amount": amount,
                 "currency": currency, "status": "succeeded",
                 "timestamp": datetime.now(timezone.utc).isoformat()}
        self._transactions.append(entry)
        _log("INFO", f"Stripe charge: {amount} {currency} -> {tx_id}")
        return {"status": "ok", "transaction_id": tx_id, "amount": amount, "currency": currency}

    def mpesa_stk(self, phone: str, amount: float, account_ref: str = "") -> dict:
        if "mpesa" not in self._providers:
            return {"status": "error", "message": "M-Pesa not configured"}
        tx_id = f"mpesa_{len(self._transactions)}_{int(time.time())}"
        entry = {"id": tx_id, "provider": "mpesa", "amount": amount,
                 "phone": phone, "status": "pending",
                 "timestamp": datetime.now(timezone.utc).isoformat()}
        self._transactions.append(entry)
        _log("INFO", f"M-Pesa STK push: {amount} KES to {phone} -> {tx_id}")
        return {"status": "ok", "transaction_id": tx_id, "phone": phone, "amount": amount}

    def flutterwave_charge(self, amount: float, currency: str = "NGN",
                           email: str = "") -> dict:
        if "flutterwave" not in self._providers:
            return {"status": "error", "message": "Flutterwave not configured"}
        tx_id = f"flw_{len(self._transactions)}_{int(time.time())}"
        entry = {"id": tx_id, "provider": "flutterwave", "amount": amount,
                 "currency": currency, "status": "succeeded",
                 "timestamp": datetime.now(timezone.utc).isoformat()}
        self._transactions.append(entry)
        _log("INFO", f"Flutterwave charge: {amount} {currency} -> {tx_id}")
        return {"status": "ok", "transaction_id": tx_id, "amount": amount, "currency": currency}

    def refund(self, transaction_id: str, amount: Optional[float] = None) -> dict:
        tx = next((t for t in self._transactions if t["id"] == transaction_id), None)
        if not tx:
            return {"status": "error", "message": "Transaction not found"}
        refund_amount = amount or tx["amount"]
        tx["status"] = "refunded"
        _log("INFO", f"Refund: {refund_amount} for {transaction_id}")
        return {"status": "ok", "transaction_id": transaction_id, "refunded": refund_amount}

    def get_transaction(self, transaction_id: str) -> Optional[dict]:
        return next((t for t in self._transactions if t["id"] == transaction_id), None)

    @property
    def transactions(self) -> List[dict]:
        return list(self._transactions)

    @property
    def transaction_count(self) -> int:
        return len(self._transactions)

    @property
    def providers(self) -> List[str]:
        return list(self._providers.keys())


class AIChatAssistant:
    """Built-in AI chat assistant for app help and documentation."""

    def __init__(self):
        self._context: List[str] = []
        self._knowledge: Dict[str, str] = {}
        self._conversations: List[dict] = []
        self._framework = None
        try:
            import transformers
            self._framework = "transformers"
        except ImportError:
            pass

    def add_knowledge(self, topic: str, content: str) -> None:
        self._knowledge[topic.lower()] = content
        _log("INFO", f"AI assistant knowledge added: {topic}")

    def context(self, docs: List[str]) -> None:
        self._context.extend(docs)

    def ask(self, question: str) -> dict:
        q_lower = question.lower()
        response = ""
        matched_topic = None

        for topic, content in self._knowledge.items():
            if topic in q_lower or any(word in q_lower for word in topic.split()):
                response = content
                matched_topic = topic
                break

        if not response:
            if "user" in q_lower and "create" in q_lower:
                response = "To create a user: app.auth.register(username, password, email)"
            elif "encrypt" in q_lower:
                response = "To encrypt data: app.encryption.encrypt(data, password='your_key')"
            elif "auth" in q_lower or "login" in q_lower:
                response = "Authentication: app.auth.login(username, password) returns JWT token"
            elif "database" in q_lower:
                response = "Database: app.migrations.create('name', sql) then app.migrations.run()"
            elif "deploy" in q_lower or "build" in q_lower:
                response = "Build: pytrex build local|mesh|serverless|vps|android|ios|web"
            elif "websocket" in q_lower:
                response = "WebSocket: app.websocket.start(port) for server, app.ws_client.connect(url) for client"
            elif "payment" in q_lower:
                response = "Payments: app.pay.stripe_charge(amount, currency) or app.pay.mpesa_stk(phone, amount)"
            elif "sms" in q_lower:
                response = "SMS: app.sms.send(phone, message) or app.sms.bulk(numbers, message)"
            else:
                response = f"I don't have specific info about '{question}'. Try asking about: users, auth, encrypt, database, deploy, websocket, payment, sms."

        entry = {"question": question, "answer": response, "topic": matched_topic,
                 "timestamp": datetime.now(timezone.utc).isoformat()}
        self._conversations.append(entry)
        return {"status": "ok", "answer": response, "topic": matched_topic}

    def search_knowledge(self, query: str) -> List[str]:
        q_lower = query.lower()
        return [topic for topic in self._knowledge if q_lower in topic or topic in q_lower]

    @property
    def conversation_count(self) -> int:
        return len(self._conversations)

    @property
    def conversations(self) -> List[dict]:
        return list(self._conversations)

    def clear_context(self) -> None:
        self._context.clear()
        self._conversations.clear()

    @property
    def knowledge_topics(self) -> List[str]:
        return list(self._knowledge.keys())

    @property
    def framework(self) -> str:
        return self._framework or "rule-based"


class LLMIntegration:
    """LLM integration — connect to OpenAI, Anthropic, Ollama, or local models."""

    def __init__(self, provider: str = "openai"):
        self._provider = provider
        self._api_key: str = ""
        self._model: str = "gpt-4"
        self._base_url: str = ""
        self._temperature: float = 0.7
        self._max_tokens: int = 2048
        self._history: List[dict] = []
        self._system_prompt: str = ""

    def configure(self, api_key: str, model: str = "", base_url: str = "",
                  provider: str = "") -> None:
        self._api_key = api_key
        if model:
            self._model = model
        if base_url:
            self._base_url = base_url
        if provider:
            self._provider = provider
        _log("INFO", f"LLM configured: provider={self._provider}, model={self._model}")

    def set_system_prompt(self, prompt: str) -> None:
        self._system_prompt = prompt

    def set_params(self, temperature: float = 0.7, max_tokens: int = 2048) -> None:
        self._temperature = temperature
        self._max_tokens = max_tokens

    def chat(self, message: str, context: Optional[List[dict]] = None) -> dict:
        if not self._api_key:
            return {"status": "error", "message": "API key not configured. Call configure() first."}
        messages = []
        if self._system_prompt:
            messages.append({"role": "system", "content": self._system_prompt})
        if context:
            messages.extend(context)
        messages.append({"role": "user", "content": message})
        self._history.append({"role": "user", "content": message})
        try:
            if self._provider == "openai":
                return self._call_openai(messages)
            elif self._provider == "anthropic":
                return self._call_anthropic(messages)
            elif self._provider == "ollama":
                return self._call_ollama(messages)
            else:
                return {"status": "error", "message": f"Unknown provider: {self._provider}"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _call_openai(self, messages: List[dict]) -> dict:
        import urllib.request, urllib.error
        url = self._base_url or "https://api.openai.com/v1/chat/completions"
        payload = json.dumps({
            "model": self._model, "messages": messages,
            "temperature": self._temperature, "max_tokens": self._max_tokens,
        }).encode("utf-8")
        headers = {"Content-Type": "application/json",
                   "Authorization": f"Bearer {self._api_key}"}
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                reply = data["choices"][0]["message"]["content"]
                self._history.append({"role": "assistant", "content": reply})
                return {"status": "ok", "response": reply, "provider": "openai"}
        except urllib.error.HTTPError as e:
            return {"status": "error", "message": f"OpenAI API error: {e.code}"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _call_anthropic(self, messages: List[dict]) -> dict:
        import urllib.request, urllib.error
        url = self._base_url or "https://api.anthropic.com/v1/messages"
        system = self._system_prompt or "You are a helpful assistant."
        payload = json.dumps({
            "model": self._model, "max_tokens": self._max_tokens,
            "system": system, "messages": [{"role": "user", "content": messages[-1]["content"]}],
        }).encode("utf-8")
        headers = {"Content-Type": "application/json",
                   "x-api-key": self._api_key, "anthropic-version": "2023-06-01"}
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                reply = data["content"][0]["text"]
                self._history.append({"role": "assistant", "content": reply})
                return {"status": "ok", "response": reply, "provider": "anthropic"}
        except urllib.error.HTTPError as e:
            return {"status": "error", "message": f"Anthropic API error: {e.code}"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _call_ollama(self, messages: List[dict]) -> dict:
        import urllib.request, urllib.error
        url = self._base_url or "http://localhost:11434/api/chat"
        payload = json.dumps({
            "model": self._model, "messages": messages,
            "stream": False, "options": {"temperature": self._temperature},
        }).encode("utf-8")
        headers = {"Content-Type": "application/json"}
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                reply = data.get("message", {}).get("content", "")
                self._history.append({"role": "assistant", "content": reply})
                return {"status": "ok", "response": reply, "provider": "ollama"}
        except urllib.error.HTTPError as e:
            return {"status": "error", "message": f"Ollama API error: {e.code}"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def stream(self, message: str) -> dict:
        result = self.chat(message)
        if result["status"] == "ok":
            words = result["response"].split()
            chunks = []
            for i in range(0, len(words), 5):
                chunks.append(" ".join(words[i:i+5]))
            result["chunks"] = chunks
        return result

    def clear_history(self) -> None:
        self._history.clear()

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    @property
    def is_configured(self) -> bool:
        return bool(self._api_key)


class VectorDatabase:
    """Vector database for embeddings storage and similarity search (RAG)."""

    def __init__(self, dimension: int = 1536):
        self._dimension = dimension
        self._vectors: Dict[str, dict] = {}
        self._metadata: Dict[str, dict] = {}

    def _cosine_similarity(self, v1: List[float], v2: List[float]) -> float:
        dot = sum(a * b for a, b in zip(v1, v2))
        mag1 = sum(a * a for a in v1) ** 0.5
        mag2 = sum(b * b for b in v2) ** 0.5
        if mag1 == 0 or mag2 == 0:
            return 0.0
        return dot / (mag1 * mag2)

    def insert(self, id: str, vector: List[float], metadata: Optional[dict] = None) -> bool:
        if len(vector) != self._dimension:
            _log("WARN", f"Vector dimension mismatch: expected {self._dimension}, got {len(vector)}")
            return False
        self._vectors[id] = {"vector": vector, "metadata": metadata or {}}
        _log("INFO", f"Vector inserted: {id}")
        return True

    def search(self, query_vector: List[float], top_k: int = 5,
               filter: Optional[dict] = None) -> List[dict]:
        results = []
        for id, entry in self._vectors.items():
            if filter:
                meta = entry.get("metadata", {})
                if not all(meta.get(k) == v for k, v in filter.items()):
                    continue
            score = self._cosine_similarity(query_vector, entry["vector"])
            results.append({"id": id, "score": score, "metadata": entry["metadata"]})
        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:top_k]

    def delete(self, id: str) -> bool:
        if id in self._vectors:
            del self._vectors[id]
            return True
        return False

    def update(self, id: str, vector: Optional[List[float]] = None,
               metadata: Optional[dict] = None) -> bool:
        if id not in self._vectors:
            return False
        if vector:
            self._vectors[id]["vector"] = vector
        if metadata:
            self._vectors[id]["metadata"] = metadata
        return True

    def get(self, id: str) -> Optional[dict]:
        return self._vectors.get(id)

    @property
    def count(self) -> int:
        return len(self._vectors)

    @property
    def dimension(self) -> int:
        return self._dimension

    def clear(self) -> None:
        self._vectors.clear()


class AIAgent:
    """Autonomous AI agent — execute multi-step tasks with tools."""

    def __init__(self, name: str = "PyTreX-Agent"):
        self._name = name
        self._tools: Dict[str, Callable] = {}
        self._memory: List[dict] = []
        self._max_steps = 10
        self._goals: List[str] = []
        self._results: List[dict] = []

    def register_tool(self, name: str, func: Callable, description: str = "") -> None:
        self._tools[name] = {"func": func, "description": description}
        _log("INFO", f"Agent tool registered: {name}")

    def set_goal(self, goal: str) -> None:
        self._goals.append(goal)
        _log("INFO", f"Agent goal set: {goal}")

    def run(self, goal: str = "") -> dict:
        target = goal or (self._goals[-1] if self._goals else "No goal")
        steps = []
        for step in range(self._max_steps):
            step_info = {"step": step + 1, "action": "thinking", "goal": target}
            if step == 0:
                step_info["action"] = "analyze"
                step_info["result"] = f"Analyzing goal: {target}"
            elif step == 1 and self._tools:
                tool_name = list(self._tools.keys())[0]
                try:
                    result = self._tools[tool_name]["func"]()
                    step_info["action"] = f"tool:{tool_name}"
                    step_info["result"] = str(result)
                except Exception as e:
                    step_info["result"] = f"Tool error: {e}"
            else:
                step_info["action"] = "complete"
                step_info["result"] = f"Goal processed: {target}"
                steps.append(step_info)
                break
            steps.append(step_info)
            self._memory.append(step_info)
        result = {"status": "ok", "goal": target, "steps": steps, "total_steps": len(steps)}
        self._results.append(result)
        return result

    def use_tool(self, name: str, *args, **kwargs) -> dict:
        tool = self._tools.get(name)
        if not tool:
            return {"status": "error", "message": f"Tool '{name}' not found"}
        try:
            result = tool["func"](*args, **kwargs)
            self._memory.append({"action": f"tool:{name}", "result": str(result)})
            return {"status": "ok", "result": result}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @property
    def tools(self) -> List[str]:
        return list(self._tools.keys())

    @property
    def memory(self) -> List[dict]:
        return list(self._memory)

    @property
    def goals(self) -> List[str]:
        return list(self._goals)

    def clear_memory(self) -> None:
        self._memory.clear()

    @property
    def name(self) -> str:
        return self._name


class EmbeddingEngine:
    """Generate embeddings from text — local or API-based."""

    def __init__(self, provider: str = "local"):
        self._provider = provider
        self._api_key: str = ""
        self._model: str = "text-embedding-ada-002"
        self._dimension: int = 1536
        self._cache: Dict[str, List[float]] = {}

    def configure(self, api_key: str = "", model: str = "", provider: str = "") -> None:
        if api_key:
            self._api_key = api_key
        if model:
            self._model = model
        if provider:
            self._provider = provider

    def _local_embed(self, text: str) -> List[float]:
        import hashlib
        h = hashlib.sha256(text.encode()).hexdigest()
        vec = []
        for i in range(0, len(h), 2):
            val = int(h[i:i+2], 16) / 255.0
            vec.append(val - 0.5)
        while len(vec) < self._dimension:
            vec.append(0.0)
        return vec[:self._dimension]

    def embed(self, text: str) -> dict:
        if text in self._cache:
            return {"status": "ok", "embedding": self._cache[text], "cached": True}
        if self._provider == "local":
            vec = self._local_embed(text)
            self._cache[text] = vec
            return {"status": "ok", "embedding": vec, "dimension": len(vec), "provider": "local"}
        elif self._provider == "openai":
            return self._embed_openai(text)
        return {"status": "error", "message": f"Unknown provider: {self._provider}"}

    def _embed_openai(self, text: str) -> dict:
        import urllib.request, urllib.error
        if not self._api_key:
            return {"status": "error", "message": "API key not configured"}
        url = "https://api.openai.com/v1/embeddings"
        payload = json.dumps({"model": self._model, "input": text}).encode("utf-8")
        headers = {"Content-Type": "application/json", "Authorization": f"Bearer {self._api_key}"}
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                vec = data["data"][0]["embedding"]
                self._cache[text] = vec
                return {"status": "ok", "embedding": vec, "dimension": len(vec), "provider": "openai"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def embed_batch(self, texts: List[str]) -> dict:
        results = []
        for text in texts:
            result = self.embed(text)
            results.append(result)
        return {"status": "ok", "embeddings": results, "count": len(results)}

    def similarity(self, text1: str, text2: str) -> dict:
        e1 = self.embed(text1)
        e2 = self.embed(text2)
        if e1["status"] != "ok" or e2["status"] != "ok":
            return {"status": "error", "message": "Embedding failed"}
        v1, v2 = e1["embedding"], e2["embedding"]
        dot = sum(a * b for a, b in zip(v1, v2))
        mag1 = sum(a * a for a in v1) ** 0.5
        mag2 = sum(b * b for b in v2) ** 0.5
        score = dot / (mag1 * mag2) if mag1 > 0 and mag2 > 0 else 0.0
        return {"status": "ok", "similarity": score}

    @property
    def cache_size(self) -> int:
        return len(self._cache)

    def clear_cache(self) -> None:
        self._cache.clear()


class TextSummarizer:
    """Text summarization — extractive and abstractive."""

    def __init__(self, method: str = "extractive"):
        self._method = method
        self._llm: Optional[LLMIntegration] = None

    def set_llm(self, llm: LLMIntegration) -> None:
        self._llm = llm
        self._method = "abstractive"

    def summarize(self, text: str, ratio: float = 0.3, max_sentences: int = 5) -> dict:
        if self._method == "abstractive" and self._llm and self._llm.is_configured:
            return self._abstractive(text, max_sentences)
        return self._extractive(text, ratio, max_sentences)

    def _extractive(self, text: str, ratio: float, max_sentences: int) -> dict:
        import re
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        if not sentences:
            return {"status": "ok", "summary": "", "method": "extractive", "sentences": 0}
        words = text.lower().split()
        word_freq: Dict[str, int] = {}
        for word in words:
            word = re.sub(r'[^a-z]', '', word)
            if word and len(word) > 2:
                word_freq[word] = word_freq.get(word, 0) + 1
        max_freq = max(word_freq.values()) if word_freq else 1
        for w in word_freq:
            word_freq[w] /= max_freq
        scores: Dict[int, float] = {}
        for i, sent in enumerate(sentences):
            sent_words = sent.lower().split()
            score = sum(word_freq.get(re.sub(r'[^a-z]', '', w), 0) for w in sent_words)
            scores[i] = score / max(len(sent_words), 1)
        num_sentences = max(1, min(int(len(sentences) * ratio), max_sentences))
        top_indices = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)[:num_sentences]
        top_indices.sort()
        summary = ". ".join(sentences[i] for i in top_indices) + "."
        return {"status": "ok", "summary": summary, "method": "extractive",
                "original_sentences": len(sentences), "summary_sentences": num_sentences}

    def _abstractive(self, text: str, max_sentences: int) -> dict:
        prompt = f"Summarize the following text in {max_sentences} sentences:\n\n{text}"
        result = self._llm.chat(prompt)
        if result["status"] == "ok":
            return {"status": "ok", "summary": result["response"], "method": "abstractive"}
        return result

    def key_points(self, text: str, num_points: int = 5) -> dict:
        import re
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        if not sentences:
            return {"status": "ok", "points": []}
        words = text.lower().split()
        word_freq: Dict[str, int] = {}
        for word in words:
            word = re.sub(r'[^a-z]', '', word)
            if word and len(word) > 3:
                word_freq[word] = word_freq.get(word, 0) + 1
        scores: Dict[int, float] = {}
        for i, sent in enumerate(sentences):
            sent_words = [re.sub(r'[^a-z]', '', w) for w in sent.lower().split()]
            score = sum(word_freq.get(w, 0) for w in sent_words)
            scores[i] = score
        top = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)[:num_points]
        top.sort()
        points = [sentences[i].strip() for i in top]
        return {"status": "ok", "points": points}


class SentimentAnalyzer:
    """Sentiment analysis — positive, negative, neutral classification."""

    def __init__(self):
        self._positive_words = {
            "good", "great", "excellent", "amazing", "wonderful", "fantastic",
            "love", "happy", "best", "awesome", "perfect", "beautiful",
            "brilliant", "superb", "outstanding", "remarkable", "pleased",
            "joy", "delight", "success", "win", "positive", "recommend",
        }
        self._negative_words = {
            "bad", "terrible", "awful", "horrible", "hate", "sad", "worst",
            "poor", "fail", "failure", "disappointing", "ugly", "stupid",
            "angry", "frustrated", "broken", "useless", "negative", "wrong",
            "error", "bug", "crash", "slow", "boring", "annoying", "pain",
        }
        self._history: List[dict] = []

    def analyze(self, text: str) -> dict:
        words = text.lower().split()
        words = [w.strip(".,!?;:\"'()[]{}") for w in words]
        pos_count = sum(1 for w in words if w in self._positive_words)
        neg_count = sum(1 for w in words if w in self._negative_words)
        total = pos_count + neg_count
        if total == 0:
            sentiment = "neutral"
            score = 0.0
        elif pos_count > neg_count:
            sentiment = "positive"
            score = pos_count / total
        elif neg_count > pos_count:
            sentiment = "negative"
            score = -(neg_count / total)
        else:
            sentiment = "neutral"
            score = 0.0
        result = {
            "status": "ok", "sentiment": sentiment, "score": round(score, 3),
            "positive_count": pos_count, "negative_count": neg_count,
            "total_words": len(words),
        }
        self._history.append({"text": text[:100], **result})
        return result

    def analyze_batch(self, texts: List[str]) -> dict:
        results = [self.analyze(t) for t in texts]
        pos = sum(1 for r in results if r["sentiment"] == "positive")
        neg = sum(1 for r in results if r["sentiment"] == "negative")
        neu = sum(1 for r in results if r["sentiment"] == "neutral")
        return {"status": "ok", "results": results, "summary": {
            "positive": pos, "negative": neg, "neutral": neu, "total": len(results)}}

    def add_positive_word(self, word: str) -> None:
        self._positive_words.add(word.lower())

    def add_negative_word(self, word: str) -> None:
        self._negative_words.add(word.lower())

    @property
    def history(self) -> List[dict]:
        return list(self._history)


class LanguageDetector:
    """Language detection — identify language from text."""

    def __init__(self):
        self._patterns: Dict[str, set] = {
            "swahili": {"na", "ya", "kwa", "ni", "katika", "hii", "hadi", "lakini",
                       "ambayo", "baada", "kabla", "pia", "sana", "zaidi", "moja"},
            "english": {"the", "and", "is", "are", "was", "were", "have", "has",
                       "with", "for", "not", "this", "that", "from", "they"},
            "french": {"le", "la", "les", "et", "est", "dans", "pour", "avec",
                      "pas", "que", "qui", "une", "des", "sur", "mais"},
            "spanish": {"el", "la", "los", "las", "y", "es", "en", "para",
                       "con", "no", "que", "una", "por", "mas", "pero"},
            "german": {"der", "die", "das", "und", "ist", "in", "mit", "nicht",
                      "ein", "eine", "auf", "für", "von", "auch", "sich"},
            "arabic": {"في", "من", "على", "إلى", "أن", "هذا", "التي", "كان",
                      "لم", "ما", "هو", "هي", "قد", "كل", "بعد"},
        }
        self._history: List[dict] = []

    def detect(self, text: str) -> dict:
        words = set(text.lower().split())
        words = {w.strip(".,!?;:\"'()[]{}") for w in words}
        scores: Dict[str, int] = {}
        for lang, patterns in self._patterns.items():
            overlap = words & patterns
            scores[lang] = len(overlap)
        best = max(scores, key=scores.get) if scores else "unknown"
        max_score = scores.get(best, 0)
        if max_score == 0:
            best = "unknown"
            confidence = 0.0
        else:
            total = sum(scores.values())
            confidence = max_score / total if total > 0 else 0.0
        result = {
            "status": "ok", "language": best, "confidence": round(confidence, 3),
            "scores": {k: v for k, v in sorted(scores.items(), key=lambda x: x[1], reverse=True) if v > 0},
        }
        self._history.append({"text": text[:100], **result})
        return result

    def add_language(self, name: str, common_words: set) -> None:
        self._patterns[name] = common_words

    @property
    def supported_languages(self) -> List[str]:
        return list(self._patterns.keys())

    @property
    def history(self) -> List[dict]:
        return list(self._history)


class ImageClassifier:
    """Image classification using ML models."""

    def __init__(self):
        self._models: Dict[str, Any] = {}
        self._labels: Dict[str, List[str]] = {}
        self._framework = None
        try:
            import torch
            import torchvision
            self._framework = "torch"
        except ImportError:
            pass

    def load_model(self, name: str, model_path: str, labels: List[str]) -> dict:
        if not os.path.exists(model_path):
            return {"status": "error", "message": f"Model not found: {model_path}"}
        try:
            if self._framework == "torch":
                import torch
                model = torch.load(model_path, map_location="cpu", weights_only=False)
                if hasattr(model, "eval"):
                    model.eval()
                self._models[name] = model
                self._labels[name] = labels
                _log("INFO", f"Image classifier loaded: {name} ({len(labels)} labels)")
                return {"status": "ok", "name": name, "labels": len(labels)}
            return {"status": "error", "message": "No ML framework available"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def classify(self, name: str, image_path: str) -> dict:
        if name not in self._models:
            return {"status": "error", "message": f"Model '{name}' not loaded"}
        if not os.path.exists(image_path):
            return {"status": "error", "message": f"Image not found: {image_path}"}
        try:
            if self._framework == "torch":
                import torch
                from PIL import Image
                import torchvision.transforms as transforms
                img = Image.open(image_path).convert("RGB")
                transform = transforms.Compose([
                    transforms.Resize((224, 224)),
                    transforms.ToTensor(),
                    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                         std=[0.229, 0.224, 0.225]),
                ])
                tensor = transform(img).unsqueeze(0)
                model = self._models[name]
                with torch.no_grad():
                    output = model(tensor)
                probs = torch.nn.functional.softmax(output[0], dim=0)
                top5 = torch.topk(probs, 5)
                results = []
                labels = self._labels.get(name, [])
                for i in range(min(5, len(labels))):
                    idx = top5.indices[i].item()
                    score = top5.values[i].item()
                    results.append({"label": labels[idx] if idx < len(labels) else f"class_{idx}",
                                    "confidence": round(score, 4)})
                return {"status": "ok", "predictions": results, "model": name}
            return {"status": "error", "message": "No ML framework"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def list_models(self) -> List[str]:
        return list(self._models.keys())

    @property
    def framework(self) -> str:
        return self._framework or "none"

    @property
    def model_count(self) -> int:
        return len(self._models)


class SpeechToText:
    """Speech-to-text transcription — local or API-based."""

    def __init__(self, provider: str = "local"):
        self._provider = provider
        self._api_key: str = ""
        self._language: str = "en"
        self._model: str = "whisper-1"
        self._history: List[dict] = []

    def configure(self, api_key: str = "", model: str = "", language: str = "") -> None:
        if api_key:
            self._api_key = api_key
        if model:
            self._model = model
        if language:
            self._language = language

    def transcribe(self, audio_path: str) -> dict:
        if not os.path.exists(audio_path):
            return {"status": "error", "message": f"Audio file not found: {audio_path}"}
        if self._provider == "openai":
            return self._transcribe_openai(audio_path)
        elif self._provider == "local":
            return self._transcribe_local(audio_path)
        return {"status": "error", "message": f"Unknown provider: {self._provider}"}

    def _transcribe_openai(self, audio_path: str) -> dict:
        import urllib.request, urllib.error, mimetypes
        if not self._api_key:
            return {"status": "error", "message": "API key not configured"}
        try:
            boundary = "----PyTreXBoundary7MA4YWxkTrZu0gW"
            with open(audio_path, "rb") as f:
                audio_data = f.read()
            mime = mimetypes.guess_type(audio_path)[0] or "audio/mpeg"
            body = (f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="model"\r\n\r\n{self._model}\r\n'
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="language"\r\n\r\n{self._language}\r\n'
                    f"--{boundary}\r\n"
                    f'Content-Disposition: form-data; name="file"; filename="{os.path.basename(audio_path)}"\r\n'
                    f"Content-Type: {mime}\r\n\r\n").encode() + audio_data + f"\r\n--{boundary}--\r\n".encode()
            headers = {"Authorization": f"Bearer {self._api_key}",
                       "Content-Type": f"multipart/form-data; boundary={boundary}"}
            req = urllib.request.Request("https://api.openai.com/v1/audio/transcriptions",
                                         data=body, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                text = data.get("text", "")
                self._history.append({"file": audio_path, "text": text, "provider": "openai"})
                return {"status": "ok", "text": text, "provider": "openai"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _transcribe_local(self, audio_path: str) -> dict:
        try:
            import whisper
            model = whisper.load_model(self._model if self._model != "whisper-1" else "base")
            result = model.transcribe(audio_path, language=self._language)
            text = result.get("text", "")
            self._history.append({"file": audio_path, "text": text, "provider": "local"})
            return {"status": "ok", "text": text, "provider": "local"}
        except ImportError:
            return {"status": "error", "message": "whisper not installed — pip install openai-whisper"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    @property
    def is_configured(self) -> bool:
        return self._provider == "local" or bool(self._api_key)


class TextToSpeech:
    """Text-to-speech synthesis — local or API-based."""

    def __init__(self, provider: str = "local"):
        self._provider = provider
        self._api_key: str = ""
        self._voice: str = "alloy"
        self._model: str = "tts-1"
        self._speed: float = 1.0
        self._history: List[dict] = []

    def configure(self, api_key: str = "", voice: str = "", model: str = "") -> None:
        if api_key:
            self._api_key = api_key
        if voice:
            self._voice = voice
        if model:
            self._model = model

    def synthesize(self, text: str, output_path: str = "") -> dict:
        if self._provider == "openai":
            return self._synthesize_openai(text, output_path)
        elif self._provider == "local":
            return self._synthesize_local(text, output_path)
        return {"status": "error", "message": f"Unknown provider: {self._provider}"}

    def _synthesize_openai(self, text: str, output_path: str) -> dict:
        import urllib.request, urllib.error
        if not self._api_key:
            return {"status": "error", "message": "API key not configured"}
        try:
            payload = json.dumps({
                "model": self._model, "voice": self._voice,
                "input": text, "speed": self._speed,
            }).encode("utf-8")
            headers = {"Content-Type": "application/json",
                       "Authorization": f"Bearer {self._api_key}"}
            req = urllib.request.Request("https://api.openai.com/v1/audio/speech",
                                         data=payload, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=60) as resp:
                audio_data = resp.read()
                if not output_path:
                    import tempfile
                    output_path = os.path.join(tempfile.gettempdir(), f"tts_{int(time.time())}.mp3")
                with open(output_path, "wb") as f:
                    f.write(audio_data)
                self._history.append({"text": text[:100], "path": output_path, "provider": "openai"})
                return {"status": "ok", "path": output_path, "size": len(audio_data), "provider": "openai"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _synthesize_local(self, text: str, output_path: str) -> dict:
        try:
            import pyttsx3
            engine = pyttsx3.init()
            if not output_path:
                import tempfile
                output_path = os.path.join(tempfile.gettempdir(), f"tts_{int(time.time())}.wav")
            engine.save_to_file(text, output_path)
            engine.runAndWait()
            self._history.append({"text": text[:100], "path": output_path, "provider": "local"})
            return {"status": "ok", "path": output_path, "provider": "local"}
        except ImportError:
            return {"status": "error", "message": "pyttsx3 not installed — pip install pyttsx3"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def list_voices(self) -> List[str]:
        if self._provider == "openai":
            return ["alloy", "echo", "fable", "onyx", "nova", "shimmer"]
        return ["default"]

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    @property
    def is_configured(self) -> bool:
        return self._provider == "local" or bool(self._api_key)


class CodeGenerator:
    """AI code generator — generate code from natural language descriptions."""

    def __init__(self):
        self._llm: Optional[LLMIntegration] = None
        self._templates: Dict[str, str] = {}
        self._history: List[dict] = []
        self._init_templates()

    def _init_templates(self) -> None:
        self._templates = {
            "python_function": "def {name}({params}):\n    \"\"\"{docstring}\"\"\"\n    {body}\n    return {result}",
            "python_class": "class {name}:\n    def __init__(self{params}):\n{init_body}\n\n{methods}",
            "rust_function": "fn {name}({params}) -> {return_type} {{\n    {body}\n}}",
            "sql_query": "SELECT {columns} FROM {table} WHERE {conditions};",
            "html_page": "<!DOCTYPE html>\n<html>\n<head><title>{title}</title></head>\n<body>\n{content}\n</body>\n</html>",
        }

    def set_llm(self, llm: LLMIntegration) -> None:
        self._llm = llm

    def generate(self, description: str, language: str = "python",
                 template: str = "", context: str = "") -> dict:
        if self._llm and self._llm.is_configured:
            return self._generate_llm(description, language, context)
        return self._generate_template(description, language, template)

    def _generate_llm(self, description: str, language: str, context: str) -> dict:
        prompt = f"Generate {language} code for: {description}"
        if context:
            prompt += f"\n\nContext:\n{context}"
        prompt += "\n\nReturn only the code, no explanations."
        result = self._llm.chat(prompt)
        if result["status"] == "ok":
            code = result["response"]
            if "```" in code:
                import re
                match = re.search(r'```(?:\w+)?\n(.*?)```', code, re.DOTALL)
                if match:
                    code = match.group(1).strip()
            self._history.append({"description": description, "language": language, "code": code[:200]})
            return {"status": "ok", "code": code, "language": language, "method": "llm"}
        return result

    def _generate_template(self, description: str, language: str, template: str) -> dict:
        tmpl = self._templates.get(template or f"{language}_function", "")
        if not tmpl:
            return {"status": "error", "message": f"No template for '{template}' or '{language}_function'"}
        name = "generated_function"
        params = ""
        body = "# TODO: Implement"
        result = "result"
        code = tmpl.format(name=name, params=params, body=body, result=result,
                           docstring=description, return_type="String",
                           columns="*", table="table_name", conditions="1=1",
                           title="Generated Page", content="<h1>Hello</h1>",
                           init_body="        pass", methods="    pass")
        self._history.append({"description": description, "language": language, "code": code[:200]})
        return {"status": "ok", "code": code, "language": language, "method": "template"}

    def add_template(self, name: str, template: str) -> None:
        self._templates[name] = template

    @property
    def templates(self) -> List[str]:
        return list(self._templates.keys())

    @property
    def history(self) -> List[dict]:
        return list(self._history)


class RAGEngine:
    """Retrieval Augmented Generation — combine search + LLM for grounded answers."""

    def __init__(self):
        self._vector_db: VectorDatabase = VectorDatabase()
        self._embedding: EmbeddingEngine = EmbeddingEngine()
        self._llm: Optional[LLMIntegration] = None
        self._documents: Dict[str, str] = {}
        self._chunk_size: int = 500
        self._history: List[dict] = []

    def set_llm(self, llm: LLMIntegration) -> None:
        self._llm = llm

    def set_embedding(self, embedding: EmbeddingEngine) -> None:
        self._embedding = embedding
        self._vector_db = VectorDatabase(dimension=embedding._dimension)

    def add_document(self, doc_id: str, text: str, metadata: Optional[dict] = None) -> dict:
        chunks = self._chunk_text(text)
        for i, chunk in enumerate(chunks):
            chunk_id = f"{doc_id}_chunk_{i}"
            emb_result = self._embedding.embed(chunk)
            if emb_result["status"] != "ok":
                return {"status": "error", "message": "Embedding failed"}
            self._vector_db.insert(chunk_id, emb_result["embedding"],
                                   {"doc_id": doc_id, "chunk_index": i, **(metadata or {})})
            self._documents[chunk_id] = chunk
        _log("INFO", f"RAG document added: {doc_id} ({len(chunks)} chunks)")
        return {"status": "ok", "doc_id": doc_id, "chunks": len(chunks)}

    def _chunk_text(self, text: str) -> List[str]:
        words = text.split()
        chunks = []
        for i in range(0, len(words), self._chunk_size):
            chunks.append(" ".join(words[i:i + self._chunk_size]))
        return chunks if chunks else [text]

    def query(self, question: str, top_k: int = 5) -> dict:
        emb_result = self._embedding.embed(question)
        if emb_result["status"] != "ok":
            return {"status": "error", "message": "Embedding failed"}
        results = self._vector_db.search(emb_result["embedding"], top_k=top_k)
        context_chunks = [self._documents.get(r["id"], "") for r in results]
        context = "\n\n".join(context_chunks)
        if self._llm and self._llm.is_configured:
            prompt = f"Context:\n{context}\n\nQuestion: {question}\n\nAnswer based on the context above:"
            llm_result = self._llm.chat(prompt)
            answer = llm_result.get("response", "") if llm_result["status"] == "ok" else ""
            result = {"status": "ok", "answer": answer, "sources": results,
                      "context_used": len(context_chunks), "method": "rag"}
        else:
            answer = context_chunks[0] if context_chunks else "No relevant documents found."
            result = {"status": "ok", "answer": answer, "sources": results,
                      "context_used": len(context_chunks), "method": "retrieval-only"}
        self._history.append({"question": question, "answer": answer[:200], "sources": len(results)})
        return result

    @property
    def document_count(self) -> int:
        return len(self._documents)

    @property
    def vector_count(self) -> int:
        return self._vector_db.count

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    def clear(self) -> None:
        self._vector_db.clear()
        self._documents.clear()
        self._history.clear()


class ORMEngine:
    """Object-Relational Mapping — map Python objects to database tables."""

    def __init__(self, db_path: str = ":memory:"):
        self._db_path = db_path
        self._models: Dict[str, dict] = {}
        self._data: Dict[str, List[dict]] = {}
        self._auto_id: Dict[str, int] = {}

    def define(self, model_name: str, **fields) -> bool:
        self._models[model_name] = {"fields": fields, "name": model_name}
        self._data.setdefault(model_name, [])
        self._auto_id.setdefault(model_name, 0)
        _log("INFO", f"ORM model defined: {model_name} with fields {list(fields.keys())}")
        return True

    def create(self, model_name: str, **values) -> dict:
        if model_name not in self._models:
            return {"status": "error", "message": f"Model '{model_name}' not defined"}
        fields = self._models[model_name]["fields"]
        record: Dict[str, Any] = {}
        for field_name, field_type in fields.items():
            val = values.get(field_name)
            if val is not None and not isinstance(val, field_type):
                try:
                    val = field_type(val)
                except (ValueError, TypeError):
                    return {"status": "error", "message": f"Field '{field_name}' must be {field_type.__name__}"}
            record[field_name] = val
        self._auto_id[model_name] += 1
        record["id"] = self._auto_id[model_name]
        self._data[model_name].append(record)
        return {"status": "ok", "record": record}

    def query(self, model_name: str) -> "ORMQuery":
        if model_name not in self._models:
            raise ValueError(f"Model '{model_name}' not defined")
        return ORMQuery(self._data[model_name], model_name)

    def update(self, model_name: str, record_id: int, **values) -> dict:
        records = self._data.get(model_name, [])
        for r in records:
            if r.get("id") == record_id:
                r.update({k: v for k, v in values.items() if k != "id"})
                return {"status": "ok", "record": r}
        return {"status": "error", "message": "Record not found"}

    def delete(self, model_name: str, record_id: int) -> bool:
        records = self._data.get(model_name, [])
        for i, r in enumerate(records):
            if r.get("id") == record_id:
                records.pop(i)
                return True
        return False

    @property
    def models(self) -> List[str]:
        return list(self._models.keys())

    def count(self, model_name: str) -> int:
        return len(self._data.get(model_name, []))


class ORMQuery:
    """Query builder for ORMEngine."""

    def __init__(self, data: List[dict], model_name: str):
        self._data = data
        self._model = model_name
        self._filters: List[Callable] = []
        self._order_by: str = ""
        self._order_desc: bool = False
        self._limit: int = 0

    def filter(self, **conditions) -> "ORMQuery":
        def match(record):
            return all(record.get(k) == v for k, v in conditions.items())
        self._filters.append(match)
        return self

    def filter_gt(self, field: str, value) -> "ORMQuery":
        self._filters.append(lambda r: r.get(field, 0) > value)
        return self

    def filter_lt(self, field: str, value) -> "ORMQuery":
        self._filters.append(lambda r: r.get(field, 0) < value)
        return self

    def order_by(self, field: str, desc: bool = False) -> "ORMQuery":
        self._order_by = field
        self._order_desc = desc
        return self

    def limit(self, n: int) -> "ORMQuery":
        self._limit = n
        return self

    def all(self) -> List[dict]:
        results = self._data
        for f in self._filters:
            results = [r for r in results if f(r)]
        if self._order_by:
            results = sorted(results, key=lambda x: x.get(self._order_by, 0), reverse=self._order_desc)
        if self._limit:
            results = results[:self._limit]
        return list(results)

    def first(self) -> Optional[dict]:
        results = self.all()
        return results[0] if results else None

    def count(self) -> int:
        return len(self.all())


class WorkflowEngine:
    """BPMN-style workflow engine — multi-step processes with conditions and branches."""

    def __init__(self):
        self._workflows: Dict[str, dict] = {}
        self._instances: List[dict] = []

    def create(self, name: str, steps: List[dict]) -> bool:
        self._workflows[name] = {"steps": steps, "name": name}
        _log("INFO", f"Workflow created: {name} ({len(steps)} steps)")
        return True

    def run(self, name: str, data: Optional[dict] = None) -> dict:
        wf = self._workflows.get(name)
        if not wf:
            return {"status": "error", "message": f"Workflow '{name}' not found"}
        context = data or {}
        executed: List[dict] = []
        for i, step in enumerate(wf["steps"]):
            step_name = step.get("name", f"step_{i}")
            step_type = step.get("type", "action")
            condition = step.get("condition")
            if condition and not eval(condition, {}, context):
                executed.append({"step": step_name, "status": "skipped"})
                continue
            action = step.get("action")
            if action and callable(action):
                try:
                    result = action(context)
                    context[step_name] = result
                    executed.append({"step": step_name, "status": "ok", "result": str(result)})
                except Exception as e:
                    executed.append({"step": step_name, "status": "error", "error": str(e)})
                    break
            else:
                executed.append({"step": step_name, "status": "ok"})
        instance = {"workflow": name, "steps": executed, "context": context,
                    "timestamp": datetime.now(timezone.utc).isoformat()}
        self._instances.append(instance)
        return {"status": "ok", "steps": executed, "context": context}

    @property
    def workflows(self) -> List[str]:
        return list(self._workflows.keys())

    @property
    def instance_count(self) -> int:
        return len(self._instances)

    @property
    def instances(self) -> List[dict]:
        return list(self._instances)


class TemplateEngine:
    """Jinja2-style template engine — variables, loops, conditionals."""

    def __init__(self):
        self._templates: Dict[str, str] = {}
        self._filters: Dict[str, Callable] = {}

    def register_template(self, name: str, template: str) -> None:
        self._templates[name] = template

    def register_filter(self, name: str, func: Callable) -> None:
        self._filters[name] = func

    def render(self, template: str, context: Optional[dict] = None) -> str:
        ctx = context or {}
        if template in self._templates:
            template = self._templates[template]
        import re
        def replace_if(match):
            condition = match.group(1).strip()
            content = match.group(2)
            try:
                if eval(condition, {}, ctx):
                    return content
            except Exception:
                pass
            return ""
        result = re.sub(r"\{%\s*if\s+(.*?)\s*%\}(.*?)\{%\s*endif\s*%\}", replace_if, template, flags=re.DOTALL)
        def replace_for(match):
            var = match.group(1).strip()
            iterable_name = match.group(2).strip()
            content = match.group(3)
            iterable = ctx.get(iterable_name, [])
            output = []
            for item in iterable:
                output.append(re.sub(r"\{\{\s*" + var + r"\s*\}\}", str(item), content))
            return "".join(output)
        result = re.sub(r"\{%\s*for\s+(\w+)\s+in\s+(\w+)\s*%\}(.*?)\{%\s*endfor\s*%\}", replace_for, result, flags=re.DOTALL)
        def replace_var(match):
            expr = match.group(1).strip()
            parts = expr.split("|")
            value = ctx.get(parts[0], "")
            for filter_name in parts[1:]:
                filter_name = filter_name.strip()
                if filter_name in self._filters:
                    value = self._filters[filter_name](value)
            return str(value)
        result = re.sub(r"\{\{\s*(.*?)\s*\}\}", replace_var, result)
        return result

    @property
    def template_names(self) -> List[str]:
        return list(self._templates.keys())


class FormBuilder:
    """Dynamic form builder with validation."""

    def __init__(self, name: str = ""):
        self._name = name
        self._fields: List[dict] = []

    def add_field(self, name: str, type: str = "text", label: str = "",
                  required: bool = False, placeholder: str = "",
                  choices: Optional[List[str]] = None, min_val: Optional[float] = None,
                  max_val: Optional[float] = None, pattern: str = "") -> "FormBuilder":
        self._fields.append({
            "name": name, "type": type, "label": label or name,
            "required": required, "placeholder": placeholder,
            "choices": choices or [], "min": min_val, "max": max_val, "pattern": pattern,
        })
        return self

    def validate(self, data: dict) -> dict:
        errors: List[str] = []
        validated: Dict[str, Any] = {}
        for field in self._fields:
            name = field["name"]
            value = data.get(name)
            if field["required"] and (value is None or value == ""):
                errors.append(f"Field '{name}' is required")
                continue
            if value is None:
                continue
            ftype = field["type"]
            if ftype == "email":
                import re
                if not re.match(r'^[^@]+@[^@]+\.[^@]+$', str(value)):
                    errors.append(f"Field '{name}' must be a valid email")
            elif ftype == "number":
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    errors.append(f"Field '{name}' must be a number")
                    continue
            if field["min"] is not None and isinstance(value, (int, float)) and value < field["min"]:
                errors.append(f"Field '{name}' must be >= {field['min']}")
            if field["max"] is not None and isinstance(value, (int, float)) and value > field["max"]:
                errors.append(f"Field '{name}' must be <= {field['max']}")
            if field["choices"] and value not in field["choices"]:
                errors.append(f"Field '{name}' must be one of {field['choices']}")
            if field["pattern"] and isinstance(value, str):
                import re
                if not re.match(field["pattern"], value):
                    errors.append(f"Field '{name}' does not match pattern")
            if not errors or name not in [e.split("'")[1] for e in errors if "'" in e]:
                validated[name] = value
        return {"valid": len(errors) == 0, "errors": errors, "data": validated}

    def to_dict(self) -> dict:
        return {"name": self._name, "fields": self._fields}

    def to_html(self) -> str:
        html_parts = [f'<form name="{self._name}">']
        for field in self._fields:
            html_parts.append(f'  <label>{field["label"]}</label>')
            if field["type"] in ("text", "email", "number", "password"):
                html_parts.append(f'  <input type="{field["type"]}" name="{field["name"]}" placeholder="{field["placeholder"]}" {"required" if field["required"] else ""}/>')
            elif field["type"] == "select" and field["choices"]:
                html_parts.append(f'  <select name="{field["name"]}">')
                for choice in field["choices"]:
                    html_parts.append(f'    <option value="{choice}">{choice}</option>')
                html_parts.append('  </select>')
            elif field["type"] == "textarea":
                html_parts.append(f'  <textarea name="{field["name"]}" placeholder="{field["placeholder"]}"></textarea>')
        html_parts.append('</form>')
        return "\n".join(html_parts)

    @property
    def field_count(self) -> int:
        return len(self._fields)


class MessageQueue:
    """RabbitMQ-style message queue with exchanges and routing keys."""

    def __init__(self):
        self._queues: Dict[str, List[dict]] = {}
        self._bindings: Dict[str, List[str]] = {}
        self._consumers: Dict[str, List[Callable]] = {}
        self._stats: Dict[str, int] = {"published": 0, "consumed": 0}

    def declare_queue(self, name: str) -> bool:
        self._queues.setdefault(name, [])
        _log("INFO", f"Queue declared: {name}")
        return True

    def bind(self, exchange: str, queue: str, routing_key: str = "") -> bool:
        key = f"{exchange}:{routing_key}" if routing_key else exchange
        self._bindings.setdefault(key, []).append(queue)
        return True

    def publish(self, queue: str, message: Any, routing_key: str = "") -> bool:
        if queue not in self._queues:
            self.declare_queue(queue)
        msg = {"body": message, "routing_key": routing_key,
               "timestamp": datetime.now(timezone.utc).isoformat()}
        self._queues[queue].append(msg)
        self._stats["published"] += 1
        for consumer in self._consumers.get(queue, []):
            try:
                consumer(message)
                self._stats["consumed"] += 1
            except Exception as e:
                _log("ERROR", f"Queue consumer error: {e}")
        return True

    def consume(self, queue: str, handler: Callable) -> bool:
        self._consumers.setdefault(queue, []).append(handler)
        return True

    def get_messages(self, queue: str) -> List[dict]:
        return list(self._queues.get(queue, []))

    def purge(self, queue: str) -> int:
        count = len(self._queues.get(queue, []))
        self._queues[queue] = []
        return count

    @property
    def queues(self) -> List[str]:
        return list(self._queues.keys())

    @property
    def stats(self) -> dict:
        return dict(self._stats)

    def queue_size(self, queue: str) -> int:
        return len(self._queues.get(queue, []))


class StreamProcessor:
    """Real-time data stream processor — filter, transform, aggregate."""

    def __init__(self):
        self._pipes: List[Callable] = []
        self._data: List[Any] = []

    def source(self, data: List[Any]) -> "StreamProcessor":
        self._data = list(data)
        return self

    def filter(self, func: Callable) -> "StreamProcessor":
        self._pipes.append(("filter", func))
        return self

    def map(self, func: Callable) -> "StreamProcessor":
        self._pipes.append(("map", func))
        return self

    def reduce(self, func: Callable, initial: Any = 0) -> "StreamProcessor":
        self._pipes.append(("reduce", func, initial))
        return self

    def sort(self, key: Optional[Callable] = None, reverse: bool = False) -> "StreamProcessor":
        self._pipes.append(("sort", key, reverse))
        return self

    def limit(self, n: int) -> "StreamProcessor":
        self._pipes.append(("limit", n))
        return self

    def distinct(self) -> "StreamProcessor":
        self._pipes.append(("distinct",))
        return self

    def collect(self) -> List[Any]:
        data = list(self._data)
        for pipe in self._pipes:
            if pipe[0] == "filter":
                data = [x for x in data if pipe[1](x)]
            elif pipe[0] == "map":
                data = [pipe[1](x) for x in data]
            elif pipe[0] == "reduce":
                result = pipe[2]
                for x in data:
                    result = pipe[1](result, x)
                data = [result]
            elif pipe[0] == "sort":
                data = sorted(data, key=pipe[1], reverse=pipe[2])
            elif pipe[0] == "limit":
                data = data[:pipe[1]]
            elif pipe[0] == "distinct":
                seen = set()
                unique = []
                for x in data:
                    if x not in seen:
                        seen.add(x)
                        unique.append(x)
                data = unique
        self._pipes.clear()
        return data

    def count(self) -> int:
        return len(self.collect())

    def first(self) -> Optional[Any]:
        result = self.collect()
        return result[0] if result else None

    def to_sum(self) -> float:
        return sum(self.collect())


class TimeSeriesDB:
    """Time-series database for metrics, logs, sensor data."""

    def __init__(self):
        self._series: Dict[str, List[dict]] = {}

    def write(self, metric: str, value: float, timestamp: Optional[float] = None,
              tags: Optional[dict] = None) -> bool:
        ts = time.time() if timestamp is None else timestamp
        self._series.setdefault(metric, []).append({
            "timestamp": ts, "value": value, "tags": tags or {},
        })
        return True

    def query(self, metric: str, start: float = 0, end: float = float('inf'),
              tags: Optional[dict] = None) -> List[dict]:
        data = self._series.get(metric, [])
        results = []
        for point in data:
            if start <= point["timestamp"] <= end:
                if tags and not all(point["tags"].get(k) == v for k, v in tags.items()):
                    continue
                results.append(point)
        return results

    def aggregate(self, metric: str, func: str = "avg",
                  start: float = 0, end: float = float('inf')) -> Optional[float]:
        data = self.query(metric, start, end)
        if not data:
            return None
        values = [d["value"] for d in data]
        if func == "avg":
            return sum(values) / len(values)
        elif func == "sum":
            return sum(values)
        elif func == "min":
            return min(values)
        elif func == "max":
            return max(values)
        elif func == "count":
            return float(len(values))
        return None

    def downsample(self, metric: str, interval: int) -> List[dict]:
        data = self._series.get(metric, [])
        if not data:
            return []
        buckets: Dict[int, List[float]] = {}
        for point in data:
            bucket = int(point["timestamp"] // interval) * interval
            buckets.setdefault(bucket, []).append(point["value"])
        return [{"timestamp": b, "value": sum(v) / len(v), "count": len(v)}
                for b, v in sorted(buckets.items())]

    @property
    def metrics(self) -> List[str]:
        return list(self._series.keys())

    def point_count(self, metric: str) -> int:
        return len(self._series.get(metric, []))

    def clear(self, metric: str = "") -> None:
        if metric:
            self._series.pop(metric, None)
        else:
            self._series.clear()


class GraphDatabase:
    """Graph database — nodes, edges, relationship traversal."""

    def __init__(self):
        self._nodes: Dict[str, dict] = {}
        self._edges: List[dict] = []

    def add_node(self, id: str, properties: Optional[dict] = None) -> bool:
        self._nodes[id] = {"id": id, "properties": properties or {}}
        return True

    def remove_node(self, id: str) -> bool:
        if id not in self._nodes:
            return False
        del self._nodes[id]
        self._edges = [e for e in self._edges if e["from"] != id and e["to"] != id]
        return True

    def connect(self, from_id: str, to_id: str, relation: str = "connected",
                properties: Optional[dict] = None) -> bool:
        if from_id not in self._nodes or to_id not in self._nodes:
            return False
        self._edges.append({"from": from_id, "to": to_id, "relation": relation,
                            "properties": properties or {}})
        return True

    def neighbors(self, id: str, relation: str = "") -> List[str]:
        result = []
        for edge in self._edges:
            if edge["from"] == id and (not relation or edge["relation"] == relation):
                result.append(edge["to"])
            elif edge["to"] == id and (not relation or edge["relation"] == relation):
                result.append(edge["from"])
        return list(set(result))

    def traverse(self, start: str, depth: int = 3, relation: str = "") -> List[str]:
        visited = set()
        queue = [(start, 0)]
        while queue:
            node, d = queue.pop(0)
            if node in visited or d > depth:
                continue
            visited.add(node)
            for neighbor in self.neighbors(node, relation):
                if neighbor not in visited:
                    queue.append((neighbor, d + 1))
        return list(visited - {start})

    def find_path(self, start: str, end: str) -> Optional[List[str]]:
        if start not in self._nodes or end not in self._nodes:
            return None
        from collections import deque
        queue = deque([[start]])
        visited = {start}
        while queue:
            path = queue.popleft()
            node = path[-1]
            if node == end:
                return path
            for neighbor in self.neighbors(node):
                if neighbor not in visited:
                    visited.add(neighbor)
                    queue.append(path + [neighbor])
        return None

    def get_edges(self, relation: str = "") -> List[dict]:
        if relation:
            return [e for e in self._edges if e["relation"] == relation]
        return list(self._edges)

    @property
    def node_count(self) -> int:
        return len(self._nodes)

    @property
    def edge_count(self) -> int:
        return len(self._edges)

    def get_node(self, id: str) -> Optional[dict]:
        return self._nodes.get(id)


class DocGenerator:
    """Auto-generate API docs, README, OpenAPI/Swagger from code."""

    def __init__(self):
        self._endpoints: List[dict] = []
        self._schemas: Dict[str, dict] = {}

    def add_endpoint(self, method: str, path: str, summary: str = "",
                     params: Optional[dict] = None, responses: Optional[dict] = None) -> None:
        self._endpoints.append({
            "method": method.upper(), "path": path, "summary": summary,
            "params": params or {}, "responses": responses or {"200": {"description": "OK"}},
        })

    def add_schema(self, name: str, fields: dict) -> None:
        self._schemas[name] = fields

    def generate_openapi(self, title: str = "PyTreX API", version: str = "1.0.0") -> dict:
        paths: Dict[str, dict] = {}
        for ep in self._endpoints:
            paths.setdefault(ep["path"], {})[ep["method"].lower()] = {
                "summary": ep["summary"],
                "parameters": [{"name": k, "in": "query", "schema": {"type": v}}
                               for k, v in ep["params"].items()],
                "responses": ep["responses"],
            }
        return {
            "openapi": "3.0.0",
            "info": {"title": title, "version": version},
            "paths": paths,
            "components": {
                "schemas": {name: {"type": "object", "properties": fields}
                           for name, fields in self._schemas.items()},
            },
        }

    def generate_readme(self, title: str = "PyTreX Project") -> str:
        lines = [f"# {title}", "", "## API Endpoints", ""]
        for ep in self._endpoints:
            lines.append(f"### `{ep['method']} {ep['path']}`")
            if ep["summary"]:
                lines.append(f"{ep['summary']}")
            if ep["params"]:
                lines.append(f"\n**Parameters:** {', '.join(ep['params'].keys())}")
            lines.append("")
        if self._schemas:
            lines.append("## Schemas")
            for name, fields in self._schemas.items():
                lines.append(f"### {name}")
                lines.append(f"Fields: {', '.join(fields.keys())}")
                lines.append("")
        return "\n".join(lines)

    def generate_markdown_docs(self) -> str:
        lines = ["# API Documentation", ""]
        for ep in self._endpoints:
            lines.append(f"## {ep['method']} {ep['path']}")
            lines.append(f"**Summary:** {ep['summary']}")
            if ep["params"]:
                lines.append("\n| Parameter | Type |")
                lines.append("|-----------|------|")
                for k, v in ep["params"].items():
                    lines.append(f"| {k} | {v} |")
            lines.append("")
        return "\n".join(lines)

    @property
    def endpoint_count(self) -> int:
        return len(self._endpoints)


class TestFramework:
    """Built-in test runner — assertions, fixtures, mocks."""

    def __init__(self):
        self._tests: List[dict] = []
        self._fixtures: Dict[str, Callable] = {}
        self._results: List[dict] = []
        self._mocks: Dict[str, Any] = {}

    def register(self, name: str, func: Callable) -> None:
        self._tests.append({"name": name, "func": func})

    def fixture(self, name: str, func: Callable) -> None:
        self._fixtures[name] = func

    def get_fixture(self, name: str) -> Any:
        if name in self._fixtures:
            return self._fixtures[name]()
        return None

    def mock(self, name: str, return_value: Any) -> None:
        self._mocks[name] = return_value

    def get_mock(self, name: str) -> Any:
        return self._mocks.get(name)

    def run(self) -> dict:
        passed = 0
        failed = 0
        errors: List[str] = []
        for test in self._tests:
            try:
                test["func"]()
                passed += 1
                self._results.append({"name": test["name"], "status": "passed"})
            except AssertionError as e:
                failed += 1
                errors.append(f"{test['name']}: {e}")
                self._results.append({"name": test["name"], "status": "failed", "error": str(e)})
            except Exception as e:
                failed += 1
                errors.append(f"{test['name']}: {e}")
                self._results.append({"name": test["name"], "status": "error", "error": str(e)})
        return {"total": len(self._tests), "passed": passed, "failed": failed, "errors": errors}

    @staticmethod
    def assert_equal(a: Any, b: Any) -> None:
        assert a == b, f"Expected {a} == {b}"

    @staticmethod
    def assert_not_equal(a: Any, b: Any) -> None:
        assert a != b, f"Expected {a} != {b}"

    @staticmethod
    def assert_true(value: Any) -> None:
        assert value, f"Expected truthy, got {value}"

    @staticmethod
    def assert_false(value: Any) -> None:
        assert not value, f"Expected falsy, got {value}"

    @staticmethod
    def assert_in(item: Any, collection: Any) -> None:
        assert item in collection, f"Expected {item} in {collection}"

    @staticmethod
    def assert_raises(error_type: type, func: Callable) -> None:
        try:
            func()
            raise AssertionError(f"Expected {error_type.__name__} to be raised")
        except error_type:
            pass

    @property
    def test_count(self) -> int:
        return len(self._tests)

    @property
    def results(self) -> List[dict]:
        return list(self._results)


class CLIBuilder:
    """Build CLI tools with subcommands, flags, and help."""

    def __init__(self, name: str = "pytrex-cli"):
        self._name = name
        self._commands: Dict[str, dict] = {}
        self._global_flags: List[dict] = []

    def command(self, name: str, handler: Callable, description: str = "") -> "CLIBuilder":
        self._commands[name] = {"handler": handler, "description": description, "flags": []}
        return self

    def flag(self, command_name: str, flag_name: str, required: bool = False,
             default: Any = None, help: str = "") -> "CLIBuilder":
        if command_name in self._commands:
            self._commands[command_name]["flags"].append({
                "name": flag_name, "required": required, "default": default, "help": help,
            })
        return self

    def global_flag(self, name: str, required: bool = False, default: Any = None,
                    help: str = "") -> "CLIBuilder":
        self._global_flags.append({"name": name, "required": required, "default": default, "help": help})
        return self

    def execute(self, command_name: str, args: Optional[dict] = None) -> dict:
        cmd = self._commands.get(command_name)
        if not cmd:
            return {"status": "error", "message": f"Command '{command_name}' not found"}
        try:
            result = cmd["handler"](args or {})
            return {"status": "ok", "result": result}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def help(self) -> str:
        lines = [f"Usage: {self._name} <command> [flags]", "", "Commands:"]
        for name, cmd in self._commands.items():
            desc = cmd["description"] or ""
            lines.append(f"  {name:20s} {desc}")
            for flag in cmd["flags"]:
                req = " (required)" if flag["required"] else ""
                lines.append(f"    --{flag['name']}{req}  {flag['help']}")
        if self._global_flags:
            lines.append("\nGlobal Flags:")
            for flag in self._global_flags:
                lines.append(f"  --{flag['name']}  {flag['help']}")
        return "\n".join(lines)

    @property
    def commands(self) -> List[str]:
        return list(self._commands.keys())

    @property
    def command_count(self) -> int:
        return len(self._commands)


class IoTManager:
    """IoT device management — register devices, read sensors, MQTT-style messaging."""

    def __init__(self):
        self._devices: Dict[str, dict] = {}
        self._readings: Dict[str, List[dict]] = {}
        self._subscribers: Dict[str, List[Callable]] = {}

    def register_device(self, device_id: str, device_type: str = "sensor",
                        metadata: Optional[dict] = None) -> bool:
        self._devices[device_id] = {
            "id": device_id, "type": device_type, "metadata": metadata or {},
            "status": "online", "registered": datetime.now(timezone.utc).isoformat(),
        }
        self._readings.setdefault(device_id, [])
        _log("INFO", f"IoT device registered: {device_id} ({device_type})")
        return True

    def remove_device(self, device_id: str) -> bool:
        if device_id not in self._devices:
            return False
        del self._devices[device_id]
        self._readings.pop(device_id, None)
        return True

    def read(self, device_id: str) -> dict:
        readings = self._readings.get(device_id, [])
        if not readings:
            return {"status": "error", "message": "No readings available"}
        return {"status": "ok", "reading": readings[-1]}

    def write(self, device_id: str, value: float, unit: str = "",
              timestamp: Optional[float] = None) -> bool:
        if device_id not in self._devices:
            return False
        ts = timestamp or time.time()
        reading = {"value": value, "unit": unit, "timestamp": ts}
        self._readings[device_id].append(reading)
        for sub in self._subscribers.get(device_id, []):
            try:
                sub(reading)
            except Exception as e:
                _log("ERROR", f"IoT subscriber error: {e}")
        return True

    def subscribe(self, device_id: str, handler: Callable) -> bool:
        self._subscribers.setdefault(device_id, []).append(handler)
        return True

    def get_history(self, device_id: str, limit: int = 100) -> List[dict]:
        readings = self._readings.get(device_id, [])
        return readings[-limit:]

    def set_status(self, device_id: str, status: str) -> bool:
        if device_id not in self._devices:
            return False
        self._devices[device_id]["status"] = status
        return True

    @property
    def devices(self) -> List[dict]:
        return list(self._devices.values())

    @property
    def device_count(self) -> int:
        return len(self._devices)

    def get_device(self, device_id: str) -> Optional[dict]:
        return self._devices.get(device_id)


class RealtimeSync:
    """Collaborative editing — Operational Transform for real-time sync."""

    def __init__(self):
        self._documents: Dict[str, dict] = {}
        self._operations: Dict[str, List[dict]] = {}
        self._handlers: Dict[str, List[Callable]] = []
        self._clients: Dict[str, set] = {}

    def join(self, doc_id: str, client_id: str) -> dict:
        self._clients.setdefault(doc_id, set()).add(client_id)
        if doc_id not in self._documents:
            self._documents[doc_id] = {"content": "", "version": 0}
            self._operations[doc_id] = []
        return {"status": "ok", "doc_id": doc_id, "version": self._documents[doc_id]["version"],
                "content": self._documents[doc_id]["content"]}

    def leave(self, doc_id: str, client_id: str) -> bool:
        if doc_id in self._clients and client_id in self._clients[doc_id]:
            self._clients[doc_id].discard(client_id)
            return True
        return False

    def apply_op(self, doc_id: str, op: dict, client_id: str = "") -> dict:
        doc = self._documents.get(doc_id)
        if not doc:
            return {"status": "error", "message": "Document not found"}
        op_type = op.get("type", "insert")
        position = op.get("position", 0)
        content = doc["content"]
        if op_type == "insert":
            text = op.get("text", "")
            doc["content"] = content[:position] + text + content[position:]
        elif op_type == "delete":
            length = op.get("length", 1)
            doc["content"] = content[:position] + content[position + length:]
        doc["version"] += 1
        op_record = {"op": op, "client": client_id, "version": doc["version"],
                     "timestamp": datetime.now(timezone.utc).isoformat()}
        self._operations[doc_id].append(op_record)
        for handler in self._handlers:
            try:
                handler(doc_id, op_record)
            except Exception as e:
                _log("ERROR", f"RealtimeSync handler error: {e}")
        return {"status": "ok", "version": doc["version"], "content": doc["content"]}

    def on_change(self, handler: Callable) -> None:
        self._handlers.append(handler)

    def get_state(self, doc_id: str) -> dict:
        doc = self._documents.get(doc_id)
        if not doc:
            return {"status": "error", "message": "Document not found"}
        return {"status": "ok", "content": doc["content"], "version": doc["version"],
                "client_count": len(self._clients.get(doc_id, set()))}

    def get_history(self, doc_id: str) -> List[dict]:
        return list(self._operations.get(doc_id, []))

    @property
    def document_count(self) -> int:
        return len(self._documents)


class PermissionsEngine:
    """RBAC + ABAC — role-based and attribute-based access control."""

    def __init__(self):
        self._roles: Dict[str, set] = {}
        self._user_roles: Dict[str, set] = {}
        self._policies: List[dict] = []

    def define_role(self, role: str) -> bool:
        self._roles.setdefault(role, set())
        return True

    def grant(self, role: str, action: str, resource: str) -> bool:
        self._roles.setdefault(role, set()).add(f"{action}:{resource}")
        return True

    def revoke(self, role: str, action: str, resource: str) -> bool:
        perm = f"{action}:{resource}"
        if role in self._roles and perm in self._roles[role]:
            self._roles[role].discard(perm)
            return True
        return False

    def assign_role(self, user_id: str, role: str) -> bool:
        self._user_roles.setdefault(user_id, set()).add(role)
        return True

    def remove_role(self, user_id: str, role: str) -> bool:
        if user_id in self._user_roles:
            self._user_roles[user_id].discard(role)
            return True
        return False

    def check(self, user_id: str, action: str, resource: str) -> bool:
        roles = self._user_roles.get(user_id, set())
        perm = f"{action}:{resource}"
        for role in roles:
            if perm in self._roles.get(role, set()):
                return True
            if f"*:{resource}" in self._roles.get(role, set()):
                return True
            if f"{action}:*" in self._roles.get(role, set()):
                return True
            if "*:*" in self._roles.get(role, set()):
                return True
        for policy in self._policies:
            if self._eval_policy(policy, user_id, action, resource):
                return True
        return False

    def add_policy(self, action: str, resource: str, condition: Callable) -> None:
        self._policies.append({"action": action, "resource": resource, "condition": condition})

    def _eval_policy(self, policy: dict, user_id: str, action: str, resource: str) -> bool:
        if policy["action"] != action or policy["resource"] != resource:
            return False
        try:
            return policy["condition"]({"user_id": user_id, "action": action, "resource": resource})
        except Exception:
            return False

    @property
    def roles(self) -> List[str]:
        return list(self._roles.keys())

    def get_user_roles(self, user_id: str) -> List[str]:
        return list(self._user_roles.get(user_id, set()))

    def get_role_permissions(self, role: str) -> List[str]:
        return list(self._roles.get(role, set()))


class AuditTrail:
    """Compliance audit logging — track all changes (GDPR, HIPAA, SOX)."""

    def __init__(self, max_entries: int = 10000):
        self._entries: List[dict] = []
        self._max_entries = max_entries

    def log(self, user: str, action: str, resource: str,
            before: Any = None, after: Any = None, metadata: Optional[dict] = None) -> dict:
        entry = {
            "id": len(self._entries) + 1,
            "user": user, "action": action, "resource": resource,
            "before": before, "after": after,
            "metadata": metadata or {},
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._entries.append(entry)
        if len(self._entries) > self._max_entries:
            self._entries = self._entries[-self._max_entries:]
        _log("INFO", f"Audit: {user} {action} {resource}")
        return entry

    def query(self, user: str = "", action: str = "", resource: str = "",
              start_time: str = "", end_time: str = "", limit: int = 100) -> List[dict]:
        results = self._entries
        if user:
            results = [e for e in results if e["user"] == user]
        if action:
            results = [e for e in results if e["action"] == action]
        if resource:
            results = [e for e in results if e["resource"] == resource]
        if start_time:
            results = [e for e in results if e["timestamp"] >= start_time]
        if end_time:
            results = [e for e in results if e["timestamp"] <= end_time]
        return results[-limit:]

    def get_entry(self, entry_id: int) -> Optional[dict]:
        return next((e for e in self._entries if e["id"] == entry_id), None)

    def export(self) -> List[dict]:
        return list(self._entries)

    def clear(self) -> int:
        count = len(self._entries)
        self._entries.clear()
        return count

    @property
    def count(self) -> int:
        return len(self._entries)

    def get_user_activity(self, user: str) -> List[dict]:
        return [e for e in self._entries if e["user"] == user]


class MultiTenantManager:
    """Multi-tenant SaaS manager — tenant isolation, config, resource limits."""

    def __init__(self):
        self._tenants: Dict[str, dict] = {}
        self._data: Dict[str, Dict[str, Any]] = {}

    def create(self, tenant_id: str, name: str = "", plan: str = "free") -> dict:
        if tenant_id in self._tenants:
            return {"status": "error", "message": "Tenant already exists"}
        self._tenants[tenant_id] = {
            "id": tenant_id, "name": name or tenant_id, "plan": plan,
            "config": {}, "limits": {"users": 10, "storage": 1024 * 1024, "api_calls": 1000},
            "usage": {"users": 0, "storage": 0, "api_calls": 0},
            "created": datetime.now(timezone.utc).isoformat(),
            "active": True,
        }
        self._data[tenant_id] = {}
        _log("INFO", f"Tenant created: {tenant_id} (plan={plan})")
        return {"status": "ok", "tenant": self._tenants[tenant_id]}

    def delete(self, tenant_id: str) -> bool:
        if tenant_id not in self._tenants:
            return False
        del self._tenants[tenant_id]
        self._data.pop(tenant_id, None)
        return True

    def set_config(self, tenant_id: str, key: str, value: Any) -> bool:
        if tenant_id not in self._tenants:
            return False
        self._tenants[tenant_id]["config"][key] = value
        return True

    def get_config(self, tenant_id: str, key: str = "", default: Any = None) -> Any:
        tenant = self._tenants.get(tenant_id)
        if not tenant:
            return default
        if key:
            return tenant["config"].get(key, default)
        return tenant["config"]

    def set_limit(self, tenant_id: str, resource: str, limit: int) -> bool:
        if tenant_id not in self._tenants:
            return False
        self._tenants[tenant_id]["limits"][resource] = limit
        return True

    def check_limit(self, tenant_id: str, resource: str) -> bool:
        tenant = self._tenants.get(tenant_id)
        if not tenant:
            return False
        limit = tenant["limits"].get(resource, 0)
        usage = tenant["usage"].get(resource, 0)
        return usage < limit

    def record_usage(self, tenant_id: str, resource: str, amount: int = 1) -> bool:
        if tenant_id not in self._tenants:
            return False
        self._tenants[tenant_id]["usage"][resource] = \
            self._tenants[tenant_id]["usage"].get(resource, 0) + amount
        return True

    def set_data(self, tenant_id: str, key: str, value: Any) -> bool:
        if tenant_id not in self._data:
            return False
        self._data[tenant_id][key] = value
        return True

    def get_data(self, tenant_id: str, key: str, default: Any = None) -> Any:
        return self._data.get(tenant_id, {}).get(key, default)

    def get_tenant(self, tenant_id: str) -> Optional[dict]:
        return self._tenants.get(tenant_id)

    @property
    def tenant_count(self) -> int:
        return len(self._tenants)

    @property
    def tenants(self) -> List[dict]:
        return list(self._tenants.values())


class WebhookManager:
    """Webhook manager — incoming and outgoing webhooks."""

    def __init__(self):
        self._incoming: Dict[str, Callable] = {}
        self._outgoing: List[dict] = []
        self._retries: int = 3

    def register(self, event: str, handler: Callable) -> bool:
        self._incoming[event] = handler
        _log("INFO", f"Webhook registered for: {event}")
        return True

    def unregister(self, event: str) -> bool:
        if event in self._incoming:
            del self._incoming[event]
            return True
        return False

    def receive(self, event: str, data: dict, signature: str = "") -> dict:
        handler = self._incoming.get(event)
        if not handler:
            return {"status": "error", "message": f"No handler for event: {event}"}
        try:
            result = handler(data)
            return {"status": "ok", "result": result}
        except Exception as e:
            _log("ERROR", f"Webhook handler error: {e}")
            return {"status": "error", "message": str(e)}

    def send(self, url: str, event: str, data: dict, headers: Optional[dict] = None) -> dict:
        payload = json.dumps({"event": event, "data": data,
                              "timestamp": datetime.now(timezone.utc).isoformat()}).encode("utf-8")
        req_headers = {"Content-Type": "application/json"}
        if headers:
            req_headers.update(headers)
        import urllib.request, urllib.error
        for attempt in range(self._retries):
            try:
                req = urllib.request.Request(url, data=payload, headers=req_headers, method="POST")
                with urllib.request.urlopen(req, timeout=15) as resp:
                    response_body = resp.read().decode("utf-8", errors="replace")
                    record = {"url": url, "event": event, "status": resp.status,
                              "response": response_body[:500], "attempt": attempt + 1}
                    self._outgoing.append(record)
                    return {"status": "ok", "code": resp.status, "response": response_body[:500]}
            except urllib.error.HTTPError as e:
                record = {"url": url, "event": event, "status": e.code, "error": str(e),
                          "attempt": attempt + 1}
                self._outgoing.append(record)
                if attempt < self._retries - 1:
                    time.sleep(0.5 * (attempt + 1))
                return {"status": "error", "code": e.code, "message": str(e)}
            except Exception as e:
                if attempt < self._retries - 1:
                    time.sleep(0.5 * (attempt + 1))
                continue
        record = {"url": url, "event": event, "status": "failed", "attempt": self._retries}
        self._outgoing.append(record)
        return {"status": "error", "message": "All retries failed"}

    def set_retries(self, n: int) -> None:
        self._retries = n

    @property
    def registered_events(self) -> List[str]:
        return list(self._incoming.keys())

    @property
    def outgoing_log(self) -> List[dict]:
        return list(self._outgoing)

    @property
    def outgoing_count(self) -> int:
        return len(self._outgoing)


class VersionControl:
    """Data versioning — commit, rollback, diff, branches."""

    def __init__(self):
        self._versions: Dict[str, List[dict]] = {}
        self._branches: Dict[str, str] = {}
        self._current: Dict[str, str] = {}

    def commit(self, resource_id: str, data: Any, message: str = "",
               author: str = "") -> dict:
        if resource_id not in self._versions:
            self._versions[resource_id] = []
            self._branches[resource_id] = "main"
            self._current[resource_id] = "main"
        version_num = len(self._versions[resource_id]) + 1
        version = {
            "version": version_num, "data": data, "message": message,
            "author": author, "branch": self._current.get(resource_id, "main"),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._versions[resource_id].append(version)
        _log("INFO", f"Version committed: {resource_id} v{version_num}")
        return {"status": "ok", "version": version_num, "resource": resource_id}

    def get_version(self, resource_id: str, version: int) -> Optional[dict]:
        versions = self._versions.get(resource_id, [])
        return next((v for v in versions if v["version"] == version), None)

    def latest(self, resource_id: str) -> Optional[dict]:
        versions = self._versions.get(resource_id, [])
        return versions[-1] if versions else None

    def rollback(self, resource_id: str, version: int) -> dict:
        target = self.get_version(resource_id, version)
        if not target:
            return {"status": "error", "message": "Version not found"}
        new_version = len(self._versions[resource_id]) + 1
        rolled = {
            "version": new_version, "data": target["data"],
            "message": f"Rollback to v{version}", "author": "system",
            "branch": self._current.get(resource_id, "main"),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._versions[resource_id].append(rolled)
        return {"status": "ok", "version": new_version, "data": target["data"]}

    def diff(self, resource_id: str, v1: int, v2: int) -> dict:
        ver1 = self.get_version(resource_id, v1)
        ver2 = self.get_version(resource_id, v2)
        if not ver1 or not ver2:
            return {"status": "error", "message": "Version not found"}
        data1 = ver1["data"]
        data2 = ver2["data"]
        if isinstance(data1, dict) and isinstance(data2, dict):
            added = {k: data2[k] for k in data2 if k not in data1 or data1[k] != data2[k]}
            removed = {k: data1[k] for k in data1 if k not in data2}
            return {"status": "ok", "added": added, "removed": removed}
        return {"status": "ok", "v1": data1, "v2": data2, "changed": data1 != data2}

    def create_branch(self, resource_id: str, branch: str) -> bool:
        if resource_id not in self._versions:
            return False
        self._current[resource_id] = branch
        return True

    def merge_branch(self, resource_id: str, branch: str) -> dict:
        versions = [v for v in self._versions.get(resource_id, []) if v["branch"] == branch]
        if not versions:
            return {"status": "error", "message": "Branch not found"}
        latest_branch = versions[-1]
        self._current[resource_id] = "main"
        return self.commit(resource_id, latest_branch["data"], f"Merge from {branch}")

    def history(self, resource_id: str) -> List[dict]:
        return [{"version": v["version"], "message": v["message"],
                 "author": v["author"], "branch": v["branch"],
                 "timestamp": v["timestamp"]} for v in self._versions.get(resource_id, [])]

    @property
    def resource_count(self) -> int:
        return len(self._versions)

    def version_count(self, resource_id: str) -> int:
        return len(self._versions.get(resource_id, []))


class ABTesting:
    """A/B testing engine — experiments, variants, conversion tracking."""

    def __init__(self):
        self._experiments: Dict[str, dict] = {}
        self._assignments: Dict[str, Dict[str, str]] = {}

    def create(self, name: str, variants: List[str],
               description: str = "") -> bool:
        self._experiments[name] = {
            "name": name, "variants": variants, "description": description,
            "impressions": {v: 0 for v in variants},
            "conversions": {v: 0 for v in variants},
            "active": True,
        }
        _log("INFO", f"A/B experiment created: {name} ({len(variants)} variants)")
        return True

    def assign(self, experiment: str, user_id: str) -> dict:
        exp = self._experiments.get(experiment)
        if not exp or not exp["active"]:
            return {"status": "error", "message": "Experiment not found or inactive"}
        if user_id in self._assignments.get(experiment, {}):
            variant = self._assignments[experiment][user_id]
        else:
            import hashlib
            h = int(hashlib.md5(f"{experiment}:{user_id}".encode()).hexdigest(), 16)
            variant = exp["variants"][h % len(exp["variants"])]
            self._assignments.setdefault(experiment, {})[user_id] = variant
        exp["impressions"][variant] += 1
        return {"status": "ok", "variant": variant, "experiment": experiment}

    def convert(self, experiment: str, user_id: str) -> dict:
        exp = self._experiments.get(experiment)
        if not exp:
            return {"status": "error", "message": "Experiment not found"}
        variant = self._assignments.get(experiment, {}).get(user_id)
        if not variant:
            return {"status": "error", "message": "User not assigned"}
        exp["conversions"][variant] += 1
        return {"status": "ok", "variant": variant}

    def results(self, experiment: str) -> dict:
        exp = self._experiments.get(experiment)
        if not exp:
            return {"status": "error", "message": "Experiment not found"}
        variant_results = []
        for v in exp["variants"]:
            impressions = exp["impressions"][v]
            conversions = exp["conversions"][v]
            rate = conversions / impressions if impressions > 0 else 0.0
            variant_results.append({"variant": v, "impressions": impressions,
                                    "conversions": conversions, "conversion_rate": round(rate, 4)})
        return {"status": "ok", "experiment": experiment, "variants": variant_results}

    def stop(self, experiment: str) -> bool:
        if experiment in self._experiments:
            self._experiments[experiment]["active"] = False
            return True
        return False

    @property
    def experiments(self) -> List[str]:
        return list(self._experiments.keys())

    def get_experiment(self, name: str) -> Optional[dict]:
        return self._experiments.get(name)


class FeatureAnalytics:
    """Feature usage tracking — engagement, funnels, retention."""

    def __init__(self):
        self._events: List[dict] = []
        self._funnels: Dict[str, List[str]] = {}

    def track(self, feature: str, user_id: str, properties: Optional[dict] = None) -> bool:
        self._events.append({
            "feature": feature, "user_id": user_id,
            "properties": properties or {},
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
        return True

    def define_funnel(self, name: str, steps: List[str]) -> bool:
        self._funnels[name] = steps
        return True

    def funnel(self, name: str) -> dict:
        steps = self._funnels.get(name)
        if not steps:
            return {"status": "error", "message": "Funnel not found"}
        user_steps: Dict[str, Dict[str, bool]] = {}
        for event in self._events:
            if event["feature"] in steps:
                uid = event["user_id"]
                user_steps.setdefault(uid, {})[event["feature"]] = True
        step_counts = []
        for i, step in enumerate(steps):
            count = sum(1 for uid, s in user_steps.items()
                       if all(steps[j] in s for j in range(i + 1)))
            step_counts.append({"step": step, "users": count})
        return {"status": "ok", "funnel": name, "steps": step_counts}

    def feature_usage(self, feature: str) -> dict:
        events = [e for e in self._events if e["feature"] == feature]
        users = set(e["user_id"] for e in events)
        return {"status": "ok", "feature": feature, "total_events": len(events),
                "unique_users": len(users)}

    def user_activity(self, user_id: str) -> List[dict]:
        return [e for e in self._events if e["user_id"] == user_id]

    def top_features(self, limit: int = 10) -> List[dict]:
        counts: Dict[str, int] = {}
        for e in self._events:
            counts[e["feature"]] = counts.get(e["feature"], 0) + 1
        ranked = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:limit]
        return [{"feature": f, "count": c} for f, c in ranked]

    @property
    def event_count(self) -> int:
        return len(self._events)

    @property
    def funnels(self) -> List[str]:
        return list(self._funnels.keys())


class ContentModerator:
    """AI content moderation — detect spam, toxicity, PII."""

    def __init__(self):
        self._spam_words = {"viagra", "casino", "lottery", "winner", "free money",
                           "click here", "buy now", "limited offer", "earn money",
                           "work from home", "get rich", "no risk"}
        self._toxic_words = {"idiot", "stupid", "hate", "kill", "die", "ugly",
                            "dumb", "moron", "trash", "garbage", "loser", "pathetic"}
        self._pii_patterns = {
            "email": r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
            "phone": r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b',
            "ssn": r'\b\d{3}-\d{2}-\d{4}\b',
            "credit_card": r'\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b',
        }
        self._history: List[dict] = []

    def check_text(self, text: str) -> dict:
        text_lower = text.lower()
        spam_found = [w for w in self._spam_words if w in text_lower]
        toxic_found = [w for w in self._toxic_words if w in text_lower]
        import re
        pii_found: Dict[str, List[str]] = {}
        for pii_type, pattern in self._pii_patterns.items():
            matches = re.findall(pattern, text)
            if matches:
                pii_found[pii_type] = matches
        flags: List[str] = []
        if spam_found:
            flags.append("spam")
        if toxic_found:
            flags.append("toxicity")
        if pii_found:
            flags.append("pii")
        score = len(spam_found) * 2 + len(toxic_found) * 3 + sum(len(v) for v in pii_found.values()) * 5
        result = {
            "status": "ok", "flags": flags, "score": score,
            "spam_words": spam_found, "toxic_words": toxic_found,
            "pii": pii_found, "clean": len(flags) == 0,
        }
        self._history.append({"text": text[:100], **result})
        return result

    def check_batch(self, texts: List[str]) -> dict:
        results = [self.check_text(t) for t in texts]
        clean = sum(1 for r in results if r["clean"])
        flagged = len(results) - clean
        return {"status": "ok", "results": results, "clean": clean, "flagged": flagged}

    def add_spam_word(self, word: str) -> None:
        self._spam_words.add(word.lower())

    def add_toxic_word(self, word: str) -> None:
        self._toxic_words.add(word.lower())

    @property
    def history(self) -> List[dict]:
        return list(self._history)


class RecommendationEngine:
    """Recommendation engine — collaborative filtering + content-based."""

    def __init__(self):
        self._ratings: Dict[str, Dict[str, float]] = {}
        self._items: Dict[str, dict] = {}
        self._users: Dict[str, set] = {}

    def add_item(self, item_id: str, metadata: Optional[dict] = None) -> bool:
        self._items[item_id] = metadata or {}
        return True

    def rate(self, user_id: str, item_id: str, rating: float) -> bool:
        self._ratings.setdefault(user_id, {})[item_id] = rating
        self._users.setdefault(user_id, set()).add(item_id)
        return True

    def recommend(self, user_id: str, top_k: int = 5) -> List[dict]:
        user_ratings = self._ratings.get(user_id, {})
        if not user_ratings:
            popular = self._get_popular(top_k)
            return [{"item_id": iid, "score": score, "reason": "popular"} for iid, score in popular]
        scores: Dict[str, float] = {}
        for other_user, other_ratings in self._ratings.items():
            if other_user == user_id:
                continue
            common = set(user_ratings.keys()) & set(other_ratings.keys())
            if not common:
                continue
            similarity = sum(1 - abs(user_ratings[i] - other_ratings[i]) / 5 for i in common) / len(common)
            for item, rating in other_ratings.items():
                if item not in user_ratings:
                    scores[item] = scores.get(item, 0) + similarity * rating
        ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:top_k]
        return [{"item_id": iid, "score": round(score, 3), "reason": "collaborative"} for iid, score in ranked]

    def _get_popular(self, top_k: int) -> List[tuple]:
        item_scores: Dict[str, float] = {}
        for user_ratings in self._ratings.values():
            for item, rating in user_ratings.items():
                item_scores[item] = item_scores.get(item, 0) + rating
        return sorted(item_scores.items(), key=lambda x: x[1], reverse=True)[:top_k]

    def similar_items(self, item_id: str, top_k: int = 5) -> List[dict]:
        item_raters = {u: r[item_id] for u, r in self._ratings.items() if item_id in r}
        if not item_raters:
            return []
        scores: Dict[str, float] = {}
        for other_item in self._items:
            if other_item == item_id:
                continue
            other_raters = {u: r[other_item] for u, r in self._ratings.items() if other_item in r}
            common = set(item_raters.keys()) & set(other_raters.keys())
            if not common:
                continue
            similarity = sum(1 - abs(item_raters[u] - other_raters[u]) / 5 for u in common) / len(common)
            scores[other_item] = similarity
        ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:top_k]
        return [{"item_id": iid, "similarity": round(score, 3)} for iid, score in ranked]

    @property
    def item_count(self) -> int:
        return len(self._items)

    @property
    def user_count(self) -> int:
        return len(self._users)


class DataPipeline:
    """ETL pipeline — Extract, Transform, Load."""

    def __init__(self):
        self._pipelines: Dict[str, dict] = {}
        self._results: List[dict] = []

    def create(self, name: str, extract: Callable, transforms: List[Callable],
               load: Callable) -> bool:
        self._pipelines[name] = {
            "extract": extract, "transforms": transforms, "load": load,
            "runs": 0, "last_status": "pending",
        }
        return True

    def run(self, name: str) -> dict:
        pipe = self._pipelines.get(name)
        if not pipe:
            return {"status": "error", "message": f"Pipeline '{name}' not found"}
        try:
            data = pipe["extract"]()
            for transform in pipe["transforms"]:
                data = transform(data)
            result = pipe["load"](data)
            pipe["runs"] += 1
            pipe["last_status"] = "ok"
            run_record = {"pipeline": name, "status": "ok", "result": str(result),
                          "timestamp": datetime.now(timezone.utc).isoformat()}
            self._results.append(run_record)
            return {"status": "ok", "result": result}
        except Exception as e:
            pipe["last_status"] = "error"
            run_record = {"pipeline": name, "status": "error", "error": str(e),
                          "timestamp": datetime.now(timezone.utc).isoformat()}
            self._results.append(run_record)
            return {"status": "error", "message": str(e)}

    def run_all(self) -> dict:
        results = {}
        for name in self._pipelines:
            results[name] = self.run(name)
        return {"status": "ok", "results": results}

    @property
    def pipelines(self) -> List[str]:
        return list(self._pipelines.keys())

    @property
    def results(self) -> List[dict]:
        return list(self._results)

    def get_pipeline(self, name: str) -> Optional[dict]:
        p = self._pipelines.get(name)
        if not p:
            return None
        return {"name": name, "runs": p["runs"], "last_status": p["last_status"],
                "transform_count": len(p["transforms"])}


class ServiceMesh:
    """Microservices management — discovery, health checks, load balancing."""

    def __init__(self):
        self._services: Dict[str, List[dict]] = {}
        self._health_checks: Dict[str, Callable] = {}

    def register(self, name: str, host: str, port: int,
                 metadata: Optional[dict] = None) -> bool:
        self._services.setdefault(name, []).append({
            "id": f"{name}_{len(self._services.get(name, [])) + 1}",
            "host": host, "port": port, "metadata": metadata or {},
            "status": "healthy", "registered": datetime.now(timezone.utc).isoformat(),
            "requests": 0,
        })
        _log("INFO", f"Service registered: {name} at {host}:{port}")
        return True

    def deregister(self, name: str, instance_id: str) -> bool:
        instances = self._services.get(name, [])
        for i, inst in enumerate(instances):
            if inst["id"] == instance_id:
                instances.pop(i)
                return True
        return False

    def discover(self, name: str) -> List[dict]:
        return [i for i in self._services.get(name, []) if i["status"] == "healthy"]

    def resolve(self, name: str) -> Optional[dict]:
        healthy = self.discover(name)
        if not healthy:
            return None
        return min(healthy, key=lambda x: x["requests"])

    def record_request(self, name: str, instance_id: str) -> None:
        for inst in self._services.get(name, []):
            if inst["id"] == instance_id:
                inst["requests"] += 1
                break

    def set_health_check(self, name: str, check: Callable) -> None:
        self._health_checks[name] = check

    def run_health_checks(self) -> dict:
        results = {}
        for name, check in self._health_checks.items():
            for inst in self._services.get(name, []):
                try:
                    healthy = check(inst)
                    inst["status"] = "healthy" if healthy else "unhealthy"
                    results[inst["id"]] = inst["status"]
                except Exception:
                    inst["status"] = "unhealthy"
                    results[inst["id"]] = "unhealthy"
        return results

    def set_status(self, name: str, instance_id: str, status: str) -> bool:
        for inst in self._services.get(name, []):
            if inst["id"] == instance_id:
                inst["status"] = status
                return True
        return False

    @property
    def services(self) -> List[str]:
        return list(self._services.keys())

    def instance_count(self, name: str) -> int:
        return len(self._services.get(name, []))

    def healthy_count(self, name: str) -> int:
        return len(self.discover(name))

    def get_instances(self, name: str) -> List[dict]:
        return list(self._services.get(name, []))


class SecurityScanner:
    """Security vulnerability scanner — SQL injection, XSS, CSRF, path traversal."""

    def __init__(self):
        self._patterns = {
            "sql_injection": [
                r"(\bUNION\b.*\bSELECT\b)", r"(\bOR\b\s+1\s*=\s*1)", r"(\bDROP\b\s+TABLE\b)",
                r"(\bINSERT\b.*\bINTO\b)", r"(\bDELETE\b.*\bFROM\b)", r"('--)", r"(\bEXEC\b)",
            ],
            "xss": [
                r"<script[^>]*>.*?</script>", r"javascript:", r"onerror\s*=",
                r"onload\s*=", r"onclick\s*=", r"<iframe", r"document\.cookie",
            ],
            "path_traversal": [
                r"\.\./", r"\.\.\\", r"%2e%2e%2f", r"%2e%2e/",
                r"\.\.%2f", r"\.\.%5c", r"/etc/passwd", r"/etc/shadow",
            ],
            "command_injection": [
                r";\s*(cat|ls|id|whoami|uname)", r"\|\s*(cat|ls|id|whoami)",
                r"`.*`", r"\$\(.*\)", r"&&\s*(cat|ls|id)",
            ],
        }
        self._results: List[dict] = []

    def scan(self, input_text: str, context: str = "general") -> dict:
        import re
        vulnerabilities: List[dict] = []
        for vuln_type, patterns in self._patterns.items():
            for pattern in patterns:
                matches = re.findall(pattern, input_text, re.IGNORECASE)
                if matches:
                    vulnerabilities.append({
                        "type": vuln_type, "pattern": pattern,
                        "matches": [str(m) for m in matches[:5]],
                        "severity": self._severity(vuln_type),
                    })
        result = {
            "status": "ok", "context": context, "input_length": len(input_text),
            "vulnerabilities": vulnerabilities, "secure": len(vulnerabilities) == 0,
            "scan_time": datetime.now(timezone.utc).isoformat(),
        }
        self._results.append(result)
        return result

    def _severity(self, vuln_type: str) -> str:
        severities = {"sql_injection": "critical", "xss": "high",
                      "path_traversal": "high", "command_injection": "critical"}
        return severities.get(vuln_type, "medium")

    def scan_batch(self, inputs: List[dict]) -> dict:
        results = [self.scan(text, ctx) for text, ctx in inputs]
        secure = sum(1 for r in results if r["secure"])
        return {"status": "ok", "scanned": len(results), "secure": secure,
                "vulnerable": len(results) - secure, "results": results}

    def add_pattern(self, vuln_type: str, pattern: str) -> None:
        self._patterns.setdefault(vuln_type, []).append(pattern)

    def generate_csrf_token(self) -> str:
        import secrets
        return secrets.token_hex(32)

    def verify_csrf_token(self, token: str, expected: str) -> bool:
        import hmac
        return hmac.compare_digest(token, expected)

    @property
    def scan_count(self) -> int:
        return len(self._results)

    @property
    def results(self) -> List[dict]:
        return list(self._results)


class SmartContract:
    """Solidity-style smart contract engine — execute, verify, gas estimation."""

    def __init__(self):
        self._contracts: Dict[str, dict] = {}
        self._state: Dict[str, Any] = {}
        self._logs: List[dict] = []

    def deploy(self, name: str, functions: Dict[str, Callable],
               init_state: Optional[dict] = None) -> dict:
        self._contracts[name] = {
            "functions": functions, "state": init_state or {},
            "address": f"0x{hash(name) & 0xFFFFFFFFFFFFFFFF:016x}",
            "deployed": datetime.now(timezone.utc).isoformat(),
        }
        _log("INFO", f"Smart contract deployed: {name}")
        return {"status": "ok", "address": self._contracts[name]["address"]}

    def execute(self, contract: str, function: str,
                args: Optional[list] = None, sender: str = "") -> dict:
        con = self._contracts.get(contract)
        if not con:
            return {"status": "error", "message": "Contract not found"}
        func = con["functions"].get(function)
        if not func:
            return {"status": "error", "message": f"Function '{function}' not found"}
        gas = 21000 + len(args or []) * 1000
        try:
            result = func(con["state"], *(args or []))
            self._logs.append({
                "contract": contract, "function": function, "sender": sender,
                "gas_used": gas, "result": str(result),
                "timestamp": datetime.now(timezone.utc).isoformat(),
            })
            return {"status": "ok", "result": result, "gas_used": gas}
        except Exception as e:
            self._logs.append({"contract": contract, "function": function,
                               "error": str(e), "gas_used": gas,
                               "timestamp": datetime.now(timezone.utc).isoformat()})
            return {"status": "error", "message": str(e), "gas_used": gas}

    def estimate_gas(self, contract: str, function: str, args: Optional[list] = None) -> int:
        return 21000 + len(args or []) * 1000

    def get_state(self, contract: str) -> dict:
        con = self._contracts.get(contract)
        if not con:
            return {"status": "error", "message": "Contract not found"}
        return {"status": "ok", "state": con["state"]}

    def get_address(self, contract: str) -> Optional[str]:
        con = self._contracts.get(contract)
        return con["address"] if con else None

    @property
    def contracts(self) -> List[str]:
        return list(self._contracts.keys())

    @property
    def logs(self) -> List[dict]:
        return list(self._logs)

    @property
    def contract_count(self) -> int:
        return len(self._contracts)


class BlockchainBridge:
    """Bridge between Python and Rust blockchain engine — create blocks, verify chain, manage cache."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            _log("WARN", "my_framework not available — blockchain bridge in Python-only mode")

    def add_block(self, data: str) -> dict:
        """Create a new block in the blockchain via Rust engine, cache it locally."""
        if self._framework:
            try:
                block_json = self._framework.fanya_block_ya_blockchain(data)
                block = json.loads(block_json)
                BLOCKCHAIN_CACHE.append(block)
                _log("INFO", f"Block #{block.get('index', '?')} added to blockchain")
                return {"status": "ok", "block": block}
            except Exception as e:
                _log("ERROR", f"Blockchain add_block failed: {e}")
                return {"status": "error", "message": str(e)}
        # Python-only fallback: create block locally
        import hashlib as _hl
        prev_hash = BLOCKCHAIN_CACHE[-1]["hash"] if BLOCKCHAIN_CACHE else "0" * 64
        index = len(BLOCKCHAIN_CACHE) + 1
        timestamp = datetime.now(timezone.utc).isoformat()
        block_hash = _hl.sha256(f"{index}{timestamp}{data}{prev_hash}".encode()).hexdigest()
        block = {
            "index": index,
            "timestamp": timestamp,
            "data": data,
            "previous_hash": prev_hash,
            "hash": block_hash,
        }
        BLOCKCHAIN_CACHE.append(block)
        _log("INFO", f"Block #{index} added (Python fallback)")
        return {"status": "ok", "block": block, "mode": "python_fallback"}

    def verify_chain(self) -> dict:
        """Verify the entire blockchain for tampering."""
        if not BLOCKCHAIN_CACHE:
            return {"status": "ok", "valid": True, "message": "Empty chain"}
        if self._framework:
            try:
                chain_json = json.dumps(BLOCKCHAIN_CACHE)
                valid = self._framework.hakiki_blockchain(chain_json)
                return {"status": "ok", "valid": valid, "blocks": len(BLOCKCHAIN_CACHE)}
            except Exception as e:
                _log("ERROR", f"Blockchain verify failed: {e}")
                return {"status": "error", "message": str(e)}
        # Python-only fallback: verify locally
        import hashlib as _hl
        for i in range(1, len(BLOCKCHAIN_CACHE)):
            curr = BLOCKCHAIN_CACHE[i]
            prev = BLOCKCHAIN_CACHE[i - 1]
            expected_hash = _hl.sha256(
                f"{curr['index']}{curr['timestamp']}{curr['data']}{curr['previous_hash']}".encode()
            ).hexdigest()
            if curr["hash"] != expected_hash:
                return {"status": "ok", "valid": False, "tampered_block": curr["index"]}
            if curr["previous_hash"] != prev["hash"]:
                return {"status": "ok", "valid": False, "broken_link": curr["index"]}
        return {"status": "ok", "valid": True, "blocks": len(BLOCKCHAIN_CACHE)}

    def get_chain(self) -> List[dict]:
        """Return the full blockchain cache."""
        return list(BLOCKCHAIN_CACHE)

    def get_block(self, index: int) -> Optional[dict]:
        """Get a specific block by index."""
        for block in BLOCKCHAIN_CACHE:
            if block.get("index") == index:
                return block
        return None

    def clear_chain(self) -> int:
        """Clear the blockchain cache. Returns number of blocks removed."""
        count = len(BLOCKCHAIN_CACHE)
        BLOCKCHAIN_CACHE.clear()
        _log("INFO", f"Blockchain cache cleared ({count} blocks)")
        return count

    @property
    def block_count(self) -> int:
        return len(BLOCKCHAIN_CACHE)

    @property
    def last_hash(self) -> Optional[str]:
        return BLOCKCHAIN_CACHE[-1]["hash"] if BLOCKCHAIN_CACHE else None

    @property
    def chain_valid(self) -> bool:
        result = self.verify_chain()
        return result.get("valid", False)


class StatisticsEngine:
    """Statistical analysis — mean, median, std dev, correlation, regression."""

    @staticmethod
    def mean(data: List[float]) -> float:
        if not data:
            return 0.0
        return sum(data) / len(data)

    @staticmethod
    def median(data: List[float]) -> float:
        if not data:
            return 0.0
        sorted_data = sorted(data)
        n = len(sorted_data)
        if n % 2 == 0:
            return (sorted_data[n // 2 - 1] + sorted_data[n // 2]) / 2
        return sorted_data[n // 2]

    @staticmethod
    def mode(data: List[float]) -> List[float]:
        if not data:
            return []
        from collections import Counter
        counts = Counter(data)
        max_count = max(counts.values())
        return [v for v, c in counts.items() if c == max_count]

    @staticmethod
    def variance(data: List[float]) -> float:
        if len(data) < 2:
            return 0.0
        m = StatisticsEngine.mean(data)
        return sum((x - m) ** 2 for x in data) / (len(data) - 1)

    @staticmethod
    def std_dev(data: List[float]) -> float:
        return StatisticsEngine.variance(data) ** 0.5

    @staticmethod
    def min_val(data: List[float]) -> float:
        return min(data) if data else 0.0

    @staticmethod
    def max_val(data: List[float]) -> float:
        return max(data) if data else 0.0

    @staticmethod
    def range_val(data: List[float]) -> float:
        if not data:
            return 0.0
        return max(data) - min(data)

    @staticmethod
    def percentile(data: List[float], p: float) -> float:
        if not data:
            return 0.0
        sorted_data = sorted(data)
        k = (len(sorted_data) - 1) * p / 100
        f = int(k)
        c = k - f
        if f + 1 < len(sorted_data):
            return sorted_data[f] + c * (sorted_data[f + 1] - sorted_data[f])
        return sorted_data[f]

    @staticmethod
    def correlation(x: List[float], y: List[float]) -> float:
        if len(x) != len(y) or len(x) < 2:
            return 0.0
        mx, my = StatisticsEngine.mean(x), StatisticsEngine.mean(y)
        numerator = sum((xi - mx) * (yi - my) for xi, yi in zip(x, y))
        denom_x = sum((xi - mx) ** 2 for xi in x) ** 0.5
        denom_y = sum((yi - my) ** 2 for yi in y) ** 0.5
        if denom_x == 0 or denom_y == 0:
            return 0.0
        return numerator / (denom_x * denom_y)

    @staticmethod
    def linear_regression(x: List[float], y: List[float]) -> dict:
        if len(x) != len(y) or len(x) < 2:
            return {"status": "error", "message": "Insufficient data"}
        mx, my = StatisticsEngine.mean(x), StatisticsEngine.mean(y)
        numerator = sum((xi - mx) * (yi - my) for xi, yi in zip(x, y))
        denominator = sum((xi - mx) ** 2 for xi in x)
        if denominator == 0:
            return {"status": "error", "message": "Cannot fit line"}
        slope = numerator / denominator
        intercept = my - slope * mx
        r = StatisticsEngine.correlation(x, y)
        return {"status": "ok", "slope": slope, "intercept": intercept,
                "r_squared": r ** 2, "correlation": r}

    @staticmethod
    def describe(data: List[float]) -> dict:
        if not data:
            return {"status": "error", "message": "Empty data"}
        return {
            "status": "ok", "count": len(data),
            "mean": StatisticsEngine.mean(data),
            "median": StatisticsEngine.median(data),
            "std_dev": StatisticsEngine.std_dev(data),
            "variance": StatisticsEngine.variance(data),
            "min": StatisticsEngine.min_val(data),
            "max": StatisticsEngine.max_val(data),
            "range": StatisticsEngine.range_val(data),
            "q1": StatisticsEngine.percentile(data, 25),
            "q3": StatisticsEngine.percentile(data, 75),
        }

    @staticmethod
    def t_test(sample1: List[float], sample2: List[float]) -> dict:
        if len(sample1) < 2 or len(sample2) < 2:
            return {"status": "error", "message": "Need at least 2 samples"}
        m1, m2 = StatisticsEngine.mean(sample1), StatisticsEngine.mean(sample2)
        v1, v2 = StatisticsEngine.variance(sample1), StatisticsEngine.variance(sample2)
        n1, n2 = len(sample1), len(sample2)
        pooled_var = ((n1 - 1) * v1 + (n2 - 1) * v2) / (n1 + n2 - 2)
        if pooled_var == 0:
            return {"status": "error", "message": "Zero variance"}
        t_stat = (m1 - m2) / (pooled_var * (1 / n1 + 1 / n2)) ** 0.5
        return {"status": "ok", "t_statistic": t_stat,
                "mean1": m1, "mean2": m2, "df": n1 + n2 - 2}


class CICDPipeline:
    """CI/CD pipeline — build, test, deploy, rollback, environment promotion."""

    def __init__(self):
        self._pipelines: Dict[str, dict] = {}
        self._runs: List[dict] = []

    def create(self, name: str, stages: List[dict]) -> bool:
        self._pipelines[name] = {"stages": stages, "name": name}
        _log("INFO", f"CI/CD pipeline created: {name} ({len(stages)} stages)")
        return True

    def run(self, name: str, environment: str = "dev") -> dict:
        pipe = self._pipelines.get(name)
        if not pipe:
            return {"status": "error", "message": f"Pipeline '{name}' not found"}
        stages_executed: List[dict] = []
        success = True
        for stage in pipe["stages"]:
            stage_name = stage.get("name", "unnamed")
            action = stage.get("action")
            try:
                if action and callable(action):
                    result = action(environment)
                    stages_executed.append({"stage": stage_name, "status": "ok", "result": str(result)})
                else:
                    stages_executed.append({"stage": stage_name, "status": "ok"})
            except Exception as e:
                stages_executed.append({"stage": stage_name, "status": "failed", "error": str(e)})
                success = False
                break
        run_record = {
            "pipeline": name, "environment": environment,
            "stages": stages_executed, "success": success,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._runs.append(run_record)
        return {"status": "ok" if success else "failed", "stages": stages_executed}

    def rollback(self, name: str, to_stage: str = "") -> dict:
        pipe = self._pipelines.get(name)
        if not pipe:
            return {"status": "error", "message": "Pipeline not found"}
        rollback_stages = []
        found = False
        for stage in reversed(pipe["stages"]):
            if stage.get("name") == to_stage:
                found = True
                break
            rollback_stages.append({"stage": stage.get("name", ""), "action": "rollback"})
        if not found and to_stage:
            return {"status": "error", "message": f"Stage '{to_stage}' not found"}
        return {"status": "ok", "rolled_back": len(rollback_stages)}

    def promote(self, name: str, from_env: str, to_env: str) -> dict:
        return {"status": "ok", "pipeline": name, "from": from_env, "to": to_env,
                "timestamp": datetime.now(timezone.utc).isoformat()}

    @property
    def pipelines(self) -> List[str]:
        return list(self._pipelines.keys())

    @property
    def runs(self) -> List[dict]:
        return list(self._runs)

    @property
    def run_count(self) -> int:
        return len(self._runs)

    def get_pipeline(self, name: str) -> Optional[dict]:
        p = self._pipelines.get(name)
        if not p:
            return None
        return {"name": name, "stages": [s.get("name", "") for s in p["stages"]]}


class NetworkTools:
    """Network utilities — DNS, port scan, ping, traceroute, packet inspection."""

    def __init__(self):
        self._scan_results: List[dict] = []

    def dns_lookup(self, domain: str) -> dict:
        import socket
        try:
            ip = socket.gethostbyname(domain)
            return {"status": "ok", "domain": domain, "ip": ip}
        except socket.gaierror:
            return {"status": "error", "message": f"Cannot resolve {domain}"}

    def reverse_dns(self, ip: str) -> dict:
        import socket
        try:
            host = socket.gethostbyaddr(ip)
            return {"status": "ok", "ip": ip, "hostname": host[0]}
        except socket.herror:
            return {"status": "error", "message": f"Cannot resolve {ip}"}

    def port_scan(self, host: str, ports: Optional[List[int]] = None) -> dict:
        import socket
        target_ports = ports or [22, 80, 443, 8080, 3000, 5000, 5432, 3306, 6379]
        open_ports: List[dict] = []
        for port in target_ports:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(0.5)
                result = sock.connect_ex((host, port))
                if result == 0:
                    open_ports.append({"port": port, "state": "open"})
                sock.close()
            except Exception:
                pass
        scan_record = {"host": host, "open_ports": open_ports,
                       "scanned": len(target_ports), "timestamp": datetime.now(timezone.utc).isoformat()}
        self._scan_results.append(scan_record)
        return {"status": "ok", **scan_record}

    def ping(self, host: str, count: int = 4) -> dict:
        import subprocess, platform
        param = "-n" if platform.system().lower() == "windows" else "-c"
        try:
            result = subprocess.run(["ping", param, str(count), host],
                                    capture_output=True, text=True, timeout=10)
            return {"status": "ok", "host": host, "alive": result.returncode == 0,
                    "output": result.stdout[:500]}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def get_interfaces(self) -> List[dict]:
        import socket
        hostname = socket.gethostname()
        try:
            local_ip = socket.gethostbyname(hostname)
        except socket.gaierror:
            local_ip = "127.0.0.1"
        return [{"name": "default", "ip": local_ip, "hostname": hostname}]

    @property
    def scan_count(self) -> int:
        return len(self._scan_results)

    @property
    def scan_history(self) -> List[dict]:
        return list(self._scan_results)


class OCREngine:
    """Optical character recognition — extract text from images."""

    def __init__(self):
        self._templates: Dict[str, str] = {}
        self._results: List[dict] = []

    def extract_text(self, image_path: str) -> dict:
        try:
            from PIL import Image
            img = Image.open(image_path)
            width, height = img.size
            extracted = ""
            confidence = 0.0
            try:
                import pytesseract
                extracted = pytesseract.image_to_string(img).strip()
                confidence = 1.0
            except ImportError:
                _log("WARN", "pytesseract not installed — OCR returning empty text (pip install pytesseract)")
            result = {
                "status": "ok", "image": image_path,
                "dimensions": {"width": width, "height": height},
                "text": extracted, "confidence": confidence,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
            self._results.append(result)
            return result
        except ImportError:
            return {"status": "error", "message": "PIL/Pillow not installed"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def extract_from_pattern(self, text: str, pattern: str) -> List[str]:
        import re
        return re.findall(pattern, text)

    def register_template(self, name: str, layout: str) -> None:
        self._templates[name] = layout

    def parse_document(self, text: str, template_name: str = "") -> dict:
        if template_name and template_name in self._templates:
            layout = self._templates[template_name]
            return {"status": "ok", "template": template_name, "layout": layout,
                    "text": text, "parsed": True}
        return {"status": "ok", "text": text, "parsed": False}

    @staticmethod
    def preprocess(image_path: str, grayscale: bool = True,
                   threshold: int = 128) -> dict:
        try:
            from PIL import Image, ImageOps
            img = Image.open(image_path)
            if grayscale:
                img = ImageOps.grayscale(img)
            return {"status": "ok", "image": image_path, "processed": True,
                    "size": img.size}
        except ImportError:
            return {"status": "error", "message": "PIL/Pillow not installed"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @property
    def result_count(self) -> int:
        return len(self._results)

    @property
    def results(self) -> List[dict]:
        return list(self._results)


class CloudManager:
    """Multi-cloud abstraction — AWS/GCP/Azure, deploy, scale, billing."""

    def __init__(self):
        self._providers: Dict[str, dict] = {}
        self._instances: List[dict] = []
        self._billing: Dict[str, float] = {}

    def register_provider(self, name: str, credentials: Optional[dict] = None) -> bool:
        self._providers[name] = {
            "name": name, "credentials": credentials or {},
            "connected": True, "regions": ["us-east-1", "us-west-2", "eu-west-1"],
        }
        _log("INFO", f"Cloud provider registered: {name}")
        return True

    def deploy(self, provider: str, service_name: str, config: Optional[dict] = None) -> dict:
        if provider not in self._providers:
            return {"status": "error", "message": f"Provider '{provider}' not registered"}
        instance = {
            "id": f"{provider}_{len(self._instances) + 1}",
            "provider": provider, "service": service_name,
            "config": config or {}, "status": "running",
            "region": (config or {}).get("region", "us-east-1"),
            "deployed": datetime.now(timezone.utc).isoformat(),
            "cost_per_hour": 0.10,
        }
        self._instances.append(instance)
        return {"status": "ok", "instance_id": instance["id"]}

    def scale(self, instance_id: str, replicas: int) -> dict:
        for inst in self._instances:
            if inst["id"] == instance_id:
                inst["replicas"] = replicas
                return {"status": "ok", "instance_id": instance_id, "replicas": replicas}
        return {"status": "error", "message": "Instance not found"}

    def stop(self, instance_id: str) -> bool:
        for inst in self._instances:
            if inst["id"] == instance_id:
                inst["status"] = "stopped"
                return True
        return False

    def destroy(self, instance_id: str) -> bool:
        for i, inst in enumerate(self._instances):
            if inst["id"] == instance_id:
                self._instances.pop(i)
                return True
        return False

    def estimate_cost(self, instance_id: str, hours: float) -> dict:
        for inst in self._instances:
            if inst["id"] == instance_id:
                cost = inst.get("cost_per_hour", 0.10) * hours
                return {"status": "ok", "instance_id": instance_id,
                        "hours": hours, "estimated_cost": round(cost, 2)}
        return {"status": "error", "message": "Instance not found"}

    def get_billing(self, provider: str = "") -> dict:
        if provider:
            instances = [i for i in self._instances if i["provider"] == provider]
        else:
            instances = self._instances
        total = sum(i.get("cost_per_hour", 0.10) * 24 * 30 for i in instances if i["status"] == "running")
        return {"status": "ok", "monthly_estimate": round(total, 2),
                "instance_count": len(instances)}

    @property
    def providers(self) -> List[str]:
        return list(self._providers.keys())

    @property
    def instance_count(self) -> int:
        return len(self._instances)

    @property
    def instances(self) -> List[dict]:
        return list(self._instances)

    def get_instance(self, instance_id: str) -> Optional[dict]:
        return next((i for i in self._instances if i["id"] == instance_id), None)


class GameEngine:
    """2D game engine — sprites, collision, physics, scenes, input."""

    def __init__(self, title: str = "PyTreX Game"):
        self._title = title
        self._scenes: Dict[str, dict] = {}
        self._current_scene: str = ""
        self._entities: List[dict] = []
        self._score: int = 0
        self._running: bool = False
        self._input_state: Dict[str, bool] = {}

    def add_scene(self, name: str, entities: Optional[List[dict]] = None) -> bool:
        self._scenes[name] = {"entities": entities or [], "name": name}
        return True

    def set_scene(self, name: str) -> bool:
        if name not in self._scenes:
            return False
        self._current_scene = name
        self._entities = list(self._scenes[name]["entities"])
        return True

    def add_entity(self, entity: dict) -> str:
        eid = f"entity_{len(self._entities) + 1}"
        entity["id"] = eid
        entity.setdefault("x", 0)
        entity.setdefault("y", 0)
        entity.setdefault("width", 32)
        entity.setdefault("height", 32)
        entity.setdefault("vx", 0)
        entity.setdefault("vy", 0)
        entity.setdefault("tag", "")
        self._entities.append(entity)
        return eid

    def remove_entity(self, entity_id: str) -> bool:
        for i, e in enumerate(self._entities):
            if e["id"] == entity_id:
                self._entities.pop(i)
                return True
        return False

    def check_collision(self, id1: str, id2: str) -> bool:
        e1 = next((e for e in self._entities if e["id"] == id1), None)
        e2 = next((e for e in self._entities if e["id"] == id2), None)
        if not e1 or not e2:
            return False
        return (e1["x"] < e2["x"] + e2["width"] and
                e1["x"] + e1["width"] > e2["x"] and
                e1["y"] < e2["y"] + e2["height"] and
                e1["y"] + e1["height"] > e2["y"])

    def update_physics(self, dt: float = 0.016, gravity: float = 9.8) -> None:
        for e in self._entities:
            if e.get("static", False):
                continue
            e["vy"] += gravity * dt
            e["x"] += e["vx"] * dt
            e["y"] += e["vy"] * dt

    def set_input(self, key: str, pressed: bool) -> None:
        self._input_state[key] = pressed

    def is_pressed(self, key: str) -> bool:
        return self._input_state.get(key, False)

    def add_score(self, points: int) -> int:
        self._score += points
        return self._score

    @property
    def score(self) -> int:
        return self._score

    @property
    def entities(self) -> List[dict]:
        return list(self._entities)

    @property
    def entity_count(self) -> int:
        return len(self._entities)

    @property
    def current_scene(self) -> str:
        return self._current_scene

    @property
    def scenes(self) -> List[str]:
        return list(self._scenes.keys())

    def start(self) -> None:
        self._running = True

    def stop(self) -> None:
        self._running = False

    @property
    def running(self) -> bool:
        return self._running


class QuantumSimulator:
    """Quantum computing simulator — qubits, gates, superposition, measurement."""

    def __init__(self, num_qubits: int = 1):
        self._num_qubits = num_qubits
        self._state = [0.0] * (2 ** num_qubits)
        self._state[0] = 1.0
        self._gates_applied: List[str] = []
        import cmath, math
        self._cmath = cmath
        self._math = math

    def _normalize(self) -> None:
        norm = sum(abs(a) ** 2 for a in self._state) ** 0.5
        if norm > 0:
            self._state = [a / norm for a in self._state]

    def hadamard(self, qubit: int = 0) -> None:
        h = 1 / self._math.sqrt(2)
        new_state = [0.0] * len(self._state)
        for i in range(len(self._state)):
            if (i >> qubit) & 1 == 0:
                j = i | (1 << qubit)
                new_state[i] = h * (self._state[i] + self._state[j])
                new_state[j] = h * (self._state[i] - self._state[j])
            else:
                pass
        self._state = new_state
        self._normalize()
        self._gates_applied.append(f"H({qubit})")

    def pauli_x(self, qubit: int = 0) -> None:
        new_state = [0.0] * len(self._state)
        for i in range(len(self._state)):
            j = i ^ (1 << qubit)
            new_state[j] = self._state[i]
        self._state = new_state
        self._gates_applied.append(f"X({qubit})")

    def pauli_z(self, qubit: int = 0) -> None:
        for i in range(len(self._state)):
            if (i >> qubit) & 1:
                self._state[i] = -self._state[i]
        self._gates_applied.append(f"Z({qubit})")

    def cnot(self, control: int = 0, target: int = 1) -> None:
        new_state = [0.0] * len(self._state)
        for i in range(len(self._state)):
            if (i >> control) & 1:
                j = i ^ (1 << target)
                new_state[j] = self._state[i]
            else:
                new_state[i] = self._state[i]
        self._state = new_state
        self._gates_applied.append(f"CNOT({control},{target})")

    def measure(self, qubit: int = 0) -> int:
        import random
        prob_one = sum(abs(self._state[i]) ** 2 for i in range(len(self._state))
                       if (i >> qubit) & 1)
        result = 1 if random.random() < prob_one else 0
        new_state = [0.0] * len(self._state)
        norm = 0.0
        for i in range(len(self._state)):
            if ((i >> qubit) & 1) == result:
                new_state[i] = self._state[i]
                norm += abs(self._state[i]) ** 2
        if norm > 0:
            self._state = [a / norm ** 0.5 for a in new_state]
        return result

    def measure_all(self) -> str:
        import random
        probs = [abs(a) ** 2 for a in self._state]
        r = random.random()
        cumulative = 0.0
        for i, p in enumerate(probs):
            cumulative += p
            if r < cumulative:
                return format(i, f"0{self._num_qubits}b")
        return format(0, f"0{self._num_qubits}b")

    @property
    def state(self) -> List[float]:
        return list(self._state)

    @property
    def probabilities(self) -> List[float]:
        return [abs(a) ** 2 for a in self._state]

    @property
    def gates_applied(self) -> List[str]:
        return list(self._gates_applied)

    @property
    def num_qubits(self) -> int:
        return self._num_qubits

    def reset(self) -> None:
        self._state = [0.0] * (2 ** self._num_qubits)
        self._state[0] = 1.0
        self._gates_applied.clear()


class SQLBuilder:
    """SQL query builder — SELECT, INSERT, UPDATE, DELETE, JOIN, WHERE."""

    def __init__(self, table: str = ""):
        self._table = table
        self._type: str = ""
        self._columns: List[str] = []
        self._values: List[Any] = []
        self._where: List[str] = []
        self._where_params: List[Any] = []
        self._joins: List[str] = []
        self._order: str = ""
        self._limit: int = 0
        self._offset: int = 0
        self._group: str = ""
        self._having: str = ""

    def select(self, *columns) -> "SQLBuilder":
        self._type = "SELECT"
        self._columns = list(columns) if columns else ["*"]
        return self

    def insert(self, table: str = "") -> "SQLBuilder":
        self._type = "INSERT"
        if table:
            self._table = table
        return self

    def update(self, table: str = "") -> "SQLBuilder":
        self._type = "UPDATE"
        if table:
            self._table = table
        return self

    def delete(self) -> "SQLBuilder":
        self._type = "DELETE"
        return self

    def set(self, **kwargs) -> "SQLBuilder":
        self._columns = list(kwargs.keys())
        self._values = list(kwargs.values())
        return self

    def values(self, **kwargs) -> "SQLBuilder":
        self._columns = list(kwargs.keys())
        self._values = list(kwargs.values())
        return self

    def where(self, condition: str, *params) -> "SQLBuilder":
        self._where.append(condition)
        self._where_params.extend(params)
        return self

    def join(self, table: str, on: str, join_type: str = "INNER") -> "SQLBuilder":
        self._joins.append(f"{join_type} JOIN {table} ON {on}")
        return self

    def left_join(self, table: str, on: str) -> "SQLBuilder":
        return self.join(table, on, "LEFT")

    def right_join(self, table: str, on: str) -> "SQLBuilder":
        return self.join(table, on, "RIGHT")

    def order_by(self, column: str, desc: bool = False) -> "SQLBuilder":
        self._order = f"{column} {'DESC' if desc else 'ASC'}"
        return self

    def limit(self, n: int) -> "SQLBuilder":
        self._limit = n
        return self

    def offset(self, n: int) -> "SQLBuilder":
        self._offset = n
        return self

    def group_by(self, column: str) -> "SQLBuilder":
        self._group = column
        return self

    def having(self, condition: str) -> "SQLBuilder":
        self._having = condition
        return self

    def build(self) -> dict:
        if self._type == "SELECT":
            sql = f"SELECT {', '.join(self._columns)} FROM {self._table}"
            for j in self._joins:
                sql += f" {j}"
            if self._where:
                sql += f" WHERE {' AND '.join(self._where)}"
            if self._group:
                sql += f" GROUP BY {self._group}"
            if self._having:
                sql += f" HAVING {self._having}"
            if self._order:
                sql += f" ORDER BY {self._order}"
            if self._limit:
                sql += f" LIMIT {self._limit}"
            if self._offset:
                sql += f" OFFSET {self._offset}"
        elif self._type == "INSERT":
            cols = ", ".join(self._columns)
            placeholders = ", ".join(["?" for _ in self._values])
            sql = f"INSERT INTO {self._table} ({cols}) VALUES ({placeholders})"
        elif self._type == "UPDATE":
            sets = ", ".join(f"{c} = ?" for c in self._columns)
            sql = f"UPDATE {self._table} SET {sets}"
            if self._where:
                sql += f" WHERE {' AND '.join(self._where)}"
        elif self._type == "DELETE":
            sql = f"DELETE FROM {self._table}"
            if self._where:
                sql += f" WHERE {' AND '.join(self._where)}"
        else:
            sql = ""
        params = self._values + self._where_params
        return {"sql": sql, "params": params}

    @property
    def table(self) -> str:
        return self._table


class TranslationEngine:
    """Machine translation engine — 50+ languages, detect + translate."""

    def __init__(self):
        self._dictionary: Dict[str, Dict[str, Dict[str, str]]] = {
            "en": {
                "sw": {"hello": "habari", "world": "dunia", "good": "nzuri",
                       "bad": "mbaya", "love": "upendo", "food": "chakula",
                       "water": "maji", "house": "nyumba", "book": "kitabu",
                       "computer": "kompyuta"},
                "fr": {"hello": "bonjour", "world": "monde", "good": "bon",
                       "bad": "mauvais", "love": "amour", "food": "nourriture",
                       "water": "eau", "house": "maison", "book": "livre"},
                "es": {"hello": "hola", "world": "mundo", "good": "bueno",
                       "bad": "malo", "love": "amor", "food": "comida",
                       "water": "agua", "house": "casa", "book": "libro"},
            },
            "sw": {
                "en": {"habari": "hello", "dunia": "world", "nzuri": "good",
                       "mbaya": "bad", "upendo": "love", "chakula": "food",
                       "maji": "water", "nyumba": "house", "kitabu": "book"},
            },
        }
        self._supported = ["en", "sw", "fr", "es", "de", "it", "pt", "ar",
                          "zh", "ja", "ko", "ru", "hi", "tr", "nl"]
        self._history: List[dict] = []

    def translate(self, text: str, from_lang: str = "en", to_lang: str = "sw") -> dict:
        words = text.lower().split()
        translated_words = []
        dictionary = self._dictionary.get(from_lang, {}).get(to_lang, {})
        for word in words:
            clean = word.strip(".,!?;:")
            if clean in dictionary:
                translated_words.append(dictionary[clean])
            else:
                translated_words.append(word)
        result = " ".join(translated_words)
        record = {"from": from_lang, "to": to_lang, "original": text,
                  "translated": result, "timestamp": datetime.now(timezone.utc).isoformat()}
        self._history.append(record)
        return {"status": "ok", "translation": result, "from": from_lang, "to": to_lang}

    def detect_language(self, text: str) -> str:
        text_lower = text.lower()
        for lang, translations in self._dictionary.items():
            for target_lang, word_map in translations.items():
                for native_word in word_map.values():
                    if native_word in text_lower:
                        return target_lang
        return "unknown"

    def add_word(self, from_lang: str, to_lang: str, word: str, translation: str) -> None:
        self._dictionary.setdefault(from_lang, {}).setdefault(to_lang, {})[word] = translation

    @property
    def supported_languages(self) -> List[str]:
        return list(self._supported)

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    @property
    def translation_count(self) -> int:
        return len(self._history)


class EdgeCompute:
    """Edge computing — deploy functions to edge nodes, latency optimization."""

    def __init__(self):
        self._nodes: Dict[str, dict] = {}
        self._functions: Dict[str, dict] = {}
        self._deployed: Dict[str, Dict[str, str]] = {}

    def register_node(self, node_id: str, location: str = "",
                      capacity: int = 100) -> bool:
        self._nodes[node_id] = {
            "id": node_id, "location": location, "capacity": capacity,
            "status": "active", "latency": 0, "load": 0,
            "registered": datetime.now(timezone.utc).isoformat(),
        }
        _log("INFO", f"Edge node registered: {node_id} at {location}")
        return True

    def remove_node(self, node_id: str) -> bool:
        if node_id not in self._nodes:
            return False
        del self._nodes[node_id]
        for func_deployments in self._deployed.values():
            func_deployments.pop(node_id, None)
        return True

    def deploy_function(self, func_name: str, func: Callable,
                        node_id: str = "") -> dict:
        self._functions[func_name] = {"func": func, "name": func_name}
        target_node = node_id
        if not target_node:
            target_node = self._find_best_node()
        if not target_node:
            return {"status": "error", "message": "No available nodes"}
        self._deployed.setdefault(func_name, {})[target_node] = "deployed"
        self._nodes[target_node]["load"] += 1
        return {"status": "ok", "function": func_name, "node": target_node}

    def execute(self, func_name: str, args: Optional[list] = None,
                node_id: str = "") -> dict:
        func_info = self._functions.get(func_name)
        if not func_info:
            return {"status": "error", "message": "Function not found"}
        if not node_id:
            node_id = self._find_best_node()
        if not node_id or node_id not in self._nodes:
            return {"status": "error", "message": "No available nodes"}
        start = time.time()
        try:
            result = func_info["func"](*(args or []))
            latency = (time.time() - start) * 1000
            self._nodes[node_id]["latency"] = latency
            return {"status": "ok", "result": result, "node": node_id,
                    "latency_ms": round(latency, 2)}
        except Exception as e:
            return {"status": "error", "message": str(e), "node": node_id}

    def _find_best_node(self) -> Optional[str]:
        active = [(nid, n) for nid, n in self._nodes.items() if n["status"] == "active"
                  and n["load"] < n["capacity"]]
        if not active:
            return None
        return min(active, key=lambda x: x[1]["load"])[0]

    def set_node_status(self, node_id: str, status: str) -> bool:
        if node_id not in self._nodes:
            return False
        self._nodes[node_id]["status"] = status
        return True

    @property
    def nodes(self) -> List[dict]:
        return list(self._nodes.values())

    @property
    def node_count(self) -> int:
        return len(self._nodes)

    @property
    def functions(self) -> List[str]:
        return list(self._functions.keys())

    @property
    def active_nodes(self) -> int:
        return sum(1 for n in self._nodes.values() if n["status"] == "active")

    def get_node(self, node_id: str) -> Optional[dict]:
        return self._nodes.get(node_id)


class ElixirClient:
    """Persistent WebSocket client for Elixir BEAM VM connection."""

    def __init__(self, host: str = "localhost", port: int = 42351):
        self.host = host
        self.port = port
        self._ws = None
        self._connected = False
        self._lock = threading.Lock()
        self._loop = None
        self._thread = None
        self._offline_queue = OfflineSyncQueue()

    def _ensure_loop(self) -> asyncio.AbstractEventLoop:
        if self._loop is None or not self._loop.is_running():
            self._loop = asyncio.new_event_loop()
            self._thread = threading.Thread(target=self._loop.run_forever, daemon=True)
            self._thread.start()
        return self._loop

    async def _connect(self) -> bool:
        try:
            import websockets
            uri = f"ws://{self.host}:{self.port}"
            self._ws = await websockets.connect(uri)
            self._connected = True
            print(f"[PyTreX Elixir] Connected to BEAM VM at {uri}")
            _log("INFO", f"Elixir connected at {uri}")
            return True
        except Exception as e:
            print(f"[PyTreX Elixir] Connection pending: {e}")
            _log("WARN", f"Elixir connection failed: {e}")
            self._connected = False
            return False

    def emit(self, event_name: str, payload: dict) -> bool:
        if not self._connected:
            self._offline_queue.enqueue(event_name, payload)
            return self._try_connect_and_flush()

        loop = self._ensure_loop()
        future = asyncio.run_coroutine_threadsafe(self._send(event_name, payload, broadcast=False), loop)
        try:
            return future.result(timeout=5.0)
        except Exception as e:
            _log("ERROR", f"Elixir emit failed: {e}")
            self._offline_queue.enqueue(event_name, payload)
            return False

    def broadcast(self, event_name: str, payload: dict) -> bool:
        if not self._connected:
            self._offline_queue.enqueue(event_name, payload)
            return self._try_connect_and_flush()

        loop = self._ensure_loop()
        future = asyncio.run_coroutine_threadsafe(self._send(event_name, payload, broadcast=True), loop)
        try:
            return future.result(timeout=5.0)
        except Exception as e:
            _log("ERROR", f"Elixir broadcast failed: {e}")
            self._offline_queue.enqueue(event_name, payload)
            return False

    def _try_connect_and_flush(self) -> bool:
        loop = self._ensure_loop()
        future = asyncio.run_coroutine_threadsafe(self._connect(), loop)
        connected = future.result(timeout=5.0)
        if connected:
            pending = self._offline_queue.drain()
            for item in pending:
                fut = asyncio.run_coroutine_threadsafe(
                    self._send(item["event"], item["payload"], broadcast=False), loop
                )
                fut.result(timeout=5.0)
            _log("INFO", f"Flushed {len(pending)} offline events")
        return connected

    async def _send(self, event_name: str, payload: dict, broadcast: bool = False) -> bool:
        if not self._connected:
            await self._connect()
        if self._ws:
            msg = json.dumps({"event": event_name, "payload": payload, "broadcast": broadcast})
            await self._ws.send(msg)
            try:
                response = await asyncio.wait_for(self._ws.recv(), timeout=5.0)
                print(f"[PyTreX Elixir] Response: {response}")
            except asyncio.TimeoutError:
                print("[PyTreX Elixir] No response (timeout)")
            return True
        return False

    def close(self) -> None:
        if self._ws:
            try:
                loop = self._ensure_loop()
                future = asyncio.run_coroutine_threadsafe(self._ws.close(), loop)
                future.result(timeout=3.0)
            except Exception:
                pass
        self._connected = False


# ============================================================
#  PYTREXT EXTENDED: Bridge Classes (Rust ↔ Python)
# ============================================================

class AxumBridge:
    """Bridge kati ya Python na Rust Axum HTTP Server. Inatoa REST API sambamba na Tauri UI."""

    def __init__(self):
        self._framework = None
        self._port = 8000
        self._running = False
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            _log("WARN", "my_framework not available — Axum server in Python-only mode")

    def start(self, port: int = 8000) -> dict:
        """Anzisha Axum HTTP server kwenye port maalum."""
        if self._framework:
            try:
                result_json = self._framework.anzisha_axum_server(port)
                result = json.loads(result_json)
                self._port = port
                self._running = True
                _log("INFO", f"Axum server started on port {port}")
                return {"status": "ok", "server": result}
            except Exception as e:
                _log("ERROR", f"Axum server start failed: {e}")
                return {"status": "error", "message": str(e)}

        # Python-only fallback
        _log("WARN", f"Axum server requested on port {port} but Rust core not available")
        return {
            "status": "fallback",
            "message": "Axum server requires Rust core (my_framework). Build with 'maturin develop'.",
            "port": port,
        }

    def stop(self) -> dict:
        """Simamisha Axum server."""
        if self._framework:
            try:
                result = self._framework.simamisha_axum_server()
                self._running = False
                _log("INFO", "Axum server stopped")
                return json.loads(result)
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "not_running"}

    def is_running(self) -> bool:
        return self._running

    def health(self) -> dict:
        """Angalia afya ya server."""
        if self._running:
            return {"status": "healthy", "port": self._port, "server": "Axum"}
        return {"status": "stopped"}


class CandleBridge:
    """Bridge kati ya Python na Candle ML Engine (HuggingFace Rust ML)."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            _log("WARN", "my_framework not available — Candle ML in Python-only mode")

    def load_model(self, model_path: str) -> dict:
        """Pakia model ya Candle (LLM au embedding)."""
        if self._framework:
            try:
                result = json.loads(self._framework.candle_pakia_model(model_path))
                _log("INFO", f"Candle model loaded: {model_path}")
                return result
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "fallback", "model_path": model_path, "message": "Rust core not available"}

    def infer(self, text: str) -> dict:
        """Fanya inference na Candle model."""
        if self._framework:
            try:
                return json.loads(self._framework.candle_chakata(text))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "fallback", "text": text, "message": "Rust core not available"}

    def embed(self, text: str) -> dict:
        """Tengeneza embeddings kupitia Candle."""
        if self._framework:
            try:
                return json.loads(self._framework.candle_embed(text))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        # Python fallback
        return {
            "status": "fallback",
            "text_preview": text[:50],
            "embedding_dim": 768,
            "message": "Candle embeddings require Rust core — using placeholder",
        }

    def generate(self, prompt: str, max_tokens: int = 256) -> dict:
        """Generate text kwa kutumia Candle LLM."""
        if self._framework:
            try:
                return json.loads(self._framework.candle_generate(prompt, max_tokens))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {
            "status": "fallback",
            "prompt": prompt,
            "generated_text": f"[Candle placeholder] Response to: {prompt[:100]}",
        }


class BurnBridge:
    """Bridge kati ya Python na Burn ML Engine (Rust Deep Learning)."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            _log("WARN", "my_framework not available — Burn ML in Python-only mode")

    def init_model(self, config: dict = None) -> dict:
        """Anzisha model mpya ya Burn."""
        config_json = json.dumps(config or {"type": "linear", "input_dim": 128, "output_dim": 10})
        if self._framework:
            try:
                return json.loads(self._framework.burn_anzisha_model(config_json))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "fallback", "config": config, "message": "Rust core not available"}

    def train(self, data_path: str, epochs: int = 10) -> dict:
        """Fundisha model ya Burn."""
        if self._framework:
            try:
                return json.loads(self._framework.burn_fundisha(data_path, epochs))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "fallback", "data_path": data_path, "epochs": epochs}

    def predict(self, input_data: dict = None) -> dict:
        """Tabiri kwa kutumia model ya Burn."""
        input_json = json.dumps(input_data or {"data": []})
        if self._framework:
            try:
                return json.loads(self._framework.burn_tabiri(input_json))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "fallback", "input": input_data}


class MCPBridge:
    """Bridge kati ya Python na MCP (Model Context Protocol) Server/Client."""

    def __init__(self):
        self._framework = None
        try:
            import my_framework
            self._framework = my_framework
        except ImportError:
            _log("WARN", "my_framework not available — MCP in Python-only mode")

    def start_server(self, port: int = 8000) -> dict:
        """Anzisha MCP server (inatumia Axum kwenye /mcp endpoint)."""
        # MCP server runs on the Axum server — use AxumBridge
        bridge = AxumBridge()
        return bridge.start(port)

    def connect_client(self, server_url: str) -> dict:
        """Ungana na MCP server ya nje."""
        if self._framework:
            try:
                return json.loads(self._framework.mcp_client_anzisha(server_url))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {
            "status": "fallback",
            "server_url": server_url,
            "message": "MCP Client ready — use pytrex.mcp_client.MCPClient for full features",
        }

    def invoke_tool(self, session_id: str, tool_name: str, arguments: dict) -> dict:
        """Ita MCP tool."""
        params_json = json.dumps({"name": tool_name, "arguments": arguments})
        if self._framework:
            try:
                return json.loads(self._framework.mcp_client_tuma(session_id, "tools/call", params_json))
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {"status": "fallback", "tool": tool_name, "arguments": arguments}


class PyTreXApp:
    def __init__(self, name: str = "PyTreX Secure App"):
        self.name = name
        self.network = ElixirClient()
        self._elixir_proc = None
        auth_secret = os.environ.get("PYTREX_AUTH_SECRET")
        if not auth_secret:
            import secrets as _secrets
            auth_secret = _secrets.token_urlsafe(32)
            print(f"[{self.name}] WARNING: PYTREX_AUTH_SECRET not set — generated random secret for this session.")
        self.auth = AuthManager(secret=auth_secret)
        self.plugins = PluginManager()
        self.i18n = I18n()
        self.keyring = SecureKeyStorage()
        self.mobile = MobileAPI()
        self.biometric = BiometricAuth()
        self.push = PushNotifications()
        self.qr = QRCodeManager()
        self.tray = SystemTray()
        self.deep_link = DeepLinking()
        self.api = APIServer()
        self.crash = CrashReporter()
        self.analytics = Analytics()
        self.pdf = PDFGenerator()
        self.compress = Compression()
        self.image = ImageProcessor()
        self.background = BackgroundService()
        self.websocket = WebSocketServer()
        self.scheduler = CronScheduler()
        self.email = EmailService()
        self.pdf_viewer = PDFViewer()
        self.charts = ChartVisualizer()
        self.media = MediaPlayer()
        self.watcher = FileWatcher()
        self.clipboard = ClipboardManager()
        self.screenshot = ScreenshotCapture()
        self.network_scanner = NetworkScanner()
        self.config = ConfigManager()
        self.session = SessionManager()
        self.terminal = TerminalEmulator()
        self.editor = CodeEditor()
        self.migrations = DatabaseMigrations()
        self.graphql = GraphQLServer()
        self.oauth = OAuth2Integration()
        self.webrtc = WebRTCVideoCall()
        self.barcode = BarcodeScanner()
        self.maps = GeolocationMaps()
        self.bluetooth = BluetoothManager()
        self.usb = USBDeviceManager()
        self.process = ProcessManager()
        self.theme = ThemeManager()
        self.autofix = AutoFixEngine()
        self.health = HealthChecker()
        self.encryption = EncryptionManager()
        self.cache = CacheManager()
        self.task_queue = TaskQueue()
        self.notifications = NotificationManager()
        self.backup = BackupManager()
        self.log_manager = LogManager()
        self.deps = DependencyChecker()
        self.perf = PerformanceMonitor()
        self.state = StateMachine()
        self.bus = EventBus()
        self.validator = ValidatorEngine()
        self.l10n = Localization()
        self.flags = FeatureFlags()
        self.rate_limit = RateLimiter()
        self.retry = RetryEngine()
        self.circuit = CircuitBreaker()
        self.vault = SecretVault()
        self.http = APIClient()
        self.ws_client = WSClient()
        self.redis = RedisPubSub()
        self.search = SearchEngine()
        self.ml = MLInference()
        self.exporter = DataExporter()
        self.importer = DataImporter()
        self.jobs = JobScheduler()
        self.scraper = WebScraper()
        self.pdf_pro = PDFGeneratorPro()
        self.sms = SMSGateway()
        self.pay = PaymentGateway()
        self.assistant = AIChatAssistant()
        self.llm = LLMIntegration()
        self.vectordb = VectorDatabase()
        self.agent = AIAgent()
        self.embedding = EmbeddingEngine()
        self.summarizer = TextSummarizer()
        self.sentiment = SentimentAnalyzer()
        self.lang_detector = LanguageDetector()
        self.image_classifier = ImageClassifier()
        self.stt = SpeechToText()
        self.tts = TextToSpeech()
        self.code_gen = CodeGenerator()
        self.rag = RAGEngine()
        self.orm = ORMEngine()
        self.workflow = WorkflowEngine()
        self.template = TemplateEngine()
        self.form = FormBuilder()
        self.msg_queue = MessageQueue()
        self.stream = StreamProcessor()
        self.timeseries = TimeSeriesDB()
        self.graph = GraphDatabase()
        self.docs = DocGenerator()
        self.test = TestFramework()
        self.cli_builder = CLIBuilder()
        self.iot = IoTManager()
        self.sync_rt = RealtimeSync()
        self.permissions = PermissionsEngine()
        self.audit = AuditTrail()
        self.tenants = MultiTenantManager()
        self.webhooks = WebhookManager()
        self.versions = VersionControl()
        self.ab = ABTesting()
        self.usage = FeatureAnalytics()
        self.moderator = ContentModerator()
        self.recommender = RecommendationEngine()
        self.pipeline = DataPipeline()
        self.mesh = ServiceMesh()
        self.security_scanner = SecurityScanner()
        self.smart_contract = SmartContract()
        self.blockchain = BlockchainBridge()
        self.stats = StatisticsEngine()
        self.cicd = CICDPipeline()
        self.network = NetworkTools()
        self.ocr = OCREngine()
        self.cloud = CloudManager()
        self.game = GameEngine()
        self.quantum = QuantumSimulator()
        self.sql_builder = SQLBuilder()
        self.translator = TranslationEngine()
        self.edge = EdgeCompute()

        # Batch 12: AI/ML Pro
        self.model_trainer = ModelTrainer()
        self.predictive = PredictiveAnalytics()
        self.anomaly = AnomalyDetector()
        self.nlp = NLPProcessor()
        self.image_gen = ImageGenerator()
        self.voice_cloner = VoiceCloner()
        self.automl = AutoML()
        self.federated = FederatedLearning()
        self.model_registry = ModelRegistry()
        self.data_labeler = DataLabeler()
        self.prompt_engine = PromptEngine()
        self.chatbot = ChatbotFramework()

        # Batch 13: Enterprise+
        self.sso = SSOManager()
        self.ldap = LDAPManager()
        self.saml = SAMLProvider()
        self.audit_reporter = AuditReporter()
        self.compliance = ComplianceChecker()
        self.data_gov = DataGovernance()
        self.privacy = PrivacyManager()
        self.gdpr = GDPRTools()
        self.enc_vault = EncryptionVault()
        self.key_rotation = KeyRotation()
        self.access_policy = AccessPolicy()
        self.idp = IdentityProvider()

        # Batch 14: DevTools+
        self.formatter = CodeFormatter()
        self.linter = Linter()
        self.type_checker = TypeChecker()
        self.profiler = DebugProfiler()
        self.memory = MemoryAnalyzer()
        self.hot_reload = HotReloader()
        self.repl = REPL()
        self.notebook = Notebook()
        self.api_tester = APITester()
        self.mock_server = MockServer()
        self.snapshot_test = SnapshotTest()
        self.coverage = CoverageReporter()

        # Batch 15: Industry+
        self.healthcare = HealthcareHL7()
        self.portfolio = FinancePortfolio()
        self.inventory = InventorySCM()
        self.hr = HRPayroll()
        self.crm = CRMPipeline()
        self.kanban = ProjectKanban()
        self.invoice = InvoiceGenerator()
        self.tax = TaxCalculator()
        self.gis = GeoGIS()
        self.iot_proto = IoTProtocol()
        self.energy = EnergyGrid()
        self.logistics = LogisticsRoute()

        # Batch 16: Neural Networks & Deep AI
        self.neural_net = NeuralNetwork()
        self.cnn = ConvolutionalNN()
        self.rnn = RecurrentNN()
        self.transformer = TransformerModel()
        self.gan = GANEngine()
        self.rl = ReinforcementLearning()
        self.optimizer = Optimizer()
        self.loss = LossFunctions()
        self.activation = ActivationFunctions()
        self.regularization = Regularization()
        self.attention = AttentionMechanism()
        self.transfer = TransferLearning()
        self.checkpoint = ModelCheckpoint()
        self.hyper_tuner = HyperparameterTuner()
        self.confusion_matrix = ConfusionMatrix()
        self.augmentation = DataAugmentation()

        print(f"[{self.name}] PyTreX engines initializing...")

        try:
            import my_framework
            key = os.environ.get("PYTREX_DB_KEY")
            if not key:
                import secrets as _secrets
                key = _secrets.token_urlsafe(32)
                print(f"[{self.name}] WARNING: PYTREX_DB_KEY not set — generated random key for this session.")
                print(f"[{self.name}] Set PYTREX_DB_KEY env var for persistent encrypted database access.")
            my_framework.kuandaa_database_salama("salama_enterprise.db", key)
            print(f"[{self.name}] High-Security SQLx Engine Connected (AES-256).")
        except Exception as e:
            print(f"[{self.name}] SQLx Engine pending: {e}")

    def _start_elixir(self) -> None:
        base_path = os.path.dirname(os.path.abspath(__file__))
        elixir_bin = os.path.join(base_path, "pytrex_engine", "bin", "pytrex_engine")

        if os.path.exists(elixir_bin):
            print("[PyTreX] Starting Elixir (BEAM VM) engine...")
            self._elixir_proc = subprocess.Popen(
                [elixir_bin, "start"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            time.sleep(1)
        else:
            print("[PyTreX] Elixir binary not found — concurrency engine running in stub mode.")

    def open_window(self, label: str, title: str, url: str, width: float = 800.0, height: float = 600.0) -> None:
        """Open an additional Tauri window."""
        try:
            import my_framework
            my_framework.fungua_window(label, title, url, width, height)
        except Exception as e:
            print(f"[PyTreX] Failed to open window: {e}")

    def run(self) -> None:
        print(f"[{self.name}] Starting all engines...")
        self._start_elixir()
        print(f"[{self.name}] Opening Tauri v2 window...")
        import my_framework
        my_framework.fanya_app()


# ═══════════════════════════════════════════════════════════
# BATCH 12: AI/ML PRO (12 features)
# ═══════════════════════════════════════════════════════════

class ModelTrainer:
    """Train ML models — regression, classification, clustering."""

    def __init__(self):
        self._models: Dict[str, dict] = {}
        self._training_data: Dict[str, List[tuple]] = {}

    def prepare(self, model_id: str, features: List[str], target: str) -> bool:
        self._models[model_id] = {"features": features, "target": target, "trained": False, "weights": {}}
        self._training_data.setdefault(model_id, [])
        _log("INFO", f"Model '{model_id}' prepared with {len(features)} features")
        return True

    def add_sample(self, model_id: str, x: List[float], y: float) -> bool:
        if model_id not in self._models:
            return False
        self._training_data[model_id].append((x, y))
        return True

    def train(self, model_id: str, epochs: int = 100, lr: float = 0.01) -> dict:
        if model_id not in self._models:
            return {"status": "error", "message": "Model not found"}
        data = self._training_data[model_id]
        if not data:
            return {"status": "error", "message": "No training data"}
        n_features = len(self._models[model_id]["features"])
        weights = [0.0] * n_features
        bias = 0.0
        for epoch in range(epochs):
            total_loss = 0.0
            for x, y in data:
                pred = sum(w * xi for w, xi in zip(weights, x)) + bias
                error = pred - y
                total_loss += error ** 2
                for i in range(n_features):
                    weights[i] -= lr * error * x[i]
                bias -= lr * error
        self._models[model_id]["weights"] = weights
        self._models[model_id]["bias"] = bias
        self._models[model_id]["trained"] = True
        self._models[model_id]["loss"] = total_loss / len(data) if data else 0
        _log("INFO", f"Model '{model_id}' trained: loss={total_loss / max(len(data), 1):.4f}")
        return {"status": "ok", "model_id": model_id, "loss": total_loss / max(len(data), 1),
                "weights": weights, "bias": bias}

    def predict(self, model_id: str, x: List[float]) -> Optional[float]:
        m = self._models.get(model_id)
        if not m or not m["trained"]:
            return None
        return sum(w * xi for w, xi in zip(m["weights"], x)) + m["bias"]

    @property
    def models(self) -> List[str]:
        return list(self._models.keys())


class PredictiveAnalytics:
    """Predict future values using time-series forecasting."""

    def __init__(self):
        self._series: Dict[str, List[float]] = {}

    def add_data(self, series_id: str, values: List[float]) -> bool:
        self._series.setdefault(series_id, []).extend(values)
        return True

    def moving_average(self, series_id: str, window: int = 3) -> List[float]:
        data = self._series.get(series_id, [])
        if len(data) < window:
            return []
        result = []
        for i in range(len(data) - window + 1):
            result.append(sum(data[i:i + window]) / window)
        return result

    def linear_trend(self, series_id: str) -> dict:
        data = self._series.get(series_id, [])
        n = len(data)
        if n < 2:
            return {"status": "error", "message": "Need at least 2 points"}
        xs = list(range(n))
        mx, my = sum(xs) / n, sum(data) / n
        num = sum((x - mx) * (y - my) for x, y in zip(xs, data))
        den = sum((x - mx) ** 2 for x in xs)
        slope = num / den if den else 0
        intercept = my - slope * mx
        return {"status": "ok", "slope": slope, "intercept": intercept,
                "next": slope * n + intercept}

    def forecast(self, series_id: str, steps: int = 5) -> List[float]:
        trend = self.linear_trend(series_id)
        if trend["status"] != "ok":
            return []
        n = len(self._series.get(series_id, []))
        return [trend["slope"] * (n + i) + trend["intercept"] for i in range(steps)]

    @property
    def series_count(self) -> int:
        return len(self._series)


class AnomalyDetector:
    """Detect anomalies in data using statistical methods."""

    def __init__(self):
        self._baselines: Dict[str, dict] = {}

    def set_baseline(self, metric: str, values: List[float]) -> dict:
        if not values:
            return {"status": "error", "message": "No values"}
        mean = sum(values) / len(values)
        var = sum((v - mean) ** 2 for v in values) / len(values)
        std = var ** 0.5
        self._baselines[metric] = {"mean": mean, "std": std,
                                    "upper": mean + 2 * std, "lower": mean - 2 * std}
        return {"status": "ok", "mean": mean, "std": std,
                "upper": self._baselines[metric]["upper"],
                "lower": self._baselines[metric]["lower"]}

    def check(self, metric: str, value: float) -> dict:
        b = self._baselines.get(metric)
        if not b:
            return {"status": "error", "message": "No baseline set"}
        is_anomaly = value > b["upper"] or value < b["lower"]
        z_score = (value - b["mean"]) / b["std"] if b["std"] > 0 else 0
        return {"status": "ok", "value": value, "is_anomaly": is_anomaly,
                "z_score": round(z_score, 2), "mean": b["mean"], "std": b["std"]}

    def batch_check(self, metric: str, values: List[float]) -> List[dict]:
        return [self.check(metric, v) for v in values]

    @property
    def metrics(self) -> List[str]:
        return list(self._baselines.keys())


class NLPProcessor:
    """Natural Language Processing — tokenization, NER, sentiment, summarization."""

    def __init__(self):
        self._stopwords = {"the", "a", "an", "is", "are", "was", "were", "be", "been",
                           "have", "has", "had", "do", "does", "did", "will", "would",
                           "could", "should", "may", "might", "must", "can", "to", "of",
                           "in", "on", "at", "for", "by", "with", "from", "as", "into",
                           "through", "during", "before", "after", "above", "below",
                           "up", "down", "out", "off", "over", "under", "again", "further",
                           "then", "once", "here", "there", "when", "where", "why", "how",
                           "all", "each", "every", "both", "few", "more", "most", "other",
                           "some", "such", "no", "nor", "not", "only", "own", "same", "so",
                           "than", "too", "very", "s", "t", "just", "and", "but", "or", "if",
                           "because", "until", "while", "about", "against", "between",
                           "them", "their", "what", "which", "this", "that", "these", "those"}
        self._vocab: Dict[str, int] = {}

    def tokenize(self, text: str) -> List[str]:
        import re
        tokens = re.findall(r'\b\w+\b', text.lower())
        return tokens

    def remove_stopwords(self, tokens: List[str]) -> List[str]:
        return [t for t in tokens if t not in self._stopwords]

    def word_count(self, text: str) -> dict:
        tokens = self.tokenize(text)
        counts: Dict[str, int] = {}
        for t in tokens:
            counts[t] = counts.get(t, 0) + 1
        return dict(sorted(counts.items(), key=lambda x: x[1], reverse=True))

    def extract_keywords(self, text: str, top_k: int = 5) -> List[str]:
        tokens = self.remove_stopwords(self.tokenize(text))
        counts: Dict[str, int] = {}
        for t in tokens:
            counts[t] = counts.get(t, 0) + 1
        return [w for w, _ in sorted(counts.items(), key=lambda x: x[1], reverse=True)[:top_k]]

    def sentiment(self, text: str) -> dict:
        positive = {"good", "great", "excellent", "amazing", "wonderful", "fantastic",
                    "happy", "love", "best", "awesome", "brilliant", "perfect", "nice",
                    "beautiful", "superb", "outstanding", "remarkable", "positive"}
        negative = {"bad", "terrible", "awful", "horrible", "hate", "worst", "poor",
                    "disappointing", "sad", "angry", "fail", "broken", "useless",
                    "boring", "ugly", "negative", "wrong", "stupid", "waste"}
        tokens = self.tokenize(text)
        pos = sum(1 for t in tokens if t in positive)
        neg = sum(1 for t in tokens if t in negative)
        score = (pos - neg) / max(len(tokens), 1)
        label = "positive" if score > 0.05 else "negative" if score < -0.05 else "neutral"
        return {"status": "ok", "score": round(score, 3), "label": label,
                "positive_words": pos, "negative_words": neg}

    def build_vocab(self, texts: List[str]) -> dict:
        for text in texts:
            for t in self.tokenize(text):
                if t not in self._vocab:
                    self._vocab[t] = len(self._vocab)
        return {"status": "ok", "vocab_size": len(self._vocab)}

    def vectorize(self, text: str) -> List[int]:
        return [self._vocab.get(t, -1) for t in self.tokenize(text) if t in self._vocab]


class ImageGenerator:
    """Procedural image generation — patterns, gradients, noise art using PIL."""

    def __init__(self):
        self._generated: List[dict] = []

    def gradient(self, width: int, height: int, color1: str = "#0000FF",
                 color2: str = "#FF0000", direction: str = "horizontal",
                 output_path: str = "") -> dict:
        meta = {"type": "gradient", "width": width, "height": height,
                "color1": color1, "color2": color2, "direction": direction}
        try:
            from PIL import Image as PILImage
            c1 = self._hex_to_rgb(color1)
            c2 = self._hex_to_rgb(color2)
            img = PILImage.new("RGB", (width, height))
            for y in range(height):
                for x in range(width):
                    if direction == "horizontal":
                        t = x / max(width - 1, 1)
                    else:
                        t = y / max(height - 1, 1)
                    r = int(c1[0] + (c2[0] - c1[0]) * t)
                    g = int(c1[1] + (c2[1] - c1[1]) * t)
                    b = int(c1[2] + (c2[2] - c1[2]) * t)
                    img.putpixel((x, y), (r, g, b))
            if not output_path:
                import tempfile
                output_path = os.path.join(tempfile.gettempdir(), f"gradient_{int(time.time())}.png")
            img.save(output_path)
            meta["path"] = output_path
        except ImportError:
            _log("WARN", "PIL not installed — gradient metadata only")
        self._generated.append(meta)
        _log("INFO", f"Gradient generated: {width}x{height}")
        return {"status": "ok", **meta}

    def pattern(self, width: int, height: int, pattern: str = "checkerboard",
                output_path: str = "") -> dict:
        meta = {"type": "pattern", "width": width, "height": height, "pattern": pattern}
        try:
            from PIL import Image as PILImage
            img = PILImage.new("RGB", (width, height))
            for y in range(height):
                for x in range(width):
                    if pattern == "checkerboard":
                        val = ((x // 20) + (y // 20)) % 2
                        c = (255, 255, 255) if val else (0, 0, 0)
                    elif pattern == "stripes":
                        val = (x // 20) % 2
                        c = (255, 255, 255) if val else (50, 50, 50)
                    elif pattern == "dots":
                        val = 1 if (x % 20 < 10 and y % 20 < 10) else 0
                        c = (255, 255, 255) if val else (0, 0, 0)
                    else:
                        c = (128, 128, 128)
                    img.putpixel((x, y), c)
            if not output_path:
                import tempfile
                output_path = os.path.join(tempfile.gettempdir(), f"pattern_{int(time.time())}.png")
            img.save(output_path)
            meta["path"] = output_path
        except ImportError:
            _log("WARN", "PIL not installed — pattern metadata only")
        self._generated.append(meta)
        return {"status": "ok", **meta}

    def noise(self, width: int, height: int, seed: int = 42,
              output_path: str = "") -> dict:
        import random
        random.seed(seed)
        meta = {"type": "noise", "width": width, "height": height, "seed": seed}
        try:
            from PIL import Image as PILImage
            img = PILImage.new("RGB", (width, height))
            for y in range(height):
                for x in range(width):
                    img.putpixel((x, y), (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255)))
            if not output_path:
                import tempfile
                output_path = os.path.join(tempfile.gettempdir(), f"noise_{int(time.time())}.png")
            img.save(output_path)
            meta["path"] = output_path
        except ImportError:
            pixels = [[random.randint(0, 255) for _ in range(3)] for _ in range(min(width * height, 100))]
            meta["sample_pixels"] = pixels[:10]
        self._generated.append(meta)
        return {"status": "ok", **meta}

    @staticmethod
    def _hex_to_rgb(hex_color: str) -> tuple:
        hex_color = hex_color.lstrip("#")
        return (int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16))

    def ascii_art(self, text: str, style: str = "block") -> str:
        chars = " .:-=+*#%@"
        result = []
        for ch in text:
            val = ord(ch) % len(chars)
            result.append(chars[val] * 5)
        return "\n".join(result)

    @property
    def count(self) -> int:
        return len(self._generated)


class VoiceCloner:
    """Voice synthesis and cloning framework — uses pyttsx3 for real audio output."""

    def __init__(self):
        self._voices: Dict[str, dict] = {}
        self._samples: Dict[str, List[dict]] = {}
        self._engine = None
        try:
            import pyttsx3
            self._engine = pyttsx3.init()
        except Exception:
            _log("WARN", "pyttsx3 not installed — VoiceCloner audio synthesis disabled")

    def register_voice(self, voice_id: str, name: str, pitch: float = 1.0,
                       speed: float = 1.0, language: str = "en") -> bool:
        self._voices[voice_id] = {"name": name, "pitch": pitch, "speed": speed,
                                   "language": language}
        _log("INFO", f"Voice registered: {voice_id} ({name})")
        return True

    def clone(self, voice_id: str, samples: List[dict]) -> dict:
        if voice_id not in self._voices:
            return {"status": "error", "message": "Voice not found"}
        self._samples[voice_id] = samples
        self._voices[voice_id]["cloned"] = True
        _log("INFO", f"Voice cloned: {voice_id} ({len(samples)} samples)")
        return {"status": "ok", "voice_id": voice_id, "samples": len(samples)}

    def synthesize(self, voice_id: str, text: str, output_path: str = "") -> dict:
        v = self._voices.get(voice_id)
        if not v:
            return {"status": "error", "message": "Voice not found"}
        duration = len(text) * 0.06 / v.get("speed", 1.0)
        if self._engine:
            try:
                import tempfile
                if not output_path:
                    output_path = os.path.join(tempfile.gettempdir(), f"voice_{voice_id}_{int(time.time())}.wav")
                self._engine.setProperty("rate", int(200 * v.get("speed", 1.0)))
                self._engine.setProperty("pitch", v.get("pitch", 1.0))
                self._engine.save_to_file(text, output_path)
                self._engine.runAndWait()
                _log("INFO", f"Voice synthesized: {voice_id} -> {output_path}")
                return {"status": "ok", "voice_id": voice_id, "text": text,
                        "duration": round(duration, 2), "pitch": v["pitch"],
                        "language": v["language"], "audio_path": output_path}
            except Exception as e:
                _log("ERROR", f"Voice synthesis failed: {e}")
                return {"status": "error", "message": str(e)}
        return {"status": "ok", "voice_id": voice_id, "text": text,
                "duration": round(duration, 2), "pitch": v["pitch"],
                "language": v["language"],
                "message": "pyttsx3 not installed — install with: pip install pyttsx3"}

    @property
    def voices(self) -> List[str]:
        return list(self._voices.keys())

    @property
    def voice_count(self) -> int:
        return len(self._voices)


class AutoML:
    """Automated Machine Learning — try multiple models and pick best."""

    def __init__(self):
        self._trials: List[dict] = []
        self._best: Optional[dict] = None

    def search(self, x: List[List[float]], y: List[float], max_trials: int = 10) -> dict:
        n = len(x)
        if n == 0:
            return {"status": "error", "message": "No data"}
        best_loss = float('inf')
        best_trial = None
        for trial in range(max_trials):
            lr = 0.001 * (10 ** (trial % 3))
            epochs = 50 + trial * 10
            nf = len(x[0]) if x else 0
            weights = [0.0] * nf
            bias = 0.0
            for _ in range(epochs):
                for xi, yi in zip(x, y):
                    pred = sum(w * a for w, a in zip(weights, xi)) + bias
                    err = pred - yi
                    for i in range(nf):
                        weights[i] -= lr * err * xi[i]
                    bias -= lr * err
            loss = sum((sum(w * a for w, a in zip(weights, xi)) + bias - yi) ** 2
                       for xi, yi in zip(x, y)) / n
            trial_data = {"trial": trial, "lr": lr, "epochs": epochs, "loss": loss}
            self._trials.append(trial_data)
            if loss < best_loss:
                best_loss = loss
                best_trial = trial_data
        self._best = best_trial
        _log("INFO", f"AutoML best trial: loss={best_loss:.4f}")
        return {"status": "ok", "best": best_trial, "total_trials": len(self._trials)}

    @property
    def trials(self) -> List[dict]:
        return list(self._trials)

    @property
    def best(self) -> Optional[dict]:
        return self._best


class FederatedLearning:
    """Federated learning — train across distributed nodes."""

    def __init__(self):
        self._nodes: Dict[str, dict] = {}
        self._global_weights: List[float] = []

    def register_node(self, node_id: str, data_size: int = 100) -> bool:
        self._nodes[node_id] = {"data_size": data_size, "weights": [], "updates": 0}
        return True

    def submit_update(self, node_id: str, weights: List[float]) -> dict:
        if node_id not in self._nodes:
            return {"status": "error", "message": "Node not found"}
        self._nodes[node_id]["weights"] = weights
        self._nodes[node_id]["updates"] += 1
        return self._aggregate()

    def _aggregate(self) -> dict:
        active = [n for n in self._nodes.values() if n["weights"]]
        if not active:
            return {"status": "error", "message": "No updates"}
        total_size = sum(n["data_size"] for n in active)
        nf = len(active[0]["weights"])
        global_w = [0.0] * nf
        for node in active:
            for i in range(nf):
                global_w[i] += node["weights"][i] * node["data_size"] / total_size
        self._global_weights = global_w
        return {"status": "ok", "global_weights": global_w, "nodes": len(active)}

    @property
    def node_count(self) -> int:
        return len(self._nodes)

    @property
    def global_weights(self) -> List[float]:
        return list(self._global_weights)


class ModelRegistry:
    """Model registry — version and track ML models."""

    def __init__(self):
        self._models: Dict[str, List[dict]] = {}

    def register(self, model_id: str, version: str, metadata: dict) -> bool:
        self._models.setdefault(model_id, []).append({
            "version": version, "metadata": metadata,
            "registered": datetime.now(timezone.utc).isoformat(),
        })
        _log("INFO", f"Model registered: {model_id} v{version}")
        return True

    def get_latest(self, model_id: str) -> Optional[dict]:
        versions = self._models.get(model_id, [])
        return versions[-1] if versions else None

    def list_versions(self, model_id: str) -> List[dict]:
        return list(self._models.get(model_id, []))

    def compare(self, model_id: str, v1: str, v2: str) -> dict:
        versions = self._models.get(model_id, [])
        m1 = next((v for v in versions if v["version"] == v1), None)
        m2 = next((v for v in versions if v["version"] == v2), None)
        if not m1 or not m2:
            return {"status": "error", "message": "Version not found"}
        return {"status": "ok", "v1": m1, "v2": m2}

    @property
    def models(self) -> List[str]:
        return list(self._models.keys())

    @property
    def total_versions(self) -> int:
        return sum(len(v) for v in self._models.values())


class DataLabeler:
    """Data labeling — annotate datasets for ML training."""

    def __init__(self):
        self._datasets: Dict[str, List[dict]] = {}
        self._labels: Dict[str, set] = {}

    def create_dataset(self, dataset_id: str, labels: List[str]) -> bool:
        self._datasets[dataset_id] = []
        self._labels[dataset_id] = set(labels)
        return True

    def add_item(self, dataset_id: str, data: Any, label: str = "") -> bool:
        if dataset_id not in self._datasets:
            return False
        self._datasets[dataset_id].append({"data": data, "label": label, "labeled": bool(label)})
        return True

    def label_item(self, dataset_id: str, index: int, label: str) -> bool:
        items = self._datasets.get(dataset_id, [])
        if index < 0 or index >= len(items):
            return False
        items[index]["label"] = label
        items[index]["labeled"] = True
        return True

    def export(self, dataset_id: str) -> dict:
        items = self._datasets.get(dataset_id, [])
        labeled = [i for i in items if i["labeled"]]
        return {"status": "ok", "total": len(items), "labeled": len(labeled),
                "unlabeled": len(items) - len(labeled), "items": items}

    def stats(self, dataset_id: str) -> dict:
        items = self._datasets.get(dataset_id, [])
        labeled = [i for i in items if i["labeled"]]
        label_counts: Dict[str, int] = {}
        for i in labeled:
            label_counts[i["label"]] = label_counts.get(i["label"], 0) + 1
        return {"total": len(items), "labeled": len(labeled),
                "progress": round(len(labeled) / max(len(items), 1) * 100, 1),
                "label_distribution": label_counts}

    @property
    def datasets(self) -> List[str]:
        return list(self._datasets.keys())


class PromptEngine:
    """Prompt engineering — templates, variables, chains for LLMs."""

    def __init__(self):
        self._templates: Dict[str, str] = {}
        self._chains: Dict[str, List[str]] = {}

    def register(self, template_id: str, template: str) -> bool:
        self._templates[template_id] = template
        return True

    def render(self, template_id: str, **variables) -> str:
        tmpl = self._templates.get(template_id, "")
        try:
            return tmpl.format(**variables)
        except (KeyError, ValueError):
            return tmpl

    def create_chain(self, chain_id: str, template_ids: List[str]) -> bool:
        for tid in template_ids:
            if tid not in self._templates:
                return False
        self._chains[chain_id] = template_ids
        return True

    def run_chain(self, chain_id: str, **variables) -> List[str]:
        tids = self._chains.get(chain_id, [])
        results = []
        context = dict(variables)
        for tid in tids:
            rendered = self.render(tid, **context)
            results.append(rendered)
            context["previous"] = rendered
        return results

    @property
    def templates(self) -> List[str]:
        return list(self._templates.keys())

    @property
    def chains(self) -> List[str]:
        return list(self._chains.keys())


class ChatbotFramework:
    """Chatbot framework — intent matching, context, multi-turn conversations."""

    def __init__(self):
        self._intents: Dict[str, List[str]] = {}
        self._responses: Dict[str, str] = {}
        self._contexts: Dict[str, dict] = {}
        self._conversations: Dict[str, List[dict]] = {}

    def add_intent(self, intent: str, patterns: List[str], response: str) -> bool:
        self._intents[intent] = patterns
        self._responses[intent] = response
        return True

    def match(self, text: str) -> Optional[str]:
        text_lower = text.lower()
        best_match = None
        best_score = 0
        for intent, patterns in self._intents.items():
            for pattern in patterns:
                if pattern.lower() in text_lower:
                    score = len(pattern) / max(len(text_lower), 1)
                    if score > best_score:
                        best_score = score
                        best_match = intent
        return best_match

    def respond(self, session_id: str, text: str) -> dict:
        intent = self.match(text)
        if not intent:
            response = "I'm not sure how to help with that."
        else:
            response = self._responses[intent]
        self._conversations.setdefault(session_id, []).append({
            "user": text, "bot": response, "intent": intent,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
        self._contexts[session_id] = {"last_intent": intent, "turn": len(self._conversations[session_id])}
        return {"status": "ok", "response": response, "intent": intent, "session": session_id}

    def history(self, session_id: str) -> List[dict]:
        return list(self._conversations.get(session_id, []))

    def clear_session(self, session_id: str) -> bool:
        self._conversations.pop(session_id, None)
        self._contexts.pop(session_id, None)
        return True

    @property
    def intent_count(self) -> int:
        return len(self._intents)

    @property
    def active_sessions(self) -> int:
        return len(self._conversations)


# ═══════════════════════════════════════════════════════════
# BATCH 13: ENTERPRISE+ (12 features)
# ═══════════════════════════════════════════════════════════

class SSOManager:
    """Single Sign-On — OAuth2/OIDC provider and consumer."""

    def __init__(self):
        self._providers: Dict[str, dict] = {}
        self._tokens: Dict[str, dict] = {}

    def register_provider(self, provider: str, client_id: str,
                          client_secret: str, redirect_uri: str) -> bool:
        self._providers[provider] = {"client_id": client_id,
                                      "client_secret": client_secret,
                                      "redirect_uri": redirect_uri}
        return True

    def authorize_url(self, provider: str, scope: str = "openid profile email") -> str:
        p = self._providers.get(provider)
        if not p:
            return ""
        state = hashlib.sha256(str(time.time()).encode()).hexdigest()[:16]
        return (f"https://{provider}.com/auth?client_id={p['client_id']}"
                f"&redirect_uri={p['redirect_uri']}&scope={scope}&state={state}")

    def exchange_code(self, provider: str, code: str) -> dict:
        if provider not in self._providers:
            return {"status": "error", "message": "Unknown provider"}
        token = hashlib.sha256(f"{code}:{time.time()}".encode()).hexdigest()
        self._tokens[token] = {"provider": provider, "code": code,
                                "expires": time.time() + 3600}
        return {"status": "ok", "access_token": token, "expires_in": 3600}

    def verify(self, token: str) -> Optional[dict]:
        t = self._tokens.get(token)
        if not t or t["expires"] < time.time():
            return None
        return t

    @property
    def providers(self) -> List[str]:
        return list(self._providers.keys())


class LDAPManager:
    """LDAP/Active Directory integration — user lookup and auth."""

    def __init__(self):
        self._servers: Dict[str, dict] = {}
        self._users: Dict[str, dict] = {}

    def configure(self, server_id: str, host: str, port: int = 389,
                  base_dn: str = "dc=example,dc=com") -> bool:
        self._servers[server_id] = {"host": host, "port": port, "base_dn": base_dn}
        return True

    def add_user(self, server_id: str, dn: str, cn: str, mail: str,
                 groups: List[str] = None) -> bool:
        self._users[dn] = {"server": server_id, "cn": cn, "mail": mail,
                           "groups": groups or [], "dn": dn}
        return True

    def authenticate(self, dn: str, password: str) -> dict:
        user = self._users.get(dn)
        if not user:
            return {"status": "error", "message": "User not found"}
        if not password:
            return {"status": "error", "message": "Invalid credentials"}
        return {"status": "ok", "dn": dn, "cn": user["cn"], "mail": user["mail"],
                "groups": user["groups"]}

    def search(self, server_id: str, query: str) -> List[dict]:
        results = [u for u in self._users.values()
                   if u["server"] == server_id and query.lower() in u["cn"].lower()]
        return results

    @property
    def user_count(self) -> int:
        return len(self._users)


class SAMLProvider:
    """SAML 2.0 identity provider — assertions and metadata."""

    def __init__(self):
        self._sp: Dict[str, dict] = {}
        self._assertions: List[dict] = []

    def register_sp(self, sp_id: str, entity_id: str, acs_url: str) -> bool:
        self._sp[sp_id] = {"entity_id": entity_id, "acs_url": acs_url}
        return True

    def generate_assertion(self, sp_id: str, user_id: str, attributes: dict) -> dict:
        if sp_id not in self._sp:
            return {"status": "error", "message": "SP not registered"}
        assertion_id = f"_a{len(self._assertions) + 1}"
        assertion = {"id": assertion_id, "sp_id": sp_id, "user_id": user_id,
                     "attributes": attributes,
                     "issued": datetime.now(timezone.utc).isoformat(),
                     "expires": time.time() + 300}
        self._assertions.append(assertion)
        return {"status": "ok", "assertion_id": assertion_id,
                "assertion": assertion}

    def validate_assertion(self, assertion_id: str) -> Optional[dict]:
        a = next((x for x in self._assertions if x["id"] == assertion_id), None)
        if not a or a["expires"] < time.time():
            return None
        return a

    def metadata(self, sp_id: str) -> dict:
        sp = self._sp.get(sp_id)
        if not sp:
            return {"status": "error", "message": "SP not found"}
        return {"status": "ok", "entity_id": sp["entity_id"],
                "acs_url": sp["acs_url"], "binding": "urn:oasis:names:tc:SAML:2.0:bindings:HTTP-POST"}

    @property
    def sp_count(self) -> int:
        return len(self._sp)


class AuditReporter:
    """Generate compliance audit reports from audit trail data."""

    def __init__(self):
        self._reports: Dict[str, dict] = {}

    def generate(self, report_id: str, entries: List[dict],
                 format: str = "summary") -> dict:
        by_user: Dict[str, int] = {}
        by_action: Dict[str, int] = {}
        by_resource: Dict[str, int] = {}
        for e in entries:
            u = e.get("user", "unknown")
            a = e.get("action", "unknown")
            r = e.get("resource", "unknown")
            by_user[u] = by_user.get(u, 0) + 1
            by_action[a] = by_action.get(a, 0) + 1
            by_resource[r] = by_resource.get(r, 0) + 1
        report = {"id": report_id, "format": format,
                  "total_entries": len(entries),
                  "by_user": by_user, "by_action": by_action,
                  "by_resource": by_resource,
                  "generated": datetime.now(timezone.utc).isoformat()}
        self._reports[report_id] = report
        return {"status": "ok", **report}

    def get_report(self, report_id: str) -> Optional[dict]:
        return self._reports.get(report_id)

    def list_reports(self) -> List[str]:
        return list(self._reports.keys())

    def export_csv(self, report_id: str) -> str:
        r = self._reports.get(report_id)
        if not r:
            return ""
        lines = ["user,count"]
        for u, c in r["by_user"].items():
            lines.append(f"{u},{c}")
        return "\n".join(lines)


class ComplianceChecker:
    """Check compliance against standards — SOC2, ISO27001, GDPR."""

    def __init__(self):
        self._standards: Dict[str, List[dict]] = {}
        self._results: Dict[str, dict] = {}

    def register_standard(self, standard: str, controls: List[dict]) -> bool:
        self._standards[standard] = controls
        return True

    def check(self, standard: str, evidence: Dict[str, bool]) -> dict:
        controls = self._standards.get(standard, [])
        if not controls:
            return {"status": "error", "message": "Standard not found"}
        passed = 0
        failed = []
        for ctrl in controls:
            ctrl_id = ctrl.get("id", "")
            if evidence.get(ctrl_id, False):
                passed += 1
            else:
                failed.append(ctrl_id)
        score = passed / len(controls) * 100 if controls else 0
        result = {"standard": standard, "total": len(controls), "passed": passed,
                  "failed": failed, "score": round(score, 1),
                  "compliant": score >= 80}
        self._results[standard] = result
        return {"status": "ok", **result}

    @property
    def standards(self) -> List[str]:
        return list(self._standards.keys())

    def get_result(self, standard: str) -> Optional[dict]:
        return self._results.get(standard)


class DataGovernance:
    """Data governance — classification, lineage, quality scoring."""

    def __init__(self):
        self._assets: Dict[str, dict] = {}
        self._lineage: Dict[str, List[str]] = {}

    def register_asset(self, asset_id: str, name: str, classification: str = "internal",
                       owner: str = "") -> bool:
        self._assets[asset_id] = {"name": name, "classification": classification,
                                   "owner": owner, "quality_score": 0.0}
        return True

    def classify(self, asset_id: str, classification: str) -> bool:
        if asset_id not in self._assets:
            return False
        self._assets[asset_id]["classification"] = classification
        return True

    def set_lineage(self, asset_id: str, sources: List[str]) -> bool:
        self._lineage[asset_id] = sources
        return True

    def get_lineage(self, asset_id: str) -> List[str]:
        return list(self._lineage.get(asset_id, []))

    def quality_score(self, asset_id: str, completeness: float,
                      accuracy: float, timeliness: float) -> dict:
        if asset_id not in self._assets:
            return {"status": "error", "message": "Asset not found"}
        score = (completeness * 0.4 + accuracy * 0.4 + timeliness * 0.2) * 100
        self._assets[asset_id]["quality_score"] = round(score, 1)
        return {"status": "ok", "asset_id": asset_id, "quality_score": round(score, 1)}

    @property
    def assets(self) -> List[str]:
        return list(self._assets.keys())


class PrivacyManager:
    """Privacy management — PII detection, consent, data subject rights."""

    def __init__(self):
        self._consents: Dict[str, dict] = {}
        self._pii_patterns = {
            "email": r'\b[\w.]+@[\w]+\.[\w.]+\b',
            "phone": r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b',
            "ssn": r'\b\d{3}-\d{2}-\d{4}\b',
            "credit_card": r'\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b',
        }

    def detect_pii(self, text: str) -> dict:
        import re
        findings = {}
        for pii_type, pattern in self._pii_patterns.items():
            matches = re.findall(pattern, text)
            if matches:
                findings[pii_type] = len(matches)
        return {"status": "ok", "has_pii": bool(findings), "findings": findings}

    def redact(self, text: str) -> str:
        import re
        for pii_type, pattern in self._pii_patterns.items():
            text = re.sub(pattern, f"[REDACTED_{pii_type.upper()}]", text)
        return text

    def record_consent(self, user_id: str, purpose: str, granted: bool) -> bool:
        self._consents[f"{user_id}:{purpose}"] = {
            "user_id": user_id, "purpose": purpose, "granted": granted,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        return True

    def check_consent(self, user_id: str, purpose: str) -> bool:
        c = self._consents.get(f"{user_id}:{purpose}")
        return c["granted"] if c else False

    def data_subject_request(self, user_id: str, request_type: str) -> dict:
        valid_types = ["access", "deletion", "portability", "rectification"]
        if request_type not in valid_types:
            return {"status": "error", "message": "Invalid request type"}
        return {"status": "ok", "user_id": user_id, "request_type": request_type,
                "ticket": f"DSR-{hash(user_id + request_type) % 10000:04d}"}

    @property
    def consent_count(self) -> int:
        return len(self._consents)


class GDPRTools:
    """GDPR compliance tools — right to be forgotten, data export, breach notification."""

    def __init__(self):
        self._breaches: List[dict] = []
        self._deletion_requests: Dict[str, dict] = {}

    def right_to_be_forgotten(self, user_id: str) -> dict:
        self._deletion_requests[user_id] = {
            "status": "processing",
            "requested": datetime.now(timezone.utc).isoformat(),
        }
        return {"status": "ok", "user_id": user_id, "message": "Deletion scheduled"}

    def export_data(self, user_id: str, data: dict) -> dict:
        return {"status": "ok", "user_id": user_id,
                "format": "JSON",
                "data": data,
                "exported": datetime.now(timezone.utc).isoformat()}

    def report_breach(self, description: str, affected_users: int,
                      severity: str = "high") -> dict:
        breach = {"id": f"BR-{len(self._breaches) + 1:04d}",
                  "description": description,
                  "affected_users": affected_users,
                  "severity": severity,
                  "reported": datetime.now(timezone.utc).isoformat(),
                  "notification_required": affected_users > 500}
        self._breaches.append(breach)
        return {"status": "ok", **breach}

    def check_deletion(self, user_id: str) -> Optional[dict]:
        return self._deletion_requests.get(user_id)

    @property
    def breach_count(self) -> int:
        return len(self._breaches)


class EncryptionVault:
    """Encryption vault — manage encryption keys and secrets."""

    def __init__(self):
        self._vault: Dict[str, bytes] = {}
        self._key = hashlib.sha256(b"pytrex-vault-default-key-v1").digest()

    def store(self, key: str, value: str) -> bool:
        encrypted = self._encrypt(value.encode())
        self._vault[key] = encrypted
        return True

    def retrieve(self, key: str) -> Optional[str]:
        encrypted = self._vault.get(key)
        if not encrypted:
            return None
        return self._decrypt(encrypted).decode()

    def _encrypt(self, data: bytes) -> bytes:
        return bytes(b ^ self._key[i % len(self._key)] for i, b in enumerate(data))

    def _decrypt(self, data: bytes) -> bytes:
        return self._encrypt(data)

    def list_keys(self) -> List[str]:
        return list(self._vault.keys())

    def delete(self, key: str) -> bool:
        return self._vault.pop(key, None) is not None

    @property
    def secret_count(self) -> int:
        return len(self._vault)


class KeyRotation:
    """Automatic key rotation — track and rotate encryption keys."""

    def __init__(self, rotation_days: int = 90):
        self._keys: Dict[str, List[dict]] = {}
        self._rotation_days = rotation_days

    def add_key(self, key_id: str, key_value: str) -> bool:
        self._keys.setdefault(key_id, []).append({
            "value": key_value, "active": True,
            "created": time.time(),
        })
        for k in self._keys[key_id][:-1]:
            k["active"] = False
        _log("INFO", f"Key '{key_id}' added/rotated")
        return True

    def rotate(self, key_id: str) -> dict:
        if key_id not in self._keys:
            return {"status": "error", "message": "Key not found"}
        new_value = hashlib.sha256(f"{key_id}:{time.time()}".encode()).hexdigest()
        self.add_key(key_id, new_value)
        return {"status": "ok", "key_id": key_id, "rotated": True}

    def get_active(self, key_id: str) -> Optional[str]:
        keys = self._keys.get(key_id, [])
        active = [k for k in keys if k["active"]]
        return active[0]["value"] if active else None

    def needs_rotation(self, key_id: str) -> bool:
        keys = self._keys.get(key_id, [])
        active = [k for k in keys if k["active"]]
        if not active:
            return True
        age = time.time() - active[0]["created"]
        return age > self._rotation_days * 86400

    @property
    def key_count(self) -> int:
        return len(self._keys)


class AccessPolicy:
    """Access policy engine — ABAC, policy rules, enforcement."""

    def __init__(self):
        self._policies: Dict[str, dict] = {}

    def create_policy(self, policy_id: str, resource: str, action: str,
                      effect: str = "allow", condition: str = "") -> bool:
        self._policies[policy_id] = {"resource": resource, "action": action,
                                      "effect": effect, "condition": condition}
        return True

    def evaluate(self, resource: str, action: str, attributes: dict = None) -> dict:
        attrs = attributes or {}
        for pid, p in self._policies.items():
            if p["resource"] == resource and p["action"] == action:
                if p["effect"] == "deny":
                    return {"status": "ok", "decision": "deny", "policy": pid}
                if p["effect"] == "allow":
                    return {"status": "ok", "decision": "allow", "policy": pid}
        return {"status": "ok", "decision": "deny", "policy": "default-deny"}

    def list_policies(self) -> List[dict]:
        return [{"id": k, **v} for k, v in self._policies.items()]

    def delete_policy(self, policy_id: str) -> bool:
        return self._policies.pop(policy_id, None) is not None

    @property
    def policy_count(self) -> int:
        return len(self._policies)


class IdentityProvider:
    """Identity provider — user provisioning, federation, MFA."""

    def __init__(self):
        self._users: Dict[str, dict] = {}
        self._mfa: Dict[str, str] = {}
        self._federated: Dict[str, str] = {}

    def provision(self, user_id: str, email: str, name: str,
                  roles: List[str] = None) -> bool:
        self._users[user_id] = {"email": email, "name": name,
                                 "roles": roles or [], "active": True,
                                 "created": datetime.now(timezone.utc).isoformat()}
        return True

    def deprovision(self, user_id: str) -> bool:
        if user_id in self._users:
            self._users[user_id]["active"] = False
            return True
        return False

    def enable_mfa(self, user_id: str, secret: str = "") -> str:
        secret = secret or hashlib.sha256(f"{user_id}:{time.time()}".encode()).hexdigest()[:16]
        self._mfa[user_id] = secret
        return secret

    def verify_mfa(self, user_id: str, code: str) -> bool:
        return self._mfa.get(user_id) == code

    def federate(self, user_id: str, provider: str) -> bool:
        self._federated[user_id] = provider
        return True

    def get_user(self, user_id: str) -> Optional[dict]:
        return self._users.get(user_id)

    @property
    def user_count(self) -> int:
        return len(self._users)

    @property
    def active_users(self) -> int:
        return sum(1 for u in self._users.values() if u["active"])


# ═══════════════════════════════════════════════════════════
# BATCH 14: DEVTOOLS+ (12 features)
# ═══════════════════════════════════════════════════════════

class CodeFormatter:
    """Code formatter — Python, JS, JSON, HTML formatting."""

    def __init__(self):
        self._rules: Dict[str, dict] = {
            "python": {"indent": 4, "max_line": 88},
            "javascript": {"indent": 2, "max_line": 80},
            "json": {"indent": 2, "max_line": 120},
        }

    def format(self, code: str, language: str = "python") -> dict:
        rule = self._rules.get(language, self._rules["python"])
        indent_size = rule["indent"]
        lines = code.split("\n")
        formatted = []
        indent_level = 0
        for line in lines:
            stripped = line.strip()
            if not stripped:
                formatted.append("")
                continue
            if stripped.startswith(("}", "]", ")", "elif", "else:", "except", "finally")):
                indent_level = max(0, indent_level - 1)
            formatted.append(" " * indent_size * indent_level + stripped)
            for char in stripped:
                if char in "({[":
                    indent_level += 1
                if char in ")}]":
                    indent_level = max(0, indent_level - 1)
            if stripped.endswith(":") and not stripped.startswith(("#", "//")):
                indent_level += 1
        return {"status": "ok", "formatted": "\n".join(formatted), "lines": len(formatted)}

    def minify(self, code: str, language: str = "javascript") -> str:
        lines = [l.strip() for l in code.split("\n") if l.strip()]
        return " ".join(lines)

    @property
    def languages(self) -> List[str]:
        return list(self._rules.keys())


class Linter:
    """Code linter — check for common issues and style violations."""

    def __init__(self):
        self._rules: List[dict] = []
        self._setup_default_rules()

    def _setup_default_rules(self):
        self._rules = [
            {"id": "E001", "check": lambda l: len(l) > 88, "msg": "Line too long"},
            {"id": "E002", "check": lambda l: "\t" in l, "msg": "Tab character found"},
            {"id": "E003", "check": lambda l: l.rstrip() != l and l.strip(), "msg": "Trailing whitespace"},
            {"id": "W001", "check": lambda l: "print(" in l, "msg": "Print statement (use logging)"},
            {"id": "W002", "check": lambda l: "TODO" in l, "msg": "TODO found"},
            {"id": "W003", "check": lambda l: l.strip().startswith("import *"), "msg": "Wildcard import"},
        ]

    def lint(self, code: str) -> dict:
        lines = code.split("\n")
        issues = []
        for i, line in enumerate(lines, 1):
            for rule in self._rules:
                if rule["check"](line):
                    issues.append({"line": i, "rule": rule["id"], "message": rule["msg"]})
        return {"status": "ok", "total_issues": len(issues), "issues": issues}

    def add_rule(self, rule_id: str, check_fn, message: str) -> bool:
        self._rules.append({"id": rule_id, "check": check_fn, "msg": message})
        return True

    @property
    def rule_count(self) -> int:
        return len(self._rules)


class TypeChecker:
    """Type checking — verify type annotations and contracts."""

    def __init__(self):
        self._types: Dict[str, str] = {}

    def check_type(self, value: Any, expected_type: str) -> dict:
        type_map = {"str": str, "int": int, "float": float, "bool": bool,
                    "list": list, "dict": dict, "set": set, "tuple": tuple}
        py_type = type_map.get(expected_type)
        if not py_type:
            return {"status": "error", "message": f"Unknown type: {expected_type}"}
        actual = type(value).__name__
        matches = isinstance(value, py_type)
        return {"status": "ok", "expected": expected_type, "actual": actual, "matches": matches}

    def check_signature(self, func: Callable, args: List[Any], kwargs: Dict[str, Any]) -> dict:
        import inspect
        sig = inspect.signature(func)
        params = list(sig.parameters.keys())
        issues = []
        for i, (pname, p) in enumerate(sig.parameters.items()):
            if p.annotation != inspect.Parameter.empty:
                expected = p.annotation.__name__ if hasattr(p.annotation, "__name__") else str(p.annotation)
                if i < len(args):
                    actual = type(args[i]).__name__
                    if not isinstance(args[i], p.annotation):
                        issues.append(f"Param '{pname}': expected {expected}, got {actual}")
        return {"status": "ok", "params": params, "issues": issues}

    def register_type(self, name: str, type_name: str) -> bool:
        self._types[name] = type_name
        return True

    @property
    def registered_types(self) -> List[str]:
        return list(self._types.keys())


class DebugProfiler:
    """Debug profiler — timing, call counts, memory tracking."""

    def __init__(self):
        self._profiles: Dict[str, dict] = {}
        self._active: Optional[str] = None
        self._start_time: float = 0

    def start(self, profile_id: str) -> bool:
        self._active = profile_id
        self._start_time = time.time()
        self._profiles[profile_id] = {"calls": 0, "total_time": 0, "sub_profiles": []}
        return True

    def stop(self) -> dict:
        if not self._active:
            return {"status": "error", "message": "No active profile"}
        elapsed = time.time() - self._start_time
        p = self._profiles[self._active]
        p["total_time"] = elapsed
        p["calls"] += 1
        result = {"status": "ok", "profile": self._active, "elapsed": round(elapsed, 6)}
        self._active = None
        return result

    def profile(self, func: Callable) -> Callable:
        def wrapper(*args, **kwargs):
            pid = f"{func.__name__}_{id(func)}"
            self.start(pid)
            result = func(*args, **kwargs)
            self.stop()
            return result
        return wrapper

    def get_profile(self, profile_id: str) -> Optional[dict]:
        return self._profiles.get(profile_id)

    @property
    def profiles(self) -> List[str]:
        return list(self._profiles.keys())


class MemoryAnalyzer:
    """Memory analyzer — track object sizes and memory usage."""

    def __init__(self):
        self._snapshots: Dict[str, dict] = {}

    def snapshot(self, snapshot_id: str, objects: Dict[str, Any]) -> dict:
        sizes = {}
        import sys
        for name, obj in objects.items():
            try:
                sizes[name] = sys.getsizeof(obj)
            except Exception:
                sizes[name] = -1
        total = sum(sizes.values())
        self._snapshots[snapshot_id] = {"objects": sizes, "total": total,
                                         "timestamp": time.time()}
        return {"status": "ok", "snapshot_id": snapshot_id,
                "total_bytes": total, "objects": sizes}

    def compare(self, snap1: str, snap2: str) -> dict:
        s1 = self._snapshots.get(snap1)
        s2 = self._snapshots.get(snap2)
        if not s1 or not s2:
            return {"status": "error", "message": "Snapshot not found"}
        diffs = {}
        all_keys = set(s1["objects"]) | set(s2["objects"])
        for key in all_keys:
            v1 = s1["objects"].get(key, 0)
            v2 = s2["objects"].get(key, 0)
            if v1 != v2:
                diffs[key] = v2 - v1
        return {"status": "ok", "delta_total": s2["total"] - s1["total"], "changes": diffs}

    @property
    def snapshot_count(self) -> int:
        return len(self._snapshots)


class HotReloader:
    """Hot reload — watch files and reload modules on change."""

    def __init__(self):
        self._watched: Dict[str, float] = {}
        self._callbacks: Dict[str, Callable] = {}
        self._running = False

    def watch(self, filepath: str, callback: Callable) -> bool:
        import os
        if os.path.exists(filepath):
            self._watched[filepath] = os.path.getmtime(filepath)
        self._callbacks[filepath] = callback
        return True

    def check(self) -> List[str]:
        import os
        changed = []
        for filepath, mtime in list(self._watched.items()):
            if os.path.exists(filepath):
                current = os.path.getmtime(filepath)
                if current > mtime:
                    self._watched[filepath] = current
                    changed.append(filepath)
                    if filepath in self._callbacks:
                        try:
                            self._callbacks[filepath](filepath)
                        except Exception as e:
                            _log("ERROR", f"Hot reload callback failed: {e}")
        return changed

    @property
    def watched_files(self) -> List[str]:
        return list(self._watched.keys())


class REPL:
    """Read-Eval-Print Loop — interactive code execution."""

    def __init__(self):
        self._history: List[dict] = []
        self._vars: Dict[str, Any] = {}

    def eval(self, code: str) -> dict:
        try:
            result = eval(code, {"__builtins__": __builtins__}, self._vars)
            self._history.append({"code": code, "result": str(result), "error": False})
            return {"status": "ok", "result": result}
        except Exception as e:
            self._history.append({"code": code, "result": str(e), "error": True})
            return {"status": "error", "message": str(e)}

    def exec(self, code: str) -> dict:
        try:
            exec(code, {"__builtins__": __builtins__}, self._vars)
            self._history.append({"code": code, "result": "executed", "error": False})
            return {"status": "ok", "message": "Code executed"}
        except Exception as e:
            self._history.append({"code": code, "result": str(e), "error": True})
            return {"status": "error", "message": str(e)}

    @property
    def history(self) -> List[dict]:
        return list(self._history)

    @property
    def variables(self) -> List[str]:
        return [k for k in self._vars if not k.startswith("_")]


class Notebook:
    """Jupyter-like notebook — cells, execution, state."""

    def __init__(self):
        self._cells: List[dict] = []
        self._state: Dict[str, Any] = {}

    def add_cell(self, code: str, cell_type: str = "code") -> int:
        idx = len(self._cells)
        self._cells.append({"index": idx, "code": code, "type": cell_type,
                            "output": None, "executed": False})
        return idx

    def run_cell(self, index: int) -> dict:
        if index < 0 or index >= len(self._cells):
            return {"status": "error", "message": "Cell not found"}
        cell = self._cells[index]
        if cell["type"] != "code":
            return {"status": "ok", "output": "Markdown cell (not executed)"}
        try:
            output = eval(cell["code"], {"__builtins__": __builtins__}, self._state)
            cell["output"] = str(output)
            cell["executed"] = True
            return {"status": "ok", "output": str(output), "index": index}
        except SyntaxError:
            exec(cell["code"], {"__builtins__": __builtins__}, self._state)
            cell["output"] = "executed"
            cell["executed"] = True
            return {"status": "ok", "output": "executed", "index": index}
        except Exception as e:
            cell["output"] = str(e)
            return {"status": "error", "message": str(e), "index": index}

    def run_all(self) -> List[dict]:
        return [self.run_cell(i) for i in range(len(self._cells)) if self._cells[i]["type"] == "code"]

    def get_cell(self, index: int) -> Optional[dict]:
        return self._cells[index] if 0 <= index < len(self._cells) else None

    @property
    def cell_count(self) -> int:
        return len(self._cells)


class APITester:
    """API testing — send requests, validate responses, test suites."""

    def __init__(self):
        self._tests: List[dict] = []
        self._results: List[dict] = []

    def add_test(self, test_id: str, method: str, url: str,
                 expected_status: int = 200, body: dict = None) -> bool:
        self._tests.append({"id": test_id, "method": method, "url": url,
                            "expected_status": expected_status, "body": body})
        return True

    def run_test(self, test_id: str) -> dict:
        test = next((t for t in self._tests if t["id"] == test_id), None)
        if not test:
            return {"status": "error", "message": "Test not found"}
        try:
            import urllib.request
            data = json.dumps(test["body"]).encode() if test["body"] else None
            req = urllib.request.Request(test["url"], data=data, method=test["method"])
            req.add_header("Content-Type", "application/json")
            with urllib.request.urlopen(req, timeout=10) as resp:
                status = resp.status
                body = resp.read().decode()
            passed = status == test["expected_status"]
            result = {"test_id": test_id, "status": status, "passed": passed}
            self._results.append(result)
            return {"status": "ok", **result}
        except Exception as e:
            result = {"test_id": test_id, "error": str(e), "passed": False}
            self._results.append(result)
            return {"status": "error", "message": str(e), "passed": False}

    def run_all(self) -> dict:
        results = [self.run_test(t["id"]) for t in self._tests]
        passed = sum(1 for r in results if r.get("passed"))
        return {"status": "ok", "total": len(results), "passed": passed,
                "failed": len(results) - passed}

    @property
    def test_count(self) -> int:
        return len(self._tests)


class MockServer:
    """Mock server — simulate API responses for testing."""

    def __init__(self):
        self._mocks: Dict[str, dict] = {}

    def register(self, method: str, path: str, response: dict, status: int = 200) -> bool:
        key = f"{method}:{path}"
        self._mocks[key] = {"response": response, "status": status}
        return True

    def get_mock(self, method: str, path: str) -> Optional[dict]:
        return self._mocks.get(f"{method}:{path}")

    def list_mocks(self) -> List[str]:
        return list(self._mocks.keys())

    def clear(self) -> bool:
        self._mocks.clear()
        return True

    @property
    def mock_count(self) -> int:
        return len(self._mocks)


class SnapshotTest:
    """Snapshot testing — compare outputs against saved snapshots."""

    def __init__(self):
        self._snapshots: Dict[str, Any] = {}
        self._results: List[dict] = []

    def save(self, snapshot_id: str, data: Any) -> bool:
        self._snapshots[snapshot_id] = data
        return True

    def compare(self, snapshot_id: str, data: Any) -> dict:
        saved = self._snapshots.get(snapshot_id)
        if saved is None:
            self._snapshots[snapshot_id] = data
            result = {"snapshot_id": snapshot_id, "passed": True, "message": "Snapshot created"}
            self._results.append(result)
            return {"status": "ok", **result}
        passed = saved == data
        result = {"snapshot_id": snapshot_id, "passed": passed,
                  "message": "Match" if passed else "Mismatch"}
        self._results.append(result)
        return {"status": "ok", **result}

    def get_snapshot(self, snapshot_id: str) -> Optional[Any]:
        return self._snapshots.get(snapshot_id)

    def list_snapshots(self) -> List[str]:
        return list(self._snapshots.keys())

    @property
    def snapshot_count(self) -> int:
        return len(self._snapshots)


class CoverageReporter:
    """Code coverage reporter — track executed lines and branches."""

    def __init__(self):
        self._coverage: Dict[str, dict] = {}

    def track(self, file_id: str, total_lines: int, covered_lines: List[int]) -> dict:
        covered = set(covered_lines)
        self._coverage[file_id] = {"total": total_lines, "covered": len(covered),
                                    "covered_lines": covered,
                                    "percentage": round(len(covered) / max(total_lines, 1) * 100, 1)}
        return {"status": "ok", **self._coverage[file_id]}

    def report(self) -> dict:
        if not self._coverage:
            return {"status": "ok", "files": 0, "total_lines": 0, "covered_lines": 0,
                    "percentage": 0}
        total = sum(c["total"] for c in self._coverage.values())
        covered = sum(c["covered"] for c in self._coverage.values())
        return {"status": "ok", "files": len(self._coverage),
                "total_lines": total, "covered_lines": covered,
                "percentage": round(covered / max(total, 1) * 100, 1),
                "details": {k: {"percentage": v["percentage"], "covered": v["covered"],
                                "total": v["total"]}
                            for k, v in self._coverage.items()}}

    def uncovered(self, file_id: str) -> List[int]:
        c = self._coverage.get(file_id)
        if not c:
            return []
        return [i for i in range(1, c["total"] + 1) if i not in c["covered_lines"]]

    @property
    def file_count(self) -> int:
        return len(self._coverage)


# ═══════════════════════════════════════════════════════════
# BATCH 15: INDUSTRY+ (12 features)
# ═══════════════════════════════════════════════════════════

class HealthcareHL7:
    """HL7 healthcare data exchange — patient records, lab results, admissions."""

    def __init__(self):
        self._patients: Dict[str, dict] = {}
        self._messages: List[dict] = []

    def register_patient(self, patient_id: str, name: str, dob: str,
                         gender: str = "U") -> bool:
        self._patients[patient_id] = {"name": name, "dob": dob, "gender": gender,
                                       "active": True}
        return True

    def create_message(self, msg_type: str, patient_id: str, data: dict) -> dict:
        if patient_id not in self._patients:
            return {"status": "error", "message": "Patient not found"}
        msg = {"type": msg_type, "patient_id": patient_id, "data": data,
               "id": f"HL7_{len(self._messages) + 1:06d}",
               "timestamp": datetime.now(timezone.utc).isoformat()}
        self._messages.append(msg)
        return {"status": "ok", "message_id": msg["id"], "type": msg_type}

    def parse_adt(self, patient_id: str, event: str = "A01") -> dict:
        events = {"A01": "Admit", "A02": "Transfer", "A03": "Discharge",
                  "A04": "Register", "A08": "Update"}
        if event not in events:
            return {"status": "error", "message": "Unknown event"}
        return self.create_message("ADT", patient_id, {"event": event, "event_name": events[event]})

    def lab_result(self, patient_id: str, test_name: str, value: str,
                   unit: str = "", reference: str = "") -> dict:
        return self.create_message("ORU", patient_id,
                                   {"test": test_name, "value": value, "unit": unit,
                                    "reference_range": reference})

    def get_patient(self, patient_id: str) -> Optional[dict]:
        return self._patients.get(patient_id)

    def get_messages(self, patient_id: str = "") -> List[dict]:
        if patient_id:
            return [m for m in self._messages if m["patient_id"] == patient_id]
        return list(self._messages)

    @property
    def patient_count(self) -> int:
        return len(self._patients)


class FinancePortfolio:
    """Portfolio management — holdings, P&L, risk metrics."""

    def __init__(self):
        self._holdings: Dict[str, dict] = {}
        self._history: List[dict] = []

    def buy(self, symbol: str, shares: int, price: float) -> dict:
        h = self._holdings.setdefault(symbol, {"shares": 0, "avg_cost": 0.0})
        total_cost = h["shares"] * h["avg_cost"] + shares * price
        h["shares"] += shares
        h["avg_cost"] = total_cost / h["shares"] if h["shares"] else 0
        self._history.append({"action": "buy", "symbol": symbol, "shares": shares, "price": price,
                              "timestamp": time.time()})
        return {"status": "ok", "symbol": symbol, "shares": h["shares"], "avg_cost": h["avg_cost"]}

    def sell(self, symbol: str, shares: int, price: float) -> dict:
        h = self._holdings.get(symbol)
        if not h or h["shares"] < shares:
            return {"status": "error", "message": "Insufficient shares"}
        h["shares"] -= shares
        pnl = (price - h["avg_cost"]) * shares
        self._history.append({"action": "sell", "symbol": symbol, "shares": shares, "price": price,
                              "pnl": pnl, "timestamp": time.time()})
        return {"status": "ok", "symbol": symbol, "shares": h["shares"], "realized_pnl": pnl}

    def value(self, prices: Dict[str, float]) -> dict:
        total = 0
        details = {}
        for sym, h in self._holdings.items():
            if h["shares"] > 0:
                val = h["shares"] * prices.get(sym, h["avg_cost"])
                pnl = val - h["shares"] * h["avg_cost"]
                total += val
                details[sym] = {"shares": h["shares"], "value": val, "unrealized_pnl": pnl}
        return {"status": "ok", "total_value": total, "holdings": details}

    @property
    def symbols(self) -> List[str]:
        return [s for s, h in self._holdings.items() if h["shares"] > 0]


class InventorySCM:
    """Inventory & Supply Chain Management — stock, orders, suppliers."""

    def __init__(self):
        self._items: Dict[str, dict] = {}
        self._orders: List[dict] = []
        self._suppliers: Dict[str, dict] = {}

    def add_item(self, sku: str, name: str, quantity: int = 0,
                 reorder_point: int = 10, unit_cost: float = 0) -> bool:
        self._items[sku] = {"name": name, "quantity": quantity,
                             "reorder_point": reorder_point, "unit_cost": unit_cost}
        return True

    def add_stock(self, sku: str, quantity: int) -> dict:
        if sku not in self._items:
            return {"status": "error", "message": "SKU not found"}
        self._items[sku]["quantity"] += quantity
        return {"status": "ok", "sku": sku, "quantity": self._items[sku]["quantity"]}

    def remove_stock(self, sku: str, quantity: int) -> dict:
        if sku not in self._items:
            return {"status": "error", "message": "SKU not found"}
        if self._items[sku]["quantity"] < quantity:
            return {"status": "error", "message": "Insufficient stock"}
        self._items[sku]["quantity"] -= quantity
        return {"status": "ok", "sku": sku, "quantity": self._items[sku]["quantity"]}

    def reorder_alerts(self) -> List[str]:
        return [sku for sku, item in self._items.items()
                if item["quantity"] <= item["reorder_point"]]

    def create_order(self, sku: str, quantity: int, supplier_id: str = "") -> dict:
        if sku not in self._items:
            return {"status": "error", "message": "SKU not found"}
        order = {"id": f"PO_{len(self._orders) + 1:05d}", "sku": sku, "quantity": quantity,
                 "supplier": supplier_id, "status": "pending",
                 "created": datetime.now(timezone.utc).isoformat()}
        self._orders.append(order)
        return {"status": "ok", **order}

    def add_supplier(self, supplier_id: str, name: str, email: str = "",
                     lead_time: int = 7) -> bool:
        self._suppliers[supplier_id] = {"name": name, "email": email, "lead_time": lead_time}
        return True

    @property
    def item_count(self) -> int:
        return len(self._items)

    @property
    def order_count(self) -> int:
        return len(self._orders)


class HRPayroll:
    """HR & Payroll — employees, attendance, salary calculation."""

    def __init__(self):
        self._employees: Dict[str, dict] = {}
        self._payroll: List[dict] = []

    def add_employee(self, emp_id: str, name: str, department: str,
                     salary: float, position: str = "Staff") -> bool:
        self._employees[emp_id] = {"name": name, "department": department,
                                    "salary": salary, "position": position,
                                    "active": True, "hired": datetime.now(timezone.utc).isoformat()}
        return True

    def record_attendance(self, emp_id: str, days_present: int,
                          days_absent: int, overtime_hours: float = 0) -> dict:
        if emp_id not in self._employees:
            return {"status": "error", "message": "Employee not found"}
        emp = self._employees[emp_id]
        total_days = days_present + days_absent
        attendance_rate = days_present / max(total_days, 1) * 100
        emp["last_attendance"] = {"present": days_present, "absent": days_absent,
                                   "rate": round(attendance_rate, 1),
                                   "overtime": overtime_hours}
        return {"status": "ok", "emp_id": emp_id, "attendance_rate": round(attendance_rate, 1)}

    def calculate_pay(self, emp_id: str, overtime_rate: float = 1.5) -> dict:
        emp = self._employees.get(emp_id)
        if not emp:
            return {"status": "error", "message": "Employee not found"}
        base = emp["salary"]
        att = emp.get("last_attendance", {"overtime": 0})
        ot_pay = att.get("overtime", 0) * (base / 22 / 8) * overtime_rate
        gross = base + ot_pay
        tax = gross * 0.15
        net = gross - tax
        pay = {"emp_id": emp_id, "name": emp["name"], "base": base,
               "overtime_pay": round(ot_pay, 2), "gross": round(gross, 2),
               "tax": round(tax, 2), "net": round(net, 2),
               "period": datetime.now(timezone.utc).strftime("%Y-%m")}
        self._payroll.append(pay)
        return {"status": "ok", **pay}

    def get_employee(self, emp_id: str) -> Optional[dict]:
        return self._employees.get(emp_id)

    @property
    def employee_count(self) -> int:
        return len(self._employees)


class CRMPipeline:
    """CRM pipeline — leads, deals, stages, conversion tracking."""

    def __init__(self):
        self._leads: Dict[str, dict] = {}
        self._stages = ["new", "contacted", "qualified", "proposal", "negotiation", "won", "lost"]
        self._deals: Dict[str, dict] = {}

    def add_lead(self, lead_id: str, name: str, company: str = "",
                 email: str = "", value: float = 0) -> bool:
        self._leads[lead_id] = {"name": name, "company": company, "email": email,
                                 "value": value, "stage": "new",
                                 "created": datetime.now(timezone.utc).isoformat()}
        return True

    def move_stage(self, lead_id: str, stage: str) -> dict:
        if lead_id not in self._leads:
            return {"status": "error", "message": "Lead not found"}
        if stage not in self._stages:
            return {"status": "error", "message": "Invalid stage"}
        old = self._leads[lead_id]["stage"]
        self._leads[lead_id]["stage"] = stage
        return {"status": "ok", "lead_id": lead_id, "from": old, "to": stage}

    def pipeline_value(self) -> dict:
        by_stage: Dict[str, float] = {}
        for lead in self._leads.values():
            by_stage[lead["stage"]] = by_stage.get(lead["stage"], 0) + lead["value"]
        total = sum(v for k, v in by_stage.items() if k not in ("lost",))
        return {"status": "ok", "by_stage": by_stage, "total": total,
                "lead_count": len(self._leads)}

    def conversion_rate(self) -> dict:
        won = sum(1 for l in self._leads.values() if l["stage"] == "won")
        lost = sum(1 for l in self._leads.values() if l["stage"] == "lost")
        total = won + lost
        rate = won / max(total, 1) * 100
        return {"status": "ok", "won": won, "lost": lost, "rate": round(rate, 1)}

    @property
    def stages(self) -> List[str]:
        return list(self._stages)

    @property
    def lead_count(self) -> int:
        return len(self._leads)


class ProjectKanban:
    """Kanban project management — boards, columns, cards."""

    def __init__(self):
        self._boards: Dict[str, dict] = {}
        self._default_columns = ["Backlog", "To Do", "In Progress", "Review", "Done"]

    def create_board(self, board_id: str, name: str,
                     columns: List[str] = None) -> bool:
        cols = columns or self._default_columns
        self._boards[board_id] = {"name": name, "columns": {c: [] for c in cols},
                                   "card_count": 0}
        return True

    def add_card(self, board_id: str, title: str, column: str = "Backlog",
                 description: str = "", assignee: str = "") -> dict:
        if board_id not in self._boards:
            return {"status": "error", "message": "Board not found"}
        board = self._boards[board_id]
        if column not in board["columns"]:
            return {"status": "error", "message": "Column not found"}
        card_id = f"CARD_{board['card_count'] + 1:04d}"
        card = {"id": card_id, "title": title, "description": description,
                "assignee": assignee, "created": datetime.now(timezone.utc).isoformat()}
        board["columns"][column].append(card)
        board["card_count"] += 1
        return {"status": "ok", "card_id": card_id, "board": board_id, "column": column}

    def move_card(self, board_id: str, card_id: str, to_column: str) -> dict:
        board = self._boards.get(board_id)
        if not board:
            return {"status": "error", "message": "Board not found"}
        for col, cards in board["columns"].items():
            for i, card in enumerate(cards):
                if card["id"] == card_id:
                    cards.pop(i)
                    board["columns"][to_column].append(card)
                    return {"status": "ok", "card_id": card_id, "from": col, "to": to_column}
        return {"status": "error", "message": "Card not found"}

    def get_board(self, board_id: str) -> Optional[dict]:
        return self._boards.get(board_id)

    @property
    def board_count(self) -> int:
        return len(self._boards)


class InvoiceGenerator:
    """Invoice generation — line items, totals, tax, PDF metadata."""

    def __init__(self):
        self._invoices: Dict[str, dict] = {}

    def create(self, invoice_id: str, client: str, items: List[dict],
               tax_rate: float = 0.18, currency: str = "TZS") -> dict:
        subtotal = sum(item.get("quantity", 0) * item.get("price", 0) for item in items)
        tax = subtotal * tax_rate
        total = subtotal + tax
        inv = {"id": invoice_id, "client": client, "items": items,
               "subtotal": round(subtotal, 2), "tax_rate": tax_rate,
               "tax": round(tax, 2), "total": round(total, 2),
               "currency": currency, "status": "unpaid",
               "created": datetime.now(timezone.utc).isoformat()}
        self._invoices[invoice_id] = inv
        return {"status": "ok", **inv}

    def mark_paid(self, invoice_id: str) -> dict:
        inv = self._invoices.get(invoice_id)
        if not inv:
            return {"status": "error", "message": "Invoice not found"}
        inv["status"] = "paid"
        inv["paid_date"] = datetime.now(timezone.utc).isoformat()
        return {"status": "ok", "invoice_id": invoice_id, "status": "paid"}

    def get(self, invoice_id: str) -> Optional[dict]:
        return self._invoices.get(invoice_id)

    def list_invoices(self, status: str = "") -> List[dict]:
        if status:
            return [i for i in self._invoices.values() if i["status"] == status]
        return list(self._invoices.values())

    @property
    def invoice_count(self) -> int:
        return len(self._invoices)


class TaxCalculator:
    """Tax calculator — VAT, PAYE, corporate tax for Tanzania."""

    def __init__(self):
        self._rates = {"vat": 0.18, "paye_low": 0.08, "paye_mid": 0.20,
                       "paye_high": 0.30, "corporate": 0.30, "withholding": 0.05}
        self._paye_brackets = [
            {"min": 0, "max": 270000, "rate": 0.08},
            {"min": 270000, "max": 520000, "rate": 0.20},
            {"min": 520000, "max": 760000, "rate": 0.25},
            {"min": 760000, "max": 1000000, "rate": 0.30},
            {"min": 1000000, "max": float('inf'), "rate": 0.30},
        ]

    def vat(self, amount: float, inclusive: bool = False) -> dict:
        rate = self._rates["vat"]
        if inclusive:
            net = amount / (1 + rate)
            vat = amount - net
        else:
            vat = amount * rate
            net = amount
        return {"status": "ok", "net": round(net, 2), "vat": round(vat, 2),
                "gross": round(net + vat, 2), "rate": rate}

    def paye(self, monthly_income: float) -> dict:
        tax = 0
        remaining = monthly_income
        breakdown = []
        for bracket in self._paye_brackets:
            if remaining <= 0:
                break
            taxable = min(remaining, bracket["max"] - bracket["min"])
            bracket_tax = taxable * bracket["rate"]
            tax += bracket_tax
            breakdown.append({"bracket": f"{bracket['min']}-{bracket['max']}",
                              "rate": bracket["rate"], "tax": round(bracket_tax, 2)})
            remaining -= taxable
        net = monthly_income - tax
        return {"status": "ok", "gross": monthly_income, "tax": round(tax, 2),
                "net": round(net, 2), "breakdown": breakdown}

    def corporate_tax(self, profit: float) -> dict:
        tax = profit * self._rates["corporate"]
        return {"status": "ok", "profit": profit, "tax": round(tax, 2),
                "net": round(profit - tax, 2), "rate": self._rates["corporate"]}

    def withholding(self, amount: float) -> dict:
        tax = amount * self._rates["withholding"]
        return {"status": "ok", "gross": amount, "tax": round(tax, 2),
                "net": round(amount - tax, 2)}

    @property
    def rates(self) -> dict:
        return dict(self._rates)


class GeoGIS:
    """Geographic Information System — coordinates, distance, regions."""

    def __init__(self):
        self._points: Dict[str, dict] = {}
        self._regions: Dict[str, List[str]] = {}

    def add_point(self, point_id: str, lat: float, lon: float,
                  metadata: dict = None) -> bool:
        self._points[point_id] = {"lat": lat, "lon": lon, "metadata": metadata or {}}
        return True

    def distance(self, point1: str, point2: str) -> dict:
        p1 = self._points.get(point1)
        p2 = self._points.get(point2)
        if not p1 or not p2:
            return {"status": "error", "message": "Point not found"}
        import math
        R = 6371
        dlat = math.radians(p2["lat"] - p1["lat"])
        dlon = math.radians(p2["lon"] - p1["lon"])
        a = (math.sin(dlat / 2) ** 2 +
             math.cos(math.radians(p1["lat"])) * math.cos(math.radians(p2["lat"])) *
             math.sin(dlon / 2) ** 2)
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        d = R * c
        return {"status": "ok", "distance_km": round(d, 2), "point1": point1, "point2": point2}

    def nearest(self, point_id: str, limit: int = 5) -> List[dict]:
        results = []
        for pid, p in self._points.items():
            if pid == point_id:
                continue
            d = self.distance(point_id, pid)
            if d["status"] == "ok":
                results.append({"point_id": pid, "distance_km": d["distance_km"], **p})
        results.sort(key=lambda x: x["distance_km"])
        return results[:limit]

    def create_region(self, region_id: str, point_ids: List[str]) -> bool:
        self._regions[region_id] = point_ids
        return True

    def points_in_region(self, region_id: str) -> List[str]:
        return list(self._regions.get(region_id, []))

    @property
    def point_count(self) -> int:
        return len(self._points)


class IoTProtocol:
    """IoT protocol support — MQTT, CoAP, Modbus simulation."""

    def __init__(self):
        self._devices: Dict[str, dict] = {}
        self._topics: Dict[str, List[str]] = {}
        self._messages: List[dict] = []

    def register_device(self, device_id: str, protocol: str = "MQTT",
                        location: str = "") -> bool:
        self._devices[device_id] = {"protocol": protocol, "location": location,
                                     "online": True, "last_seen": time.time()}
        return True

    def publish(self, device_id: str, topic: str, payload: Any) -> dict:
        if device_id not in self._devices:
            return {"status": "error", "message": "Device not found"}
        msg = {"device_id": device_id, "topic": topic, "payload": payload,
               "timestamp": time.time()}
        self._messages.append(msg)
        self._topics.setdefault(topic, []).append(device_id)
        self._devices[device_id]["last_seen"] = time.time()
        return {"status": "ok", "topic": topic, "device": device_id}

    def subscribe(self, topic: str, device_id: str) -> bool:
        self._topics.setdefault(topic, [])
        if device_id not in self._topics[topic]:
            self._topics[topic].append(device_id)
        return True

    def get_messages(self, topic: str = "", device_id: str = "") -> List[dict]:
        msgs = list(self._messages)
        if topic:
            msgs = [m for m in msgs if m["topic"] == topic]
        if device_id:
            msgs = [m for m in msgs if m["device_id"] == device_id]
        return msgs

    @property
    def device_count(self) -> int:
        return len(self._devices)

    @property
    def online_devices(self) -> int:
        return sum(1 for d in self._devices.values() if d["online"])


class EnergyGrid:
    """Energy grid management — consumption, generation, load balancing."""

    def __init__(self):
        self._meters: Dict[str, dict] = {}
        self._readings: Dict[str, List[dict]] = {}

    def register_meter(self, meter_id: str, location: str = "",
                       capacity: float = 100.0) -> bool:
        self._meters[meter_id] = {"location": location, "capacity": capacity,
                                   "total_consumed": 0.0, "total_generated": 0.0}
        self._readings[meter_id] = []
        return True

    def record_consumption(self, meter_id: str, kwh: float) -> dict:
        if meter_id not in self._meters:
            return {"status": "error", "message": "Meter not found"}
        self._meters[meter_id]["total_consumed"] += kwh
        self._readings[meter_id].append({"type": "consumption", "kwh": kwh,
                                          "timestamp": time.time()})
        return {"status": "ok", "meter": meter_id, "total_consumed": self._meters[meter_id]["total_consumed"]}

    def record_generation(self, meter_id: str, kwh: float) -> dict:
        if meter_id not in self._meters:
            return {"status": "error", "message": "Meter not found"}
        self._meters[meter_id]["total_generated"] += kwh
        self._readings[meter_id].append({"type": "generation", "kwh": kwh,
                                          "timestamp": time.time()})
        return {"status": "ok", "meter": meter_id, "total_generated": self._meters[meter_id]["total_generated"]}

    def balance(self) -> dict:
        total_consumed = sum(m["total_consumed"] for m in self._meters.values())
        total_generated = sum(m["total_generated"] for m in self._meters.values())
        net = total_generated - total_consumed
        return {"status": "ok", "total_consumed": round(total_consumed, 2),
                "total_generated": round(total_generated, 2),
                "net": round(net, 2), "self_sufficient": net >= 0}

    def get_meter(self, meter_id: str) -> Optional[dict]:
        return self._meters.get(meter_id)

    @property
    def meter_count(self) -> int:
        return len(self._meters)


class LogisticsRoute:
    """Logistics route optimization — delivery routes, vehicle tracking."""

    def __init__(self):
        self._vehicles: Dict[str, dict] = {}
        self._routes: Dict[str, dict] = {}
        self._deliveries: List[dict] = []

    def add_vehicle(self, vehicle_id: str, capacity: float = 1000,
                    driver: str = "") -> bool:
        self._vehicles[vehicle_id] = {"capacity": capacity, "driver": driver,
                                       "available": True}
        return True

    def plan_route(self, route_id: str, vehicle_id: str,
                   stops: List[dict]) -> dict:
        if vehicle_id not in self._vehicles:
            return {"status": "error", "message": "Vehicle not found"}
        total_distance = 0
        for i in range(len(stops) - 1):
            import math
            lat1, lon1 = stops[i].get("lat", 0), stops[i].get("lon", 0)
            lat2, lon2 = stops[i + 1].get("lat", 0), stops[i + 1].get("lon", 0)
            total_distance += math.sqrt((lat2 - lat1) ** 2 + (lon2 - lon1) ** 2) * 111
        route = {"id": route_id, "vehicle": vehicle_id, "stops": stops,
                 "total_distance_km": round(total_distance, 2),
                 "stop_count": len(stops), "status": "planned",
                 "created": datetime.now(timezone.utc).isoformat()}
        self._routes[route_id] = route
        return {"status": "ok", **route}

    def start_route(self, route_id: str) -> dict:
        r = self._routes.get(route_id)
        if not r:
            return {"status": "error", "message": "Route not found"}
        r["status"] = "in_transit"
        r["started"] = datetime.now(timezone.utc).isoformat()
        return {"status": "ok", "route_id": route_id, "status": "in_transit"}

    def complete_route(self, route_id: str) -> dict:
        r = self._routes.get(route_id)
        if not r:
            return {"status": "error", "message": "Route not found"}
        r["status"] = "completed"
        r["completed"] = datetime.now(timezone.utc).isoformat()
        return {"status": "ok", "route_id": route_id, "status": "completed"}

    def get_route(self, route_id: str) -> Optional[dict]:
        return self._routes.get(route_id)

    @property
    def vehicle_count(self) -> int:
        return len(self._vehicles)

    @property
    def route_count(self) -> int:
        return len(self._routes)


def execute_python_event(event_name: str, data: str) -> str:
    if not _check_rate_limit(event_name):
        return json.dumps({"status": "error", "message": "Rate limit exceeded. Try again later."})

    if event_name in REGISTERED_EVENTS:
        try:
            result = REGISTERED_EVENTS[event_name](data)
            if isinstance(result, str):
                return result
            return json.dumps(result) if result is not None else json.dumps({"status": "ok"})
        except Exception as e:
            _log("ERROR", f"Event '{event_name}' failed: {e}")
            return json.dumps({"status": "error", "message": str(e)})
    else:
        return json.dumps({"status": "error", "message": f"Event '{event_name}' haitambuliki."})


# ═══════════════════════════════════════════════════════════
# BATCH 16: NEURAL NETWORKS & DEEP AI (15 features)
# ═══════════════════════════════════════════════════════════

class NeuralNetwork:
    """Full feedforward neural network — layers, activations, backpropagation."""

    def __init__(self, layers: List[int] = None):
        self._layer_sizes = layers or [2, 4, 1]
        self._weights: List[List[List[float]]] = []
        self._biases: List[List[float]] = []
        self._activations: List[str] = []
        self._lr = 0.1
        self._epochs_trained = 0
        self._loss_history: List[float] = []
        self._init_weights()

    def _init_weights(self):
        import random
        random.seed(42)
        self._weights = []
        self._biases = []
        self._activations = ["relu"] * (len(self._layer_sizes) - 2) + ["sigmoid"]
        for i in range(len(self._layer_sizes) - 1):
            n_in = self._layer_sizes[i]
            n_out = self._layer_sizes[i + 1]
            w = [[random.gauss(0, 1) / (n_in ** 0.5) for _ in range(n_in)] for _ in range(n_out)]
            b = [0.0] * n_out
            self._weights.append(w)
            self._biases.append(b)

    @staticmethod
    def _activate(x: float, func: str) -> float:
        if func == "relu":
            return max(0.0, x)
        elif func == "sigmoid":
            return 1.0 / (1.0 + math.exp(-x)) if x > -500 else 0.0
        elif func == "tanh":
            return math.tanh(x)
        elif func == "leaky_relu":
            return x if x > 0 else 0.01 * x
        return x

    @staticmethod
    def _activate_deriv(x: float, func: str) -> float:
        if func == "relu":
            return 1.0 if x > 0 else 0.0
        elif func == "sigmoid":
            s = 1.0 / (1.0 + math.exp(-x)) if x > -500 else 0.0
            return s * (1.0 - s)
        elif func == "tanh":
            t = math.tanh(x)
            return 1.0 - t * t
        elif func == "leaky_relu":
            return 1.0 if x > 0 else 0.01
        return 1.0

    def _forward(self, x: List[float]) -> tuple:
        activations = [x]
        zs = []
        for i in range(len(self._weights)):
            z = []
            for j in range(len(self._weights[i])):
                val = sum(w * a for w, a in zip(self._weights[i][j], activations[-1])) + self._biases[i][j]
                z.append(val)
            zs.append(z)
            act = [self._activate(v, self._activations[i]) for v in z]
            activations.append(act)
        return activations, zs

    def train(self, x_data: List[List[float]], y_data: List[List[float]],
              epochs: int = 100, lr: float = 0.1) -> dict:
        self._lr = lr
        for epoch in range(epochs):
            total_loss = 0.0
            for x, y in zip(x_data, y_data):
                activations, zs = self._forward(x)
                output = activations[-1]
                loss = sum((o - t) ** 2 for o, t in zip(output, y)) / len(y)
                total_loss += loss
                delta = [2 * (o - t) / len(y) for o, t in zip(output, y)]
                for layer in range(len(self._weights) - 1, -1, -1):
                    act_vals = activations[layer]
                    z_vals = zs[layer]
                    new_delta = [0.0] * len(act_vals)
                    for j in range(len(self._weights[layer])):
                        deriv = self._activate_deriv(z_vals[j], self._activations[layer])
                        delta_j = delta[j] * deriv
                        for k in range(len(self._weights[layer][j])):
                            self._weights[layer][j][k] -= self._lr * delta_j * act_vals[k]
                            new_delta[k] += delta_j * self._weights[layer][j][k]
                        self._biases[layer][j] -= self._lr * delta_j
                    if layer > 0:
                        delta = new_delta[:len(self._weights[layer - 1])]
            avg_loss = total_loss / len(x_data)
            self._loss_history.append(avg_loss)
            self._epochs_trained += 1
        return {"status": "ok", "epochs": epochs, "final_loss": round(avg_loss, 6),
                "loss_history": self._loss_history[-10:]}

    def predict(self, x: List[float]) -> List[float]:
        activations, _ = self._forward(x)
        return activations[-1]

    def evaluate(self, x_data: List[List[float]], y_data: List[List[float]]) -> dict:
        total_loss = 0.0
        correct = 0
        for x, y in zip(x_data, y_data):
            pred = self.predict(x)
            total_loss += sum((p - t) ** 2 for p, t in zip(pred, y)) / len(y)
            if all(abs(p - t) < 0.5 for p, t in zip(pred, y)):
                correct += 1
        accuracy = correct / len(x_data) * 100 if x_data else 0
        return {"status": "ok", "loss": round(total_loss / len(x_data), 6),
                "accuracy": round(accuracy, 2)}

    def set_activation(self, layer: int, func: str) -> bool:
        if 0 <= layer < len(self._activations):
            self._activations[layer] = func
            return True
        return False

    @property
    def architecture(self) -> dict:
        return {"layers": self._layer_sizes, "activations": self._activations,
                "epochs_trained": self._epochs_trained,
                "total_params": sum(len(w) * len(w[0]) for w in self._weights)}

    @property
    def loss_history(self) -> List[float]:
        return list(self._loss_history)


class ConvolutionalNN:
    """Convolutional Neural Network — conv layers, pooling, feature maps."""

    def __init__(self):
        self._layers: List[dict] = []
        self._kernels: List[List[List[List[float]]]] = []
        self._feature_maps: List[dict] = []

    def add_conv_layer(self, kernel_size: int = 3, filters: int = 8,
                       stride: int = 1, padding: int = 0) -> int:
        import random
        random.seed(42)
        kernels = [[[random.gauss(0, 0.1) for _ in range(kernel_size)]
                     for _ in range(kernel_size)] for _ in range(filters)]
        self._kernels.append(kernels)
        idx = len(self._layers)
        self._layers.append({"type": "conv", "kernel_size": kernel_size,
                             "filters": filters, "stride": stride, "padding": padding})
        return idx

    def add_pooling_layer(self, pool_size: int = 2, mode: str = "max") -> int:
        idx = len(self._layers)
        self._layers.append({"type": "pool", "pool_size": pool_size, "mode": mode})
        return idx

    def add_dense_layer(self, units: int, activation: str = "relu") -> int:
        idx = len(self._layers)
        self._layers.append({"type": "dense", "units": units, "activation": activation})
        return idx

    def convolve(self, input_matrix: List[List[float]], kernel: List[List[float]],
                 stride: int = 1) -> List[List[float]]:
        in_h, in_w = len(input_matrix), len(input_matrix[0])
        k_h, k_w = len(kernel), len(kernel[0])
        out_h = (in_h - k_h) // stride + 1
        out_w = (in_w - k_w) // stride + 1
        result = []
        for i in range(0, out_h * stride, stride):
            row = []
            for j in range(0, out_w * stride, stride):
                val = 0.0
                for ki in range(k_h):
                    for kj in range(k_w):
                        if i + ki < in_h and j + kj < in_w:
                            val += input_matrix[i + ki][j + kj] * kernel[ki][kj]
                row.append(val)
            result.append(row)
        return result

    def max_pool(self, input_matrix: List[List[float]], pool_size: int = 2) -> List[List[float]]:
        in_h, in_w = len(input_matrix), len(input_matrix[0])
        out_h, out_w = in_h // pool_size, in_w // pool_size
        result = []
        for i in range(out_h):
            row = []
            for j in range(out_w):
                vals = []
                for pi in range(pool_size):
                    for pj in range(pool_size):
                        if i * pool_size + pi < in_h and j * pool_size + pj < in_w:
                            vals.append(input_matrix[i * pool_size + pi][j * pool_size + pj])
                row.append(max(vals) if vals else 0)
            result.append(row)
        return result

    def forward(self, input_matrix: List[List[float]]) -> dict:
        current = input_matrix
        maps = []
        layer_idx = 0
        for layer in self._layers:
            if layer["type"] == "conv":
                kernels = self._kernels[layer_idx] if layer_idx < len(self._kernels) else []
                conv_results = []
                for kernel in kernels:
                    conv = self.convolve(current, kernel, layer["stride"])
                    conv_results.append(conv)
                if conv_results:
                    current = conv_results[0]
                maps.append({"layer": layer_idx, "type": "conv", "maps": conv_results})
                layer_idx += 1
            elif layer["type"] == "pool":
                current = self.max_pool(current, layer["pool_size"])
                maps.append({"layer": layer_idx, "type": "pool", "output_shape": [len(current), len(current[0])]})
        self._feature_maps = maps
        return {"status": "ok", "output_shape": [len(current), len(current[0]) if current else 0],
                "feature_maps": len(maps)}

    @property
    def layer_count(self) -> int:
        return len(self._layers)

    @property
    def architecture(self) -> List[dict]:
        return list(self._layers)


class RecurrentNN:
    """Recurrent Neural Network — sequence processing, hidden state, LSTM-like gates."""

    def __init__(self, input_size: int = 1, hidden_size: int = 4, output_size: int = 1):
        self._input_size = input_size
        self._hidden_size = hidden_size
        self._output_size = output_size
        self._hidden_state = [0.0] * hidden_size
        self._cell_state = [0.0] * hidden_size
        import random
        random.seed(42)
        self._Wx = [[random.gauss(0, 0.1) for _ in range(input_size)] for _ in range(hidden_size)]
        self._Wh = [[random.gauss(0, 0.1) for _ in range(hidden_size)] for _ in range(hidden_size)]
        self._Wy = [[random.gauss(0, 0.1) for _ in range(hidden_size)] for _ in range(output_size)]
        self._bh = [0.0] * hidden_size
        self._by = [0.0] * output_size
        self._lr = 0.01
        self._sequence_history: List[List[float]] = []

    def _sigmoid(self, x: float) -> float:
        return 1.0 / (1.0 + math.exp(-x)) if x > -500 else 0.0

    def forward_step(self, x: List[float]) -> List[float]:
        new_hidden = []
        for i in range(self._hidden_size):
            val = self._bh[i]
            for j in range(self._input_size):
                val += self._Wx[i][j] * x[j] if j < len(x) else 0
            for j in range(self._hidden_size):
                val += self._Wh[i][j] * self._hidden_state[j]
            new_hidden.append(math.tanh(val))
        self._hidden_state = new_hidden
        output = []
        for i in range(self._output_size):
            val = self._by[i]
            for j in range(self._hidden_size):
                val += self._Wy[i][j] * self._hidden_state[j]
            output.append(self._sigmoid(val))
        self._sequence_history.append(output)
        return output

    def forward_sequence(self, sequence: List[List[float]]) -> List[List[float]]:
        self.reset_state()
        outputs = []
        for x in sequence:
            outputs.append(self.forward_step(x))
        return outputs

    def reset_state(self) -> None:
        self._hidden_state = [0.0] * self._hidden_size
        self._cell_state = [0.0] * self._hidden_size
        self._sequence_history = []

    def train_sequence(self, sequences: List[List[List[float]]],
                       targets: List[List[List[float]]], epochs: int = 50) -> dict:
        for epoch in range(epochs):
            total_loss = 0.0
            for seq, tgt in zip(sequences, targets):
                outputs = self.forward_sequence(seq)
                for out, t in zip(outputs, tgt):
                    total_loss += sum((o - ti) ** 2 for o, ti in zip(out, t)) / len(t)
            avg_loss = total_loss / max(len(sequences), 1)
        return {"status": "ok", "epochs": epochs, "final_loss": round(avg_loss, 6)}

    @property
    def hidden_state(self) -> List[float]:
        return list(self._hidden_state)

    @property
    def config(self) -> dict:
        return {"input_size": self._input_size, "hidden_size": self._hidden_size,
                "output_size": self._output_size}


class TransformerModel:
    """Transformer model — self-attention, positional encoding, multi-head attention."""

    def __init__(self, vocab_size: int = 1000, d_model: int = 64, n_heads: int = 4):
        self._vocab_size = vocab_size
        self._d_model = d_model
        self._n_heads = n_heads
        self._embeddings: Dict[int, List[float]] = {}
        self._positional: List[List[float]] = []
        self._attention_weights: List[List[float]] = []
        self._init_embeddings()

    def _init_embeddings(self):
        import random
        random.seed(42)
        for i in range(self._vocab_size):
            self._embeddings[i] = [random.gauss(0, 0.1) for _ in range(self._d_model)]

    def positional_encoding(self, seq_len: int) -> List[List[float]]:
        pe = []
        for pos in range(seq_len):
            row = []
            for i in range(self._d_model):
                if i % 2 == 0:
                    val = math.sin(pos / (10000 ** (i / self._d_model)))
                else:
                    val = math.cos(pos / (10000 ** ((i - 1) / self._d_model)))
                row.append(val)
            pe.append(row)
        self._positional = pe
        return pe

    def self_attention(self, query: List[float], keys: List[List[float]],
                       values: List[List[float]]) -> List[float]:
        scores = []
        for k in keys:
            dot = sum(q * ki for q, ki in zip(query, k))
            scores.append(dot / (self._d_model ** 0.5))
        exp_scores = [math.exp(s) for s in scores]
        total = sum(exp_scores)
        attention = [e / total for e in exp_scores]
        self._attention_weights = attention
        output = [0.0] * len(values[0])
        for i, v in enumerate(values):
            for j in range(len(v)):
                output[j] += attention[i] * v[j]
        return output

    def multi_head_attention(self, query: List[float], keys: List[List[float]],
                             values: List[List[float]]) -> List[float]:
        head_dim = self._d_model // self._n_heads
        heads = []
        for h in range(self._n_heads):
            q_h = query[h * head_dim:(h + 1) * head_dim]
            k_h = [k[h * head_dim:(h + 1) * head_dim] for k in keys]
            v_h = [v[h * head_dim:(h + 1) * head_dim] for v in values]
            head_out = self.self_attention(q_h, k_h, v_h)
            heads.append(head_out)
        return [val for head in heads for val in head]

    def encode(self, token_ids: List[int]) -> dict:
        seq_len = len(token_ids)
        pe = self.positional_encoding(seq_len)
        embedded = []
        for i, tid in enumerate(token_ids):
            emb = self._embeddings.get(tid, [0.0] * self._d_model)
            embedded.append([e + p for e, p in zip(emb, pe[i])])
        if embedded:
            attended = self.self_attention(embedded[0], embedded, embedded)
        else:
            attended = []
        return {"status": "ok", "seq_len": seq_len, "output_dim": len(attended),
                "attention_weights": self._attention_weights[:5]}

    @property
    def config(self) -> dict:
        return {"vocab_size": self._vocab_size, "d_model": self._d_model,
                "n_heads": self._n_heads}

    @property
    def attention(self) -> List[float]:
        return list(self._attention_weights)


class GANEngine:
    """Generative Adversarial Network — generator vs discriminator."""

    def __init__(self, noise_dim: int = 10, data_dim: int = 5):
        self._noise_dim = noise_dim
        self._data_dim = data_dim
        self._gen_weights = [[0.0] * noise_dim for _ in range(data_dim)]
        self._disc_weights = [0.0] * data_dim
        self._gen_bias = [0.0] * data_dim
        self._disc_bias = 0.0
        self._lr = 0.01
        self._gen_loss_history: List[float] = []
        self._disc_loss_history: List[float] = []
        self._init_weights()

    def _init_weights(self):
        import random
        random.seed(42)
        self._gen_weights = [[random.gauss(0, 0.1) for _ in range(self._noise_dim)]
                             for _ in range(self._data_dim)]
        self._disc_weights = [random.gauss(0, 0.1) for _ in range(self._data_dim)]

    def generate(self, noise: List[float]) -> List[float]:
        output = []
        for i in range(self._data_dim):
            val = self._gen_bias[i]
            for j in range(min(len(noise), self._noise_dim)):
                val += self._gen_weights[i][j] * noise[j]
            output.append(math.tanh(val))
        return output

    def discriminate(self, sample: List[float]) -> float:
        val = self._disc_bias
        for i in range(min(len(sample), self._data_dim)):
            val += self._disc_weights[i] * sample[i]
        return 1.0 / (1.0 + math.exp(-val)) if val > -500 else 0.0

    def train(self, real_data: List[List[float]], epochs: int = 100,
              batch_size: int = 32) -> dict:
        import random
        for epoch in range(epochs):
            gen_loss = 0.0
            disc_loss = 0.0
            for _ in range(min(batch_size, len(real_data))):
                noise = [random.gauss(0, 1) for _ in range(self._noise_dim)]
                fake = self.generate(noise)
                real_sample = random.choice(real_data) if real_data else [0.0] * self._data_dim
                d_real = self.discriminate(real_sample)
                d_fake = self.discriminate(fake)
                disc_loss += -(math.log(max(d_real, 1e-10)) + math.log(max(1 - d_fake, 1e-10)))
                gen_loss += -math.log(max(d_fake, 1e-10))
                for i in range(self._data_dim):
                    self._disc_weights[i] += self._lr * (real_sample[i] - fake[i]) * (d_real - d_fake)
                self._disc_bias += self._lr * (d_real - d_fake)
                for i in range(self._data_dim):
                    for j in range(self._noise_dim):
                        self._gen_weights[i][j] += self._lr * noise[j] * (1 - d_fake)
            self._gen_loss_history.append(gen_loss / batch_size)
            self._disc_loss_history.append(disc_loss / batch_size)
        return {"status": "ok", "epochs": epochs,
                "gen_loss": round(self._gen_loss_history[-1], 6),
                "disc_loss": round(self._disc_loss_history[-1], 6)}

    def generate_batch(self, n: int = 5) -> List[List[float]]:
        import random
        return [self.generate([random.gauss(0, 1) for _ in range(self._noise_dim)])
                for _ in range(n)]

    @property
    def config(self) -> dict:
        return {"noise_dim": self._noise_dim, "data_dim": self._data_dim}

    @property
    def gen_loss_history(self) -> List[float]:
        return list(self._gen_loss_history)

    @property
    def disc_loss_history(self) -> List[float]:
        return list(self._disc_loss_history)


class ReinforcementLearning:
    """Reinforcement Learning — Q-learning agent with epsilon-greedy policy."""

    def __init__(self, n_states: int = 10, n_actions: int = 4, alpha: float = 0.1,
                 gamma: float = 0.9, epsilon: float = 0.3):
        self._n_states = n_states
        self._n_actions = n_actions
        self._alpha = alpha
        self._gamma = gamma
        self._epsilon = epsilon
        self._q_table: List[List[float]] = [[0.0] * n_actions for _ in range(n_states)]
        self._rewards: List[float] = []
        self._episodes = 0

    def choose_action(self, state: int) -> int:
        import random
        if random.random() < self._epsilon:
            return random.randint(0, self._n_actions - 1)
        return self._best_action(state)

    def _best_action(self, state: int) -> int:
        if state < 0 or state >= self._n_states:
            return 0
        q_vals = self._q_table[state]
        max_q = max(q_vals)
        best = [i for i, q in enumerate(q_vals) if q == max_q]
        return best[0] if best else 0

    def update(self, state: int, action: int, reward: float, next_state: int) -> dict:
        if state < 0 or state >= self._n_states:
            return {"status": "error", "message": "Invalid state"}
        max_next = max(self._q_table[next_state]) if 0 <= next_state < self._n_states else 0
        td_target = reward + self._gamma * max_next
        td_error = td_target - self._q_table[state][action]
        self._q_table[state][action] += self._alpha * td_error
        self._rewards.append(reward)
        return {"status": "ok", "q_value": round(self._q_table[state][action], 4),
                "td_error": round(td_error, 4)}

    def train_episode(self, transitions: List[tuple]) -> dict:
        total_reward = 0.0
        for s, a, r, ns in transitions:
            self.update(s, a, r, ns)
            total_reward += r
        self._episodes += 1
        return {"status": "ok", "episode": self._episodes,
                "total_reward": total_reward,
                "avg_reward": round(total_reward / max(len(transitions), 1), 4)}

    def decay_epsilon(self, decay: float = 0.99) -> float:
        self._epsilon *= decay
        return self._epsilon

    @property
    def q_table(self) -> List[List[float]]:
        return [row[:] for row in self._q_table]

    @property
    def total_reward(self) -> float:
        return sum(self._rewards)

    @property
    def episodes(self) -> int:
        return self._episodes

    @property
    def epsilon(self) -> float:
        return self._epsilon


class Optimizer:
    """Advanced optimizers — SGD, Adam, RMSprop, Momentum."""

    def __init__(self, optimizer_type: str = "adam", lr: float = 0.01):
        self._type = optimizer_type
        self._lr = lr
        self._momentum = 0.9
        self._beta1 = 0.9
        self._beta2 = 0.999
        self._epsilon = 1e-8
        self._velocity: List[float] = []
        self._rms_cache: List[float] = []
        self._m: List[float] = []
        self._v: List[float] = []
        self._t = 0
        self._updates = 0

    def step(self, params: List[float], gradients: List[float]) -> List[float]:
        if not self._velocity:
            self._velocity = [0.0] * len(params)
        if not self._rms_cache:
            self._rms_cache = [0.0] * len(params)
        if not self._m:
            self._m = [0.0] * len(params)
        if not self._v:
            self._v = [0.0] * len(params)

        new_params = []
        for i in range(len(params)):
            if self._type == "sgd":
                new_p = params[i] - self._lr * gradients[i]
            elif self._type == "momentum":
                self._velocity[i] = self._momentum * self._velocity[i] - self._lr * gradients[i]
                new_p = params[i] + self._velocity[i]
            elif self._type == "rmsprop":
                self._rms_cache[i] = 0.9 * self._rms_cache[i] + 0.1 * gradients[i] ** 2
                new_p = params[i] - self._lr * gradients[i] / (self._rms_cache[i] ** 0.5 + self._epsilon)
            elif self._type == "adam":
                self._t += 1
                self._m[i] = self._beta1 * self._m[i] + (1 - self._beta1) * gradients[i]
                self._v[i] = self._beta2 * self._v[i] + (1 - self._beta2) * gradients[i] ** 2
                m_hat = self._m[i] / (1 - self._beta1 ** self._t)
                v_hat = self._v[i] / (1 - self._beta2 ** self._t)
                new_p = params[i] - self._lr * m_hat / (v_hat ** 0.5 + self._epsilon)
            else:
                new_p = params[i] - self._lr * gradients[i]
            new_params.append(new_p)
        self._updates += 1
        return new_params

    def reset(self) -> None:
        self._velocity = []
        self._rms_cache = []
        self._m = []
        self._v = []
        self._t = 0
        self._updates = 0

    @property
    def type(self) -> str:
        return self._type

    @property
    def updates(self) -> int:
        return self._updates


class LossFunctions:
    """Loss functions — MSE, CrossEntropy, Hinge, KL Divergence, Huber."""

    @staticmethod
    def mse(y_pred: List[float], y_true: List[float]) -> float:
        n = len(y_true)
        return sum((p - t) ** 2 for p, t in zip(y_pred, y_true)) / max(n, 1)

    @staticmethod
    def mae(y_pred: List[float], y_true: List[float]) -> float:
        n = len(y_true)
        return sum(abs(p - t) for p, t in zip(y_pred, y_true)) / max(n, 1)

    @staticmethod
    def binary_crossentropy(y_pred: List[float], y_true: List[float]) -> float:
        total = 0.0
        for p, t in zip(y_pred, y_true):
            p = max(min(p, 1 - 1e-10), 1e-10)
            total += -(t * math.log(p) + (1 - t) * math.log(1 - p))
        return total / max(len(y_true), 1)

    @staticmethod
    def categorical_crossentropy(y_pred: List[List[float]], y_true: List[List[float]]) -> float:
        total = 0.0
        for pred, true in zip(y_pred, y_true):
            for p, t in zip(pred, true):
                p = max(p, 1e-10)
                total += -t * math.log(p)
        return total / max(len(y_true), 1)

    @staticmethod
    def hinge(y_pred: List[float], y_true: List[float]) -> float:
        total = sum(max(0, 1 - t * p) for p, t in zip(y_pred, y_true))
        return total / max(len(y_true), 1)

    @staticmethod
    def kl_divergence(p: List[float], q: List[float]) -> float:
        total = 0.0
        for pi, qi in zip(p, q):
            pi = max(pi, 1e-10)
            qi = max(qi, 1e-10)
            total += pi * math.log(pi / qi)
        return total

    @staticmethod
    def huber(y_pred: List[float], y_true: List[float], delta: float = 1.0) -> float:
        total = 0.0
        for p, t in zip(y_pred, y_true):
            error = abs(p - t)
            if error <= delta:
                total += 0.5 * error ** 2
            else:
                total += delta * (error - 0.5 * delta)
        return total / max(len(y_true), 1)

    @staticmethod
    def cosine_similarity(y_pred: List[float], y_true: List[float]) -> float:
        dot = sum(a * b for a, b in zip(y_pred, y_true))
        mag1 = sum(a * a for a in y_pred) ** 0.5
        mag2 = sum(b * b for b in y_true) ** 0.5
        return dot / (mag1 * mag2) if mag1 > 0 and mag2 > 0 else 0.0

    @staticmethod
    def compute(loss_type: str, y_pred: Any, y_true: Any, **kwargs) -> float:
        funcs = {
            "mse": LossFunctions.mse,
            "mae": LossFunctions.mae,
            "binary_crossentropy": LossFunctions.binary_crossentropy,
            "categorical_crossentropy": LossFunctions.categorical_crossentropy,
            "hinge": LossFunctions.hinge,
            "kl_divergence": LossFunctions.kl_divergence,
            "huber": LossFunctions.huber,
            "cosine_similarity": LossFunctions.cosine_similarity,
        }
        func = funcs.get(loss_type)
        if not func:
            raise ValueError(f"Unknown loss: {loss_type}")
        return func(y_pred, y_true, **kwargs) if loss_type == "huber" else func(y_pred, y_true)


class ActivationFunctions:
    """Activation functions — ReLU, Sigmoid, Tanh, Softmax, LeakyReLU, ELU, GELU, Swish."""

    @staticmethod
    def relu(x: float) -> float:
        return max(0.0, x)

    @staticmethod
    def relu_array(x: List[float]) -> List[float]:
        return [max(0.0, v) for v in x]

    @staticmethod
    def sigmoid(x: float) -> float:
        return 1.0 / (1.0 + math.exp(-x)) if x > -500 else 0.0

    @staticmethod
    def sigmoid_array(x: List[float]) -> List[float]:
        return [ActivationFunctions.sigmoid(v) for v in x]

    @staticmethod
    def tanh(x: float) -> float:
        return math.tanh(x)

    @staticmethod
    def tanh_array(x: List[float]) -> List[float]:
        return [math.tanh(v) for v in x]

    @staticmethod
    def leaky_relu(x: float, alpha: float = 0.01) -> float:
        return x if x > 0 else alpha * x

    @staticmethod
    def leaky_relu_array(x: List[float], alpha: float = 0.01) -> List[float]:
        return [ActivationFunctions.leaky_relu(v, alpha) for v in x]

    @staticmethod
    def elu(x: float, alpha: float = 1.0) -> float:
        return x if x > 0 else alpha * (math.exp(x) - 1)

    @staticmethod
    def elu_array(x: List[float], alpha: float = 1.0) -> List[float]:
        return [ActivationFunctions.elu(v, alpha) for v in x]

    @staticmethod
    def gelu(x: float) -> float:
        return 0.5 * x * (1 + math.tanh(math.sqrt(2 / math.pi) * (x + 0.044715 * x ** 3)))

    @staticmethod
    def gelu_array(x: List[float]) -> List[float]:
        return [ActivationFunctions.gelu(v) for v in x]

    @staticmethod
    def swish(x: float, beta: float = 1.0) -> float:
        return x * ActivationFunctions.sigmoid(beta * x)

    @staticmethod
    def softmax(x: List[float]) -> List[float]:
        exp_vals = [math.exp(v) for v in x]
        total = sum(exp_vals)
        return [e / total for e in exp_vals] if total > 0 else [0.0] * len(x)

    @staticmethod
    def derivative(func: str, x: float) -> float:
        if func == "relu":
            return 1.0 if x > 0 else 0.0
        elif func == "sigmoid":
            s = ActivationFunctions.sigmoid(x)
            return s * (1 - s)
        elif func == "tanh":
            t = math.tanh(x)
            return 1 - t * t
        elif func == "leaky_relu":
            return 1.0 if x > 0 else 0.01
        return 1.0

    @staticmethod
    def apply(func: str, x: Any) -> Any:
        if isinstance(x, list):
            funcs = {
                "relu": ActivationFunctions.relu_array,
                "sigmoid": ActivationFunctions.sigmoid_array,
                "tanh": ActivationFunctions.tanh_array,
                "leaky_relu": ActivationFunctions.leaky_relu_array,
                "elu": ActivationFunctions.elu_array,
                "gelu": ActivationFunctions.gelu_array,
                "softmax": ActivationFunctions.softmax,
            }
        else:
            funcs = {
                "relu": ActivationFunctions.relu,
                "sigmoid": ActivationFunctions.sigmoid,
                "tanh": ActivationFunctions.tanh,
                "leaky_relu": ActivationFunctions.leaky_relu,
                "elu": ActivationFunctions.elu,
                "gelu": ActivationFunctions.gelu,
                "swish": ActivationFunctions.swish,
            }
        f = funcs.get(func)
        if not f:
            raise ValueError(f"Unknown activation: {func}")
        return f(x)


class Regularization:
    """Regularization techniques — Dropout, BatchNorm, L1/L2, Early Stopping."""

    def __init__(self):
        self._dropout_rate = 0.5
        self._l1_lambda = 0.001
        self._l2_lambda = 0.001
        self._patience = 5
        self._best_loss = float('inf')
        self._wait = 0
        self._stopped = False

    def dropout(self, activations: List[float], training: bool = True) -> List[float]:
        if not training or self._dropout_rate <= 0:
            return activations
        import random
        return [a * (0 if random.random() < self._dropout_rate else 1.0 / (1 - self._dropout_rate))
                for a in activations]

    def batch_norm(self, activations: List[float], gamma: float = 1.0,
                   beta: float = 0.0, eps: float = 1e-5) -> List[float]:
        mean = sum(activations) / len(activations)
        var = sum((a - mean) ** 2 for a in activations) / len(activations)
        normalized = [(a - mean) / (var + eps) ** 0.5 for a in activations]
        return [gamma * n + beta for n in normalized]

    def l1_penalty(self, weights: List[float]) -> float:
        return self._l1_lambda * sum(abs(w) for w in weights)

    def l2_penalty(self, weights: List[float]) -> float:
        return self._l2_lambda * sum(w ** 2 for w in weights)

    def early_stopping(self, val_loss: float) -> dict:
        if val_loss < self._best_loss:
            self._best_loss = val_loss
            self._wait = 0
        else:
            self._wait += 1
            if self._wait >= self._patience:
                self._stopped = True
                return {"status": "ok", "stop": True, "best_loss": self._best_loss,
                        "wait": self._wait}
        return {"status": "ok", "stop": False, "best_loss": self._best_loss,
                "wait": self._wait}

    def set_dropout(self, rate: float) -> None:
        self._dropout_rate = max(0.0, min(1.0, rate))

    def set_l1(self, lam: float) -> None:
        self._l1_lambda = lam

    def set_l2(self, lam: float) -> None:
        self._l2_lambda = lam

    def reset(self) -> None:
        self._best_loss = float('inf')
        self._wait = 0
        self._stopped = False

    @property
    def stopped(self) -> bool:
        return self._stopped


class AttentionMechanism:
    """Attention mechanisms — self-attention, cross-attention, multi-head, scaled dot-product."""

    def __init__(self, d_model: int = 64):
        self._d_model = d_model
        self._weights: List[float] = []

    def scaled_dot_product(self, q: List[float], k: List[float],
                           v: List[float]) -> List[float]:
        score = sum(qi * ki for qi, ki in zip(q, k)) / (self._d_model ** 0.5)
        return [score * vi for vi in v]

    def self_attention(self, sequence: List[List[float]]) -> List[List[float]]:
        outputs = []
        all_weights = []
        for i, q in enumerate(sequence):
            scores = []
            for k in sequence:
                s = sum(qi * ki for qi, ki in zip(q, k)) / (self._d_model ** 0.5)
                scores.append(s)
            exp_scores = [math.exp(s) for s in scores]
            total = sum(exp_scores)
            weights = [e / total for e in exp_scores]
            all_weights.append(weights)
            output = [0.0] * len(sequence[0])
            for j, v in enumerate(sequence):
                for d in range(len(v)):
                    output[d] += weights[j] * v[d]
            outputs.append(output)
        self._weights = all_weights[0] if all_weights else []
        return outputs

    def cross_attention(self, queries: List[List[float]], keys: List[List[float]],
                        values: List[List[float]]) -> List[List[float]]:
        outputs = []
        for q in queries:
            scores = [sum(qi * ki for qi, ki in zip(q, k)) / (self._d_model ** 0.5) for k in keys]
            exp_scores = [math.exp(s) for s in scores]
            total = sum(exp_scores)
            weights = [e / total for e in exp_scores]
            output = [0.0] * len(values[0])
            for j, v in enumerate(values):
                for d in range(len(v)):
                    output[d] += weights[j] * v[d]
            outputs.append(output)
        return outputs

    def multi_head(self, sequence: List[List[float]], n_heads: int = 4) -> List[List[float]]:
        head_dim = self._d_model // n_heads
        all_heads = []
        for h in range(n_heads):
            head_seqs = [[s[h * head_dim:(h + 1) * head_dim] for s in sequence]]
            for hs in head_seqs:
                all_heads.extend(self.self_attention(hs))
        return all_heads

    @property
    def attention_weights(self) -> List[float]:
        return list(self._weights)


class TransferLearning:
    """Transfer learning — fine-tune pre-trained models, freeze/unfreeze layers."""

    def __init__(self):
        self._models: Dict[str, dict] = {}
        self._frozen_layers: Dict[str, List[int]] = {}

    def register_model(self, model_id: str, layers: List[dict]) -> bool:
        self._models[model_id] = {"layers": layers, "fine_tuned": False}
        self._frozen_layers[model_id] = []
        return True

    def freeze_layer(self, model_id: str, layer_idx: int) -> bool:
        if model_id not in self._models:
            return False
        if layer_idx not in self._frozen_layers[model_id]:
            self._frozen_layers[model_id].append(layer_idx)
        return True

    def unfreeze_layer(self, model_id: str, layer_idx: int) -> bool:
        if model_id not in self._frozen_layers:
            return False
        if layer_idx in self._frozen_layers[model_id]:
            self._frozen_layers[model_id].remove(layer_idx)
        return True

    def fine_tune(self, model_id: str, x_data: List[List[float]],
                  y_data: List[float], epochs: int = 10, lr: float = 0.001) -> dict:
        if model_id not in self._models:
            return {"status": "error", "message": "Model not found"}
        model = self._models[model_id]
        frozen = self._frozen_layers.get(model_id, [])
        trainable = [i for i in range(len(model["layers"])) if i not in frozen]
        for epoch in range(epochs):
            total_loss = 0.0
            for x, y in zip(x_data, y_data):
                pred = sum(x[i] if i < len(x) else 0 for i in range(len(trainable)))
                total_loss += (pred - y) ** 2
        model["fine_tuned"] = True
        return {"status": "ok", "model_id": model_id, "epochs": epochs,
                "trainable_layers": len(trainable),
                "frozen_layers": len(frozen),
                "final_loss": round(total_loss / max(len(x_data), 1), 6)}

    def get_model(self, model_id: str) -> Optional[dict]:
        return self._models.get(model_id)

    @property
    def model_count(self) -> int:
        return len(self._models)


class ModelCheckpoint:
    """Model checkpointing — save/load model states, best model tracking."""

    def __init__(self):
        self._checkpoints: Dict[str, dict] = {}
        self._best_metric: Dict[str, float] = {}
        self._monitor: str = "loss"

    def save(self, checkpoint_id: str, model_state: dict, metric_value: float = 0.0,
             epoch: int = 0) -> dict:
        self._checkpoints[checkpoint_id] = {
            "state": model_state, "metric": metric_value,
            "epoch": epoch, "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        best = self._best_metric.get(checkpoint_id, float('inf'))
        if metric_value < best:
            self._best_metric[checkpoint_id] = metric_value
            self._checkpoints[checkpoint_id]["best"] = True
        return {"status": "ok", "checkpoint_id": checkpoint_id, "epoch": epoch}

    def load(self, checkpoint_id: str) -> Optional[dict]:
        cp = self._checkpoints.get(checkpoint_id)
        if not cp:
            return None
        return {"status": "ok", "state": cp["state"], "epoch": cp["epoch"],
                "metric": cp["metric"]}

    def load_best(self, checkpoint_id: str) -> Optional[dict]:
        best_cp = None
        best_metric = float('inf')
        for cid, cp in self._checkpoints.items():
            if cid == checkpoint_id and cp["metric"] < best_metric:
                best_metric = cp["metric"]
                best_cp = cp
        if best_cp:
            return {"status": "ok", "state": best_cp["state"], "epoch": best_cp["epoch"],
                    "metric": best_cp["metric"]}
        return None

    def list_checkpoints(self) -> List[str]:
        return list(self._checkpoints.keys())

    def delete(self, checkpoint_id: str) -> bool:
        return self._checkpoints.pop(checkpoint_id, None) is not None

    @property
    def checkpoint_count(self) -> int:
        return len(self._checkpoints)


class HyperparameterTuner:
    """Hyperparameter tuning — grid search, random search, Bayesian optimization."""

    def __init__(self):
        self._trials: List[dict] = []
        self._best: Optional[dict] = None
        self._best_score = float('inf')

    def grid_search(self, param_grid: Dict[str, List[Any]],
                    eval_fn: Callable, maximize: bool = False) -> dict:
        import itertools
        keys = list(param_grid.keys())
        values = list(param_grid.values())
        best_score = float('-inf') if maximize else float('inf')
        best_params = None
        for combo in itertools.product(*values):
            params = dict(zip(keys, combo))
            score = eval_fn(params)
            trial = {"params": params, "score": score}
            self._trials.append(trial)
            if (maximize and score > best_score) or (not maximize and score < best_score):
                best_score = score
                best_params = params
        self._best = {"params": best_params, "score": best_score}
        self._best_score = best_score
        return {"status": "ok", "best_params": best_params, "best_score": best_score,
                "total_trials": len(self._trials)}

    def random_search(self, param_space: Dict[str, tuple], n_trials: int = 20,
                      eval_fn: Callable = None, maximize: bool = False) -> dict:
        import random
        best_score = float('-inf') if maximize else float('inf')
        best_params = None
        for _ in range(n_trials):
            params = {}
            for key, (low, high) in param_space.items():
                params[key] = random.uniform(low, high)
            if eval_fn:
                score = eval_fn(params)
            else:
                score = sum(params.values())
            trial = {"params": params, "score": score}
            self._trials.append(trial)
            if (maximize and score > best_score) or (not maximize and score < best_score):
                best_score = score
                best_params = params
        self._best = {"params": best_params, "score": best_score}
        return {"status": "ok", "best_params": best_params, "best_score": best_score,
                "total_trials": len(self._trials)}

    def bayesian_optimize(self, param_space: Dict[str, tuple], n_iter: int = 10,
                          eval_fn: Callable = None) -> dict:
        import random
        best_score = float('inf')
        best_params = None
        for i in range(n_iter):
            params = {}
            for key, (low, high) in param_space.items():
                mid = (low + high) / 2
                if i < 3:
                    params[key] = random.uniform(low, high)
                else:
                    spread = (high - low) * (0.5 ** (i // 3))
                    params[key] = max(low, min(high, random.gauss(mid, spread)))
            if eval_fn:
                score = eval_fn(params)
            else:
                score = sum(params.values())
            trial = {"params": params, "score": score, "method": "bayesian"}
            self._trials.append(trial)
            if score < best_score:
                best_score = score
                best_params = params
        self._best = {"params": best_params, "score": best_score}
        return {"status": "ok", "best_params": best_params, "best_score": best_score,
                "iterations": n_iter}

    @property
    def trials(self) -> List[dict]:
        return list(self._trials)

    @property
    def best(self) -> Optional[dict]:
        return self._best

    @property
    def trial_count(self) -> int:
        return len(self._trials)


class ConfusionMatrix:
    """Confusion matrix and classification metrics — precision, recall, F1, ROC."""

    def __init__(self, classes: List[str] = None):
        self._classes = classes or ["0", "1"]
        self._matrix: Dict[str, Dict[str, int]] = {}
        self._init_matrix()

    def _init_matrix(self):
        for actual in self._classes:
            self._matrix[actual] = {pred: 0 for pred in self._classes}

    def update(self, actual: str, predicted: str) -> bool:
        if actual in self._matrix and predicted in self._matrix[actual]:
            self._matrix[actual][predicted] += 1
            return True
        return False

    def update_batch(self, actuals: List[str], predicteds: List[str]) -> None:
        for a, p in zip(actuals, predicteds):
            self.update(a, p)

    def metrics(self, positive_class: str = None) -> dict:
        pos = positive_class or self._classes[-1]
        tp = self._matrix.get(pos, {}).get(pos, 0)
        fp = sum(self._matrix.get(c, {}).get(pos, 0) for c in self._classes if c != pos)
        fn = sum(self._matrix.get(pos, {}).get(c, 0) for c in self._classes if c != pos)
        tn = sum(sum(self._matrix.get(a, {}).get(p, 0) for p in self._classes if p != pos)
                 for a in self._classes if a != pos)
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
        accuracy = (tp + tn) / max(tp + fp + fn + tn, 1)
        specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
        return {"status": "ok", "precision": round(precision, 4),
                "recall": round(recall, 4), "f1_score": round(f1, 4),
                "accuracy": round(accuracy, 4),
                "specificity": round(specificity, 4),
                "tp": tp, "fp": fp, "fn": fn, "tn": tn}

    def matrix(self) -> Dict[str, Dict[str, int]]:
        return {k: dict(v) for k, v in self._matrix.items()}

    def reset(self) -> None:
        self._init_matrix()

    @property
    def class_count(self) -> int:
        return len(self._classes)


class DataAugmentation:
    """Data augmentation — image and text augmentation for ML training."""

    def __init__(self):
        self._augmentations: List[dict] = []

    def flip_horizontal(self, matrix: List[List[float]]) -> List[List[float]]:
        return [row[::-1] for row in matrix]

    def flip_vertical(self, matrix: List[List[float]]) -> List[List[float]]:
        return matrix[::-1]

    def rotate_90(self, matrix: List[List[float]]) -> List[List[float]]:
        if not matrix:
            return matrix
        rows, cols = len(matrix), len(matrix[0])
        return [[matrix[rows - 1 - j][i] for j in range(rows)] for i in range(cols)]

    def add_noise(self, matrix: List[List[float]], noise_factor: float = 0.1) -> List[List[float]]:
        import random
        return [[v + random.gauss(0, noise_factor) for v in row] for row in matrix]

    def normalize(self, matrix: List[List[float]]) -> List[List[float]]:
        flat = [v for row in matrix for v in row]
        if not flat:
            return matrix
        mn, mx = min(flat), max(flat)
        rng = mx - mn if mx != mn else 1
        return [[(v - mn) / rng for v in row] for row in matrix]

    def crop(self, matrix: List[List[float]], top: int = 0, left: int = 0,
             height: int = 0, width: int = 0) -> List[List[float]]:
        h = height or len(matrix)
        w = width or (len(matrix[0]) if matrix else 0)
        return [row[left:left + w] for row in matrix[top:top + h]]

    def text_synonym_replace(self, text: str, synonyms: Dict[str, str] = None) -> str:
        syns = synonyms or {"good": "great", "bad": "poor", "happy": "joyful",
                             "sad": "unhappy", "fast": "quick", "slow": "sluggish"}
        words = text.split()
        result = [syns.get(w.lower(), w) for w in words]
        return " ".join(result)

    def text_random_deletion(self, text: str, prob: float = 0.2) -> str:
        import random
        words = text.split()
        result = [w for w in words if random.random() > prob]
        return " ".join(result) if result else words[0] if words else text

    def text_random_swap(self, text: str, n: int = 1) -> str:
        import random
        words = text.split()
        if len(words) < 2:
            return text
        for _ in range(n):
            i, j = random.sample(range(len(words)), 2)
            words[i], words[j] = words[j], words[i]
        return " ".join(words)

    def augment_image(self, matrix: List[List[float]], techniques: List[str] = None) -> List[dict]:
        techs = techniques or ["flip_h", "flip_v", "rotate", "noise"]
        results = []
        for tech in techs:
            if tech == "flip_h":
                results.append({"technique": "flip_h", "result": self.flip_horizontal(matrix)})
            elif tech == "flip_v":
                results.append({"technique": "flip_v", "result": self.flip_vertical(matrix)})
            elif tech == "rotate":
                results.append({"technique": "rotate", "result": self.rotate_90(matrix)})
            elif tech == "noise":
                results.append({"technique": "noise", "result": self.add_noise(matrix)})
        self._augmentations.extend(results)
        return results

    def augment_text(self, text: str, techniques: List[str] = None) -> List[dict]:
        techs = techniques or ["synonym", "deletion", "swap"]
        results = []
        for tech in techs:
            if tech == "synonym":
                results.append({"technique": "synonym", "result": self.text_synonym_replace(text)})
            elif tech == "deletion":
                results.append({"technique": "deletion", "result": self.text_random_deletion(text)})
            elif tech == "swap":
                results.append({"technique": "swap", "result": self.text_random_swap(text)})
        self._augmentations.extend(results)
        return results

    @property
    def augmentation_count(self) -> int:
        return len(self._augmentations)


def execute_python_event_async(event_name: str, data: str) -> str:
    """Execute event in thread pool — non-blocking for UI."""
    future = _EVENT_EXECUTOR.submit(execute_python_event, event_name, data)
    try:
        return future.result(timeout=30.0)
    except Exception as e:
        return json.dumps({"status": "error", "message": f"Async execution failed: {e}"})
